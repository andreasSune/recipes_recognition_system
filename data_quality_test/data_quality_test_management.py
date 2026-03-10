"""
Pipeline de test et de validatiion des donnees de recettes de cuisine

"""

import pandas as pd
import numpy as np
import ast
import os
import json
from typing import List, Dict, Tuple
from datetime import datetime
import matplotlib.pyplot as plt
import seaborn as sns
from pathlib import Path


NOTEBOOK_DIR = Path().resolve()
PROJECT_ROOT = NOTEBOOK_DIR.parent   
DATA_DIR = PROJECT_ROOT / "data"



"""
Pipeline Stratégie 2 : Validation Structurelle 
Tests de cohérence des graphes de gestes culinaires

"""

# ==============================================================================
# SECTION 1 : UTILITAIRES
# ==============================================================================

def parse_actions_column(actions_value):
    """
    Parse robustement la colonne 'actions' qui peut être sous différents formats
    
    Args:
        actions_value: Valeur de la colonne actions (string, list, ou autre)
    
    Returns:
        Liste d'actions
    """
    if pd.isna(actions_value):
        return []
    
    # Si c'est déjà une liste
    if isinstance(actions_value, list):
        return actions_value
    
    # Si c'est une string représentant une liste
    if isinstance(actions_value, str):
        try:
            # Essayer de parser comme du JSON/Python
            parsed = ast.literal_eval(actions_value)
            if isinstance(parsed, list):
                return parsed
        except (ValueError, SyntaxError):
            # Si échec, essayer de split par virgules
            if ',' in actions_value:
                return [a.strip() for a in actions_value.split(',') if a.strip()]
            # Sinon retourner comme élément unique
            return [actions_value.strip()] if actions_value.strip() else []
    
    return []


def create_output_directory(output_dir: str):
    """Crée le répertoire de sortie et ses sous-répertoires"""
    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(os.path.join(output_dir, 'statistics'), exist_ok=True)
    os.makedirs(os.path.join(output_dir, 'visualizations'), exist_ok=True)
    print(f"✅ Répertoires créés: {output_dir}/")


def classify_cuisine_type(title: str) -> str:
    """
    Classifie le type de cuisine d'une recette uniquement à partir du titre
    
    Args:
        title: Titre de la recette
    
    Returns:
        'bakery', 'stew', 'quick_prep', ou 'other'
    """
    if pd.isna(title):
        return 'other'
    
    title_lower = str(title).lower()
    
    # Mots-clés pour la pâtisserie/boulangerie
    bakery_keywords = [
        # Gâteaux et desserts
        'cake', 'cupcake', 'cheesecake', 'torte', 'gateau',
        # Cookies et biscuits
        'cookie', 'biscuit', 'brownie', 'bar', 'square',
        # Pains et viennoiseries
        'bread', 'bun', 'roll', 'loaf', 'baguette', 'croissant', 'danish', 'scone', 'muffin',
        # Tartes et pâtisseries
        'pie', 'tart', 'pastry', 'puff', 'strudel', 'turnover', 'cobbler', 'crisp', 'crumble',
        # Autres desserts au four
        'pudding', 'souffle', 'meringue', 'macaron', 'eclair', 'profiterole',
        # Ingrédients typiques
        'chocolate', 'vanilla', 'cinnamon', 'frosting', 'icing'
    ]
    
    # Mots-clés pour plats mijotés/cuisinés lentement
    stew_keywords = [
        # Types de plats
        'stew', 'soup', 'broth', 'chowder', 'bisque', 'consommé', 'gumbo', 'chili',
        'casserole', 'pot roast', 'ragout', 'tagine',
        # Techniques de cuisson lente
        'braised', 'slow cooked', 'slow-cooked', 'crock pot', 'crockpot',
        # Plats typiques
        'curry', 'goulash', 'bourguignon', 'stroganoff', 'fricassee',
        # Indicateurs
        'slow cooker', 'dutch oven'
    ]
    
    # Mots-clés pour préparations rapides
    quick_prep_keywords = [
        # Salades
        'salad', 'slaw', 'coleslaw', 'caesar', 'greek salad', 'caprese',
        # Sandwichs et wraps
        'sandwich', 'panini', 'sub', 'hoagie', 'wrap', 'burrito', 'quesadilla', 'taco',
        # Bols et assemblages
        'bowl', 'poke', 'buddha bowl', 'grain bowl', 'rice bowl', 'noodle bowl',
        # Smoothies et boissons
        'smoothie', 'shake', 'juice', 'drink', 'beverage', 'cocktail', 'mocktail',
        # Snacks et apéritifs
        'dip', 'spread', 'hummus', 'guacamole', 'salsa', 'bruschetta', 'crostini',
        'appetizer', 'finger food',
        # Préparations crues/froides
        'no-bake', 'no bake', 'raw', 'fresh', 'cold', 'chilled',
        'ceviche', 'tartare', 'carpaccio',
        # Indicateurs de rapidité
        'quick', 'easy', 'simple', '5 minute', '10 minute', '15 minute', 'instant',
        '5-minute', '10-minute', '15-minute'
    ]
    
    # Compter les occurrences de chaque catégorie
    bakery_score = sum(1 for keyword in bakery_keywords if keyword in title_lower)
    stew_score = sum(1 for keyword in stew_keywords if keyword in title_lower)
    quick_score = sum(1 for keyword in quick_prep_keywords if keyword in title_lower)
    
    # Retourner la catégorie avec le score le plus élevé
    scores = {
        'bakery': bakery_score,
        'stew': stew_score,
        'quick_prep': quick_score
    }
    
    max_score = max(scores.values())
    
    # Si aucun mot-clé trouvé, retourner 'other'
    if max_score == 0:
        return 'other'
    
    # Retourner la catégorie avec le score maximum
    for category, score in scores.items():
        if score == max_score:
            return category
    
    return 'other'


def classify_complexity(steps) -> str:
    """
    Classifie la complexité d'une recette selon le nombre d'étapes
    
    Args:
        steps: Nombre d'étapes (number_of_steps)
    
    Returns:
        'simple', 'moyenne', 'elevee', ou 'unknown'
    """
    if pd.isna(steps):
        return 'unknown'
    
    try:
        steps = int(steps)
        if steps <= 10:
            return 'simple'
        elif steps <= 20:
            return 'moyenne'
        else:
            return 'elevee'
    except (ValueError, TypeError):
        return 'unknown'

# ==============================================================================
# SECTION 2 : TEST 1 - CALCUL DE LA TAILLE DES LISTES D'ACTIONS
# ==============================================================================


def test_1_calculate_action_lengths(
    data_df: pd.DataFrame,
    recipes_df: pd.DataFrame
) -> Tuple[Dict, pd.DataFrame]:
    """
    Test 1: Calcule les statistiques de taille des listes d'actions
    AMÉLIORÉ avec statistiques complètes par variante, category et complexity
    
    Pour chaque niveau (global, par variante, par category, par complexity):
    - Longueur moyenne, médiane, écart-type, min, max, Q1, Q3
    - Outliers détectés (count et pourcentage)
    
    Args:
        data_df: DataFrame nettoyé avec toutes les actions
        recipes_df: DataFrame des recettes avec category et complexity
    
    Returns:
        Tuple (stats_dict, dataset_resume_df)
    """
    print(f"\n{'='*80}")
    print(f"TEST 1 : CALCUL DE LA TAILLE DES LISTES D'ACTIONS")
    print(f"{'='*80}\n")
    
    # Parser les actions
    data_df = data_df.copy()
    data_df['actions_parsed'] = data_df['actions'].apply(parse_actions_column)
    data_df['action_length'] = data_df['actions_parsed'].apply(len)
    
    # Merger avec recipes pour obtenir category et complexity
    data_df = data_df.merge(
        recipes_df[['id', 'category', 'complexity']], 
        on='id', 
        how='left'
    )
    
    def calculate_detailed_stats(subset_df, subset_name=""):
        """
        Calcule les statistiques complètes pour un sous-ensemble de données
        
        Args:
            subset_df: DataFrame filtré
            subset_name: Nom du sous-ensemble pour l'affichage
        
        Returns:
            Dictionnaire de statistiques
        """
        if len(subset_df) == 0:
            return None
        
        # Statistiques de base
        stats = {
            'count': int(len(subset_df)),
            'longueur_min': int(subset_df['action_length'].min()),
            'longueur_max': int(subset_df['action_length'].max()),
            'longueur_moyenne': float(subset_df['action_length'].mean()),
            'longueur_mediane': float(subset_df['action_length'].median()),
            'ecart_type': float(subset_df['action_length'].std()),
            'q1': float(subset_df['action_length'].quantile(0.25)),
            'q3': float(subset_df['action_length'].quantile(0.75))
        }
        
        # Calcul des outliers (> Q3 + 3*IQR ou < Q1 - 3*IQR)
        iqr = stats['q3'] - stats['q1']
        upper_bound = stats['q3'] + 3 * iqr
        lower_bound = stats['q1'] - 3 * iqr
        
        outliers_mask = (subset_df['action_length'] > upper_bound) | (subset_df['action_length'] < lower_bound)
        outliers_count = outliers_mask.sum()
        
        stats['outliers_count'] = int(outliers_count)
        stats['outliers_percentage'] = float(outliers_count / len(subset_df) * 100) if len(subset_df) > 0 else 0.0
        stats['upper_bound'] = float(upper_bound)
        stats['lower_bound'] = float(lower_bound)
        
        return stats
    
    # ========== STATISTIQUES GLOBALES ==========
    print("📊 STATISTIQUES GLOBALES")
    print("-" * 80)
    
    stats_globales = calculate_detailed_stats(data_df, "GLOBAL")
    
    print(f"Total graphes: {stats_globales['count']:,}")
    print(f"Longueur moyenne: {stats_globales['longueur_moyenne']:.2f} ± {stats_globales['ecart_type']:.2f} actions")
    print(f"Médiane: {stats_globales['longueur_mediane']:.1f}")
    print(f"Min: {stats_globales['longueur_min']}, Max: {stats_globales['longueur_max']}")
    print(f"Quartiles: Q1={stats_globales['q1']:.1f}, Q3={stats_globales['q3']:.1f}")
    print(f"Outliers détectés: {stats_globales['outliers_count']:,} ({stats_globales['outliers_percentage']:.2f}%)")
    print(f"  Bornes: [{stats_globales['lower_bound']:.1f}, {stats_globales['upper_bound']:.1f}]")
    
    # ========== STATISTIQUES PAR VARIANTE ==========
    print(f"\n📊 STATISTIQUES PAR VARIANTE")
    print("-" * 80)
    
    stats_by_variant = {}
    for variant_type in sorted(data_df['type_2'].dropna().unique()):
        subset = data_df[data_df['type_2'] == variant_type]
        stats = calculate_detailed_stats(subset, variant_type)
        
        if stats:
            stats_by_variant[variant_type] = stats
            print(f"\n  📌 {variant_type.upper()}")
            print(f"     Count: {stats['count']:,}")
            print(f"     Moyenne: {stats['longueur_moyenne']:.2f} ± {stats['ecart_type']:.2f}")
            print(f"     Médiane: {stats['longueur_mediane']:.1f}")
            print(f"     Min-Max: [{stats['longueur_min']}, {stats['longueur_max']}]")
            print(f"     Quartiles: Q1={stats['q1']:.1f}, Q3={stats['q3']:.1f}")
            print(f"     Outliers: {stats['outliers_count']:,} ({stats['outliers_percentage']:.2f}%)")
            print(f"     Bornes: [{stats['lower_bound']:.1f}, {stats['upper_bound']:.1f}]")
    
    # ========== STATISTIQUES PAR CATEGORY ==========
    print(f"\n📊 STATISTIQUES PAR CATÉGORIE DE CUISINE")
    print("-" * 80)
    
    stats_by_category = {}
    for category in sorted(data_df['category'].dropna().unique()):
        subset = data_df[data_df['category'] == category]
        stats = calculate_detailed_stats(subset, category)
        
        if stats:
            stats_by_category[category] = stats
            print(f"\n  📌 {category.upper()}")
            print(f"     Count: {stats['count']:,}")
            print(f"     Moyenne: {stats['longueur_moyenne']:.2f} ± {stats['ecart_type']:.2f}")
            print(f"     Médiane: {stats['longueur_mediane']:.1f}")
            print(f"     Min-Max: [{stats['longueur_min']}, {stats['longueur_max']}]")
            print(f"     Quartiles: Q1={stats['q1']:.1f}, Q3={stats['q3']:.1f}")
            print(f"     Outliers: {stats['outliers_count']:,} ({stats['outliers_percentage']:.2f}%)")
            print(f"     Bornes: [{stats['lower_bound']:.1f}, {stats['upper_bound']:.1f}]")
    
    # ========== STATISTIQUES PAR COMPLEXITY ==========
    print(f"\n📊 STATISTIQUES PAR NIVEAU DE COMPLEXITÉ")
    print("-" * 80)
    
    stats_by_complexity = {}
    complexity_order = ['simple', 'moyenne', 'elevee', 'unknown']
    for complexity in complexity_order:
        subset = data_df[data_df['complexity'] == complexity]
        if len(subset) > 0:
            stats = calculate_detailed_stats(subset, complexity)
            
            if stats:
                stats_by_complexity[complexity] = stats
                print(f"\n  📌 {complexity.upper()}")
                print(f"     Count: {stats['count']:,}")
                print(f"     Moyenne: {stats['longueur_moyenne']:.2f} ± {stats['ecart_type']:.2f}")
                print(f"     Médiane: {stats['longueur_mediane']:.1f}")
                print(f"     Min-Max: [{stats['longueur_min']}, {stats['longueur_max']}]")
                print(f"     Quartiles: Q1={stats['q1']:.1f}, Q3={stats['q3']:.1f}")
                print(f"     Outliers: {stats['outliers_count']:,} ({stats['outliers_percentage']:.2f}%)")
                print(f"     Bornes: [{stats['lower_bound']:.1f}, {stats['upper_bound']:.1f}]")
    
    print(f"\n✅ Test 1 terminé avec succès")
    
    # Compiler toutes les stats
    all_stats = {
        'globales': stats_globales,
        'par_variante': stats_by_variant,
        'par_category': stats_by_category,
        'par_complexity': stats_by_complexity
    }
    
    # Créer le dataset_resume pour ce dataset
    dataset_resume_partial = data_df[['id', 'action_length', 'type_2']].copy()
    dataset_resume_partial.columns = ['id', 'nombre_actions', 'variante']
    
    return all_stats, dataset_resume_partial


# ==============================================================================
# SECTION 3 : TEST 2 - VARIANTE PRINCIPALE VS NOMBRE D'INSTRUCTIONS 
# ==============================================================================

def test_2_validate_principale_vs_steps(
    recipes_df: pd.DataFrame,
    data_df: pd.DataFrame
) -> Tuple[Dict, pd.DataFrame]:
    """
    Test 2: Vérifie la cohérence entre le nombre d'actions de la variante principale
    et le nombre d'instructions de la recette
    
    MODIFICATIONS:
    - Implémentation des flags critiques pour ratio < 0.3 ou > 5.0
    
    Args:
        recipes_df: DataFrame des recettes
        data_df: DataFrame nettoyé avec toutes les actions
    
    Returns:
        Tuple (stats_dict, flags_df)
    """
    print(f"\n{'='*80}")
    print(f"TEST 2 : COMPARAISON VARIANTE PRINCIPALE VS NOMBRE D'INSTRUCTIONS")
    print(f"{'='*80}\n")
    
    # Filtrer uniquement les variantes principales
    principales = data_df[data_df['type_2'] == 'variante_principale'].copy()
    
    # Parser les actions
    principales['actions_parsed'] = principales['actions'].apply(parse_actions_column)
    principales['action_length'] = principales['actions_parsed'].apply(len)
    
    # Merger avec recipes
    merged = principales.merge(
        recipes_df[['id', 'number_of_steps']], 
        on='id', 
        how='left'
    )
    
    # Calculer le ratio
    merged['ratio'] = merged['action_length'] / merged['number_of_steps'].replace(0, np.nan)
    
    # Définir les seuils de validation
    RATIO_MIN = 0.5
    RATIO_MAX = 2.5
    RATIO_CRITIQUE_MIN = 0.3
    RATIO_CRITIQUE_MAX = 5.0
    
    # Identifier les flags (ordre important!)
    merged['flag'] = 'CONFORME'
    
    # Flags critiques d'abord
    merged.loc[(merged['ratio'] < RATIO_CRITIQUE_MIN) | (merged['ratio'] > RATIO_CRITIQUE_MAX), 'flag'] = 'FLAG_CRITIQUE'
    merged.loc[merged['ratio'].isna(), 'flag'] = 'FLAG_CRITIQUE'  # number_of_steps = 0 ou NaN
    
    # Ensuite flags d'avertissement (seulement si pas déjà critique)
    merged.loc[(merged['ratio'] < RATIO_MIN) & (merged['ratio'] >= RATIO_CRITIQUE_MIN) & (merged['flag'] != 'FLAG_CRITIQUE'), 'flag'] = 'FLAG_RATIO_FAIBLE'
    merged.loc[(merged['ratio'] > RATIO_MAX) & (merged['ratio'] <= RATIO_CRITIQUE_MAX) & (merged['flag'] != 'FLAG_CRITIQUE'), 'flag'] = 'FLAG_RATIO_ELEVE'
    
    # Statistiques
    stats = {
        'total_principales': int(len(merged)),
        'ratio_moyen': float(merged['ratio'].mean()),
        'ratio_median': float(merged['ratio'].median()),
        'ratio_std': float(merged['ratio'].std()),
        'CONFORME_count': int((merged['flag'] == 'CONFORME').sum()),
        'CONFORME_pct': float((merged['flag'] == 'CONFORME').sum() / len(merged) * 100) if len(merged) > 0 else 0.0,
        'FLAG_RATIO_FAIBLE_count': int((merged['flag'] == 'FLAG_RATIO_FAIBLE').sum()),
        'FLAG_RATIO_FAIBLE_pct': float((merged['flag'] == 'FLAG_RATIO_FAIBLE').sum() / len(merged) * 100) if len(merged) > 0 else 0.0,
        'FLAG_RATIO_ELEVE_count': int((merged['flag'] == 'FLAG_RATIO_ELEVE').sum()),
        'FLAG_RATIO_ELEVE_pct': float((merged['flag'] == 'FLAG_RATIO_ELEVE').sum() / len(merged) * 100) if len(merged) > 0 else 0.0,
        'FLAG_CRITIQUE_count': int((merged['flag'] == 'FLAG_CRITIQUE').sum()),
        'FLAG_CRITIQUE_pct': float((merged['flag'] == 'FLAG_CRITIQUE').sum() / len(merged) * 100) if len(merged) > 0 else 0.0
    }
    
    print(f"Total variantes principales: {stats['total_principales']:,}")
    print(f"Ratio moyen (actions/instructions): {stats['ratio_moyen']:.2f}")
    print(f"Ratio médian: {stats['ratio_median']:.2f}")
    print(f"\nDistribution des flags:")
    print(f"  ✅ CONFORME ({RATIO_MIN} ≤ ratio ≤ {RATIO_MAX}): {stats['CONFORME_count']:,} ({stats['CONFORME_pct']:.2f}%)")
    print(f"  ⚠️  FLAG_RATIO_FAIBLE ({RATIO_CRITIQUE_MIN} ≤ ratio < {RATIO_MIN}): {stats['FLAG_RATIO_FAIBLE_count']:,} ({stats['FLAG_RATIO_FAIBLE_pct']:.2f}%)")
    print(f"  ⚠️  FLAG_RATIO_ELEVE ({RATIO_MAX} < ratio ≤ {RATIO_CRITIQUE_MAX}): {stats['FLAG_RATIO_ELEVE_count']:,} ({stats['FLAG_RATIO_ELEVE_pct']:.2f}%)")
    print(f"  ❌ FLAG_CRITIQUE (ratio < {RATIO_CRITIQUE_MIN} ou > {RATIO_CRITIQUE_MAX}): {stats['FLAG_CRITIQUE_count']:,} ({stats['FLAG_CRITIQUE_pct']:.2f}%)")

    print(f"\n✅ Test 2 terminé")
    
    # Préparer le dataframe des flags pour export
    flags_df = merged[['id', 'flag', 'ratio']].copy()
    flags_df['test'] = 2
    flags_df['type_2'] = 'variante_principale'
    
    return stats, flags_df


# ==============================================================================
# SECTION 4 : TEST 3 - VARIANTE INGREDIENTS 
# ==============================================================================

def test_3_validate_ingredients_variant(
    recipes_df: pd.DataFrame,
    data_df: pd.DataFrame
) -> Tuple[Dict, pd.DataFrame]:
    """
    Test 3: Vérifie que la variante ingredients a bien ajouté des gestes
    par rapport à la variante principale
    
    RÉIMPLÉMENTATION COMPLÈTE avec les nouvelles règles:
    - delta = taille_variante_ingredients - taille_variante_principale
    - FLAG_CRITIQUE_NEGATIF: delta < 0
    - FLAG_AUCUN_AJOUT: delta == 0
    - FLAG_TROP_AJOUTS: delta > number_of_ingredients × 2
    
    Args:
        recipes_df: DataFrame des recettes
        data_df: DataFrame nettoyé avec toutes les actions
    
    Returns:
        Tuple (stats_dict, flags_df)
    """
    print(f"\n{'='*80}")
    print(f"TEST 3 : VALIDATION VARIANTE INGREDIENTS")
    print(f"{'='*80}\n")
    
    # Extraire principales et ingredients
    principales = data_df[data_df['type_2'] == 'variante_principale'].copy()
    ingredients = data_df[data_df['type_2'] == 'variante_ingredients'].copy()
    
    # Parser les actions
    principales['actions_parsed'] = principales['actions'].apply(parse_actions_column)
    principales['taille_principale'] = principales['actions_parsed'].apply(len)
    
    ingredients['actions_parsed'] = ingredients['actions'].apply(parse_actions_column)
    ingredients['taille_ingredients'] = ingredients['actions_parsed'].apply(len)
    
    # Merger avec recipes pour avoir number_of_ingredients
    merged = principales[['id', 'taille_principale']].merge(
        ingredients[['id', 'taille_ingredients']],
        on='id',
        how='inner'
    ).merge(
        recipes_df[['id', 'number_of_ingredients', 'number_of_steps']],
        on='id',
        how='left'
    )
    
    # Calculer le delta
    merged['delta'] = merged['taille_ingredients'] - merged['taille_principale']
    
    # Calculer les bornes théoriques
    merged['borne_sup_delta'] = merged['number_of_ingredients'] * 2
    merged['borne_sup_stricte'] = (merged['number_of_steps'] * 3) + (merged['number_of_ingredients'] * 2)
    
    # Identifier les flags
    merged['flag'] = 'CONFORME'
    
    # Flag critique: delta négatif (impossible!)
    merged.loc[merged['delta'] < 0, 'flag'] = 'FLAG_CRITIQUE_NEGATIF'
    
    # Flags d'avertissement
    merged.loc[(merged['delta'] == 0) & (merged['flag'] == 'CONFORME'), 'flag'] = 'FLAG_AUCUN_AJOUT'
    merged.loc[(merged['delta'] > merged['borne_sup_delta']) & (merged['flag'] == 'CONFORME'), 'flag'] = 'FLAG_TROP_AJOUTS'
    
    # Optionnel: vérifier borne stricte
    merged.loc[(merged['taille_ingredients'] > merged['borne_sup_stricte']) & (merged['flag'] == 'CONFORME'), 'flag'] = 'FLAG_DEPASSEMENT_BORNE_STRICTE'
    
    # Statistiques
    stats = {
        'total_paires': int(len(merged)),
        'delta_moyen': float(merged['delta'].mean()),
        'delta_median': float(merged['delta'].median()),
        'delta_std': float(merged['delta'].std()),
        'delta_min': int(merged['delta'].min()),
        'delta_max': int(merged['delta'].max()),
        'CONFORME_count': int((merged['flag'] == 'CONFORME').sum()),
        'CONFORME_pct': float((merged['flag'] == 'CONFORME').sum() / len(merged) * 100) if len(merged) > 0 else 0.0,
        'FLAG_CRITIQUE_NEGATIF_count': int((merged['flag'] == 'FLAG_CRITIQUE_NEGATIF').sum()),
        'FLAG_CRITIQUE_NEGATIF_pct': float((merged['flag'] == 'FLAG_CRITIQUE_NEGATIF').sum() / len(merged) * 100) if len(merged) > 0 else 0.0,
        'FLAG_AUCUN_AJOUT_count': int((merged['flag'] == 'FLAG_AUCUN_AJOUT').sum()),
        'FLAG_AUCUN_AJOUT_pct': float((merged['flag'] == 'FLAG_AUCUN_AJOUT').sum() / len(merged) * 100) if len(merged) > 0 else 0.0,
        'FLAG_TROP_AJOUTS_count': int((merged['flag'] == 'FLAG_TROP_AJOUTS').sum()),
        'FLAG_TROP_AJOUTS_pct': float((merged['flag'] == 'FLAG_TROP_AJOUTS').sum() / len(merged) * 100) if len(merged) > 0 else 0.0,
        'FLAG_DEPASSEMENT_BORNE_STRICTE_count': int((merged['flag'] == 'FLAG_DEPASSEMENT_BORNE_STRICTE').sum()),
        'FLAG_DEPASSEMENT_BORNE_STRICTE_pct': float((merged['flag'] == 'FLAG_DEPASSEMENT_BORNE_STRICTE').sum() / len(merged) * 100) if len(merged) > 0 else 0.0
    }
    
    print(f"Total paires principale-ingredients: {stats['total_paires']:,}")
    print(f"Delta moyen: {stats['delta_moyen']:.2f}")
    print(f"Delta médian: {stats['delta_median']:.1f}")
    print(f"Delta [min, max]: [{stats['delta_min']}, {stats['delta_max']}]")
    print(f"\nDistribution des flags:")
    print(f"  ✅ CONFORME (delta ≥ 0 et ≤ nb_ingredients × 2): {stats['CONFORME_count']:,} ({stats['CONFORME_pct']:.2f}%)")
    print(f"  ❌ FLAG_CRITIQUE_NEGATIF (delta < 0): {stats['FLAG_CRITIQUE_NEGATIF_count']:,} ({stats['FLAG_CRITIQUE_NEGATIF_pct']:.2f}%)")
    print(f"  ⚠️  FLAG_AUCUN_AJOUT (delta == 0): {stats['FLAG_AUCUN_AJOUT_count']:,} ({stats['FLAG_AUCUN_AJOUT_pct']:.2f}%)")
    print(f"  ⚠️  FLAG_TROP_AJOUTS (delta > nb_ingredients × 2): {stats['FLAG_TROP_AJOUTS_count']:,} ({stats['FLAG_TROP_AJOUTS_pct']:.2f}%)")
    print(f"  ⚠️  FLAG_DEPASSEMENT_BORNE_STRICTE: {stats['FLAG_DEPASSEMENT_BORNE_STRICTE_count']:,} ({stats['FLAG_DEPASSEMENT_BORNE_STRICTE_pct']:.2f}%)")
    
    print(f"\n✅ Test 3 terminé")
    
    # Préparer le dataframe des flags pour export
    flags_df = merged[['id', 'flag', 'delta']].copy()
    flags_df['test'] = 3
    flags_df['type_2'] = 'variante_ingredients'
    
    return stats, flags_df


# ==============================================================================
# SECTION 5 : TEST 4A - VARIANTE PERMUTATION 
# ==============================================================================

def test_4a_validate_permutation_variant(
    data_df: pd.DataFrame
) -> Tuple[Dict, pd.DataFrame]:
    """
    Test 4A: Vérifie que la variante permutation contient les mêmes actions
    que la variante principale avec un ordre différent
    
    Utilise:
    - Similarité de Jaccard (contenu): vérifie que les actions sont les mêmes
    - Distance de Levenshtein (ordre): vérifie que l'ordre a changé
    
    Conditions:
    - CONFORME: overlap ∈ [0.6, 1.0] ET levenshtein > 0
    - FLAG_CRITIQUE_SIMILARITÉ_BASSE: overlap < 0.6
    - FLAG_IDENTIQUE: levenshtein = 0 (pas de permutation réelle)
    
    Args:
        data_df: DataFrame nettoyé avec toutes les actions
    
    Returns:
        Tuple (stats_dict, flags_df)
    """
    print(f"\n{'='*80}")
    print(f"TEST 4A : VALIDATION VARIANTE PERMUTATION")
    print(f"{'='*80}\n")
    
    # Extraire principales et permutations
    principales = data_df[data_df['type_2'] == 'variante_principale'].copy()
    permutations = data_df[data_df['type_2'] == 'variante_permutation'].copy()
    
    # Parser les actions
    principales['actions_parsed'] = principales['actions'].apply(parse_actions_column)
    permutations['actions_parsed'] = permutations['actions'].apply(parse_actions_column)
    
    # Merger
    merged = principales[['id', 'actions_parsed']].merge(
        permutations[['id', 'actions_parsed']],
        on='id',
        how='inner',
        suffixes=('_princ', '_perm')
    )
    
    print(f"Total paires principale-permutation à analyser: {len(merged):,}")
    
    # Calculer Jaccard et Levenshtein pour chaque paire
    def calculate_jaccard(set1, set2):
        """Calcule la similarité de Jaccard entre deux ensembles"""
        if len(set1) == 0 and len(set2) == 0:
            return 1.0
        
        intersection = len(set1 & set2)
        union = len(set1 | set2)
        
        return intersection / union if union > 0 else 0.0
    
    def calculate_levenshtein(seq1, seq2):
        """Calcule la distance de Levenshtein entre deux séquences"""
        if len(seq1) == 0:
            return len(seq2)
        if len(seq2) == 0:
            return len(seq1)
        
        # Matrice de programmation dynamique
        dp = [[0] * (len(seq2) + 1) for _ in range(len(seq1) + 1)]
        
        # Initialisation
        for i in range(len(seq1) + 1):
            dp[i][0] = i
        for j in range(len(seq2) + 1):
            dp[0][j] = j
        
        # Calcul
        for i in range(1, len(seq1) + 1):
            for j in range(1, len(seq2) + 1):
                if seq1[i-1] == seq2[j-1]:
                    dp[i][j] = dp[i-1][j-1]
                else:
                    dp[i][j] = 1 + min(
                        dp[i-1][j],      # Suppression
                        dp[i][j-1],      # Insertion
                        dp[i-1][j-1]     # Substitution
                    )
        
        return dp[len(seq1)][len(seq2)]
    
    # Calculer les métriques
    merged['jaccard'] = merged.apply(
        lambda row: calculate_jaccard(
            set(row['actions_parsed_princ']),
            set(row['actions_parsed_perm'])
        ),
        axis=1
    )
    
    merged['levenshtein'] = merged.apply(
        lambda row: calculate_levenshtein(
            row['actions_parsed_princ'],
            row['actions_parsed_perm']
        ),
        axis=1
    )
    
    # Assigner les flags selon les conditions
    def assign_flag(row):
        overlap = row['jaccard']
        levenshtein = row['levenshtein']
        
        # Condition 1: overlap < 0.6 → FLAG_CRITIQUE_SIMILARITÉ_BASSE
        if overlap < 0.6:
            return 'FLAG_CRITIQUE_SIMILARITE_BASSE'
        
        # Condition 2: levenshtein = 0 → FLAG_IDENTIQUE (pas de permutation)
        if levenshtein == 0:
            return 'FLAG_IDENTIQUE'
        
        # Condition 3: overlap ∈ [0.6, 1.0] ET levenshtein > 0 → CONFORME
        if 0.6 <= overlap <= 1.0 and levenshtein > 0:
            return 'CONFORME'
        
        # Cas par défaut (ne devrait pas arriver)
        return 'FLAG_UNKNOWN'
    
    merged['flag'] = merged.apply(assign_flag, axis=1)
    
    # Statistiques
    total = len(merged)
    
    stats = {
        'total_paires': int(total),
        'CONFORME_count': int((merged['flag'] == 'CONFORME').sum()),
        'CONFORME_pct': float((merged['flag'] == 'CONFORME').sum() / total * 100) if total > 0 else 0.0,
        'FLAG_CRITIQUE_SIMILARITE_BASSE_count': int((merged['flag'] == 'FLAG_CRITIQUE_SIMILARITE_BASSE').sum()),
        'FLAG_CRITIQUE_SIMILARITE_BASSE_pct': float((merged['flag'] == 'FLAG_CRITIQUE_SIMILARITE_BASSE').sum() / total * 100) if total > 0 else 0.0,
        'FLAG_IDENTIQUE_count': int((merged['flag'] == 'FLAG_IDENTIQUE').sum()),
        'FLAG_IDENTIQUE_pct': float((merged['flag'] == 'FLAG_IDENTIQUE').sum() / total * 100) if total > 0 else 0.0,
        'jaccard_mean': float(merged['jaccard'].mean()),
        'jaccard_min': float(merged['jaccard'].min()),
        'jaccard_max': float(merged['jaccard'].max()),
        'levenshtein_mean': float(merged['levenshtein'].mean()),
        'levenshtein_min': int(merged['levenshtein'].min()),
        'levenshtein_max': int(merged['levenshtein'].max())
    }
    
    print(f"Total paires principale-permutation: {stats['total_paires']:,}")
    print(f"\nDistribution des flags:")
    print(f"  ✅ CONFORME (overlap ∈ [0.6, 1.0] ET levenshtein > 0): {stats['CONFORME_count']:,} ({stats['CONFORME_pct']:.2f}%)")
    print(f"  ❌ FLAG_CRITIQUE_SIMILARITE_BASSE (overlap < 0.6): {stats['FLAG_CRITIQUE_SIMILARITE_BASSE_count']:,} ({stats['FLAG_CRITIQUE_SIMILARITE_BASSE_pct']:.2f}%)")
    print(f"  ⚠️  FLAG_IDENTIQUE (levenshtein = 0): {stats['FLAG_IDENTIQUE_count']:,} ({stats['FLAG_IDENTIQUE_pct']:.2f}%)")
    
    print(f"\nMétriques de similarité (Jaccard):")
    print(f"  Moyenne: {stats['jaccard_mean']:.3f}")
    print(f"  Min: {stats['jaccard_min']:.3f}, Max: {stats['jaccard_max']:.3f}")
    
    print(f"\nMétriques d'ordre (Levenshtein):")
    print(f"  Moyenne: {stats['levenshtein_mean']:.1f}")
    print(f"  Min: {stats['levenshtein_min']}, Max: {stats['levenshtein_max']}")
    
    print(f"\n✅ Test 4A terminé")
    
    # Préparer le dataframe des flags pour export (seulement les non-conformes)
    flags_df = merged[merged['flag'] != 'CONFORME'][['id', 'flag', 'jaccard', 'levenshtein']].copy()
    flags_df['test'] = '4A'
    flags_df['type_2'] = 'variante_permutation'
    flags_df.rename(columns={'jaccard': 'jaccard_similarity', 'levenshtein': 'levenshtein_distance'}, inplace=True)
    
    return stats, flags_df


# ==============================================================================
# SECTION 6 : TEST 4B - SIMILARITÉ VARIANTE INGREDIENTS 
# ==============================================================================

def test_4b_validate_ingredients_similarity(
    data_df: pd.DataFrame
) -> Tuple[Dict, pd.DataFrame]:
    """
    Test 4B: Calcule uniquement la similarité entre variante ingredients et principale
    
    MODIFICATIONS:
    - Suppression de l'évaluation du nombre de gestes (déjà fait en Test 2)
    - Focus uniquement sur overlap (Jaccard similarity)
    - Nouveaux flags basés sur overlap
    
    Args:
        data_df: DataFrame nettoyé avec toutes les actions
    
    Returns:
        Tuple (stats_dict, flags_df)
    """
    print(f"\n{'='*80}")
    print(f"TEST 4B : VARIANTE INGREDIENTS VS PRINCIPALE (SIMILARITÉ)")
    print(f"{'='*80}\n")
    
    def calculate_overlap(list1, list2):
        """Calcule l'overlap (Jaccard similarity)"""
        if len(list1) == 0 and len(list2) == 0:
            return 1.0
        set1, set2 = set(list1), set(list2)
        intersection = len(set1 & set2)
        union = len(set1 | set2)
        return intersection / union if union > 0 else 0.0
    
    # Extraire principales et ingredients
    principales = data_df[data_df['type_2'] == 'variante_principale'].copy()
    ingredients = data_df[data_df['type_2'] == 'variante_ingredients'].copy()
    
    # Parser les actions
    principales['actions_parsed'] = principales['actions'].apply(parse_actions_column)
    ingredients['actions_parsed'] = ingredients['actions'].apply(parse_actions_column)
    
    # Merger
    merged = ingredients[['id', 'actions_parsed']].merge(
        principales[['id', 'actions_parsed']],
        on='id',
        how='inner',
        suffixes=('_ingr', '_princ')
    )
    
    # Calculer l'overlap
    merged['overlap'] = merged.apply(
        lambda row: calculate_overlap(row['actions_parsed_ingr'], row['actions_parsed_princ']),
        axis=1
    )
    
    # Identifier les flags basés uniquement sur l'overlap
    merged['flag'] = 'CONFORME'
    merged.loc[(merged['overlap'] >= 0.5) & (merged['overlap'] < 0.7), 'flag'] = 'FLAG_SIMILARITE_MOYENNE'
    merged.loc[merged['overlap'] < 0.5, 'flag'] = 'FLAG_CRITIQUE_SIMILARITE_BASSE'
    
    # Statistiques
    stats = {
        'total_paires': int(len(merged)),
        'overlap_moyen': float(merged['overlap'].mean()),
        'overlap_median': float(merged['overlap'].median()),
        'overlap_std': float(merged['overlap'].std()),
        'overlap_min': float(merged['overlap'].min()),
        'overlap_max': float(merged['overlap'].max()),
        'CONFORME_count': int((merged['flag'] == 'CONFORME').sum()),
        'CONFORME_pct': float((merged['flag'] == 'CONFORME').sum() / len(merged) * 100) if len(merged) > 0 else 0.0,
        'FLAG_SIMILARITE_MOYENNE_count': int((merged['flag'] == 'FLAG_SIMILARITE_MOYENNE').sum()),
        'FLAG_SIMILARITE_MOYENNE_pct': float((merged['flag'] == 'FLAG_SIMILARITE_MOYENNE').sum() / len(merged) * 100) if len(merged) > 0 else 0.0,
        'FLAG_CRITIQUE_SIMILARITE_BASSE_count': int((merged['flag'] == 'FLAG_CRITIQUE_SIMILARITE_BASSE').sum()),
        'FLAG_CRITIQUE_SIMILARITE_BASSE_pct': float((merged['flag'] == 'FLAG_CRITIQUE_SIMILARITE_BASSE').sum() / len(merged) * 100) if len(merged) > 0 else 0.0
    }
    
    print(f"Total paires principale-ingredients: {stats['total_paires']:,}")
    print(f"Overlap moyen: {stats['overlap_moyen']:.3f}")
    print(f"Overlap médian: {stats['overlap_median']:.3f}")
    print(f"Overlap [min, max]: [{stats['overlap_min']:.3f}, {stats['overlap_max']:.3f}]")
    print(f"\nDistribution des flags:")
    print(f"  ✅ CONFORME (overlap ≥ 0.7): {stats['CONFORME_count']:,} ({stats['CONFORME_pct']:.2f}%)")
    print(f"  ⚠️  FLAG_SIMILARITE_MOYENNE (0.5 ≤ overlap < 0.7): {stats['FLAG_SIMILARITE_MOYENNE_count']:,} ({stats['FLAG_SIMILARITE_MOYENNE_pct']:.2f}%)")
    print(f"  ❌ FLAG_CRITIQUE_SIMILARITE_BASSE (overlap < 0.5): {stats['FLAG_CRITIQUE_SIMILARITE_BASSE_count']:,} ({stats['FLAG_CRITIQUE_SIMILARITE_BASSE_pct']:.2f}%)")
    
    print(f"\n✅ Test 4B terminé")
    
    # Préparer le dataframe des flags pour export
    flags_df = merged[['id', 'flag', 'overlap']].copy()
    flags_df['test'] = '4B'
    flags_df['type_2'] = 'variante_ingredients'
    
    return stats, flags_df


# ==============================================================================
# SECTION 7 : TEST 6 - COHÉRENCE GLOBALE PAR RECETTE
# ==============================================================================

def test_6_validate_recipe_coherence(
    data_df: pd.DataFrame
) -> Tuple[Dict, pd.DataFrame]:
    """
    Test 6: Vérifie la cohérence globale par recette
    (chaque recette doit avoir exactement 1 variante principale)
    
    Args:
        data_df: DataFrame avec colonnes 'id', 'type_2', 'actions'
    
    Returns:
        Tuple (stats_dict, flags_df)
    """
    print(f"\n{'='*80}")
    print(f"TEST 6 : COHÉRENCE GLOBALE PAR RECETTE")
    print(f"{'='*80}\n")
    
    # Compter les variantes par recette
    variant_counts = data_df.groupby(['id', 'type_2']).size().unstack(fill_value=0)
    
    # Créer un dataframe des recettes
    recipe_coherence = pd.DataFrame({
        'id': variant_counts.index
    })
    
    # Vérifier la présence de la variante principale
    if 'variante_principale' in variant_counts.columns:
        recipe_coherence['nb_principale'] = variant_counts['variante_principale'].values
    else:
        recipe_coherence['nb_principale'] = 0
    
    # Assigner les flags
    recipe_coherence['flag'] = 'CONFORME'
    recipe_coherence.loc[recipe_coherence['nb_principale'] == 0, 'flag'] = 'FLAG_CRITIQUE_NO_PRINCIPALE'
    recipe_coherence.loc[recipe_coherence['nb_principale'] > 1, 'flag'] = 'FLAG_CRITIQUE_MULTIPLE_PRINCIPALE'
    
    # Statistiques
    total_recettes = len(recipe_coherence)
    conforme_count = (recipe_coherence['flag'] == 'CONFORME').sum()
    no_principale_count = (recipe_coherence['flag'] == 'FLAG_CRITIQUE_NO_PRINCIPALE').sum()
    multiple_principale_count = (recipe_coherence['flag'] == 'FLAG_CRITIQUE_MULTIPLE_PRINCIPALE').sum()
    
    # ✅ CORRECTION: Retourner les COUNTS ET les POURCENTAGES
    stats = {
        'total_recettes': int(total_recettes),
        'CONFORME_count': int(conforme_count),
        'CONFORME_pct': float(conforme_count / total_recettes * 100) if total_recettes > 0 else 0.0,
        'FLAG_CRITIQUE_NO_PRINCIPALE_count': int(no_principale_count),
        'FLAG_CRITIQUE_NO_PRINCIPALE_pct': float(no_principale_count / total_recettes * 100) if total_recettes > 0 else 0.0,
        'FLAG_CRITIQUE_MULTIPLE_PRINCIPALE_count': int(multiple_principale_count),
        'FLAG_CRITIQUE_MULTIPLE_PRINCIPALE_pct': float(multiple_principale_count / total_recettes * 100) if total_recettes > 0 else 0.0
    }
    
    print(f"Total recettes analysées: {stats['total_recettes']:,}")
    print(f"\nDistribution des flags:")
    print(f"  ❌ FLAG_CRITIQUE_NO_PRINCIPALE: {stats['FLAG_CRITIQUE_NO_PRINCIPALE_count']:,} ({stats['FLAG_CRITIQUE_NO_PRINCIPALE_pct']:.2f}%)")
    print(f"  ❌ FLAG_CRITIQUE_MULTIPLE_PRINCIPALE: {stats['FLAG_CRITIQUE_MULTIPLE_PRINCIPALE_count']:,} ({stats['FLAG_CRITIQUE_MULTIPLE_PRINCIPALE_pct']:.2f}%)")
    print(f"  ✅ CONFORME: {stats['CONFORME_count']:,} ({stats['CONFORME_pct']:.2f}%)")
    
    print(f"\n✅ Test 6 terminé")
    
    # Préparer le dataframe des flags critiques uniquement
    flags_df = recipe_coherence[recipe_coherence['flag'] != 'CONFORME'][['id', 'flag', 'nb_principale']].copy()
    flags_df['test'] = 6
    flags_df['type_2'] = 'N/A'  # Test 6 ne concerne pas une variante spécifique
    
    return stats, flags_df

# ==============================================================================
# SECTION 8 : EXPORT DU DATASET DES FLAGS CRITIQUES
# ==============================================================================

def export_critical_flags_dataset(
    all_flags_dfs: List[pd.DataFrame],
    output_dir: str
) -> str:
    """
    Crée le dataset des recettes avec flags critiques
    
    Colonnes: id, test, type_2
    
    Args:
        all_flags_dfs: Liste des DataFrames de flags de chaque test
        output_dir: Répertoire de sortie
    
    Returns:
        Chemin du fichier exporté
    """
    print(f"\n{'='*80}")
    print(f"EXPORT DU DATASET DES FLAGS CRITIQUES")
    print(f"{'='*80}\n")

    output_dir = DATA_DIR/output_dir
    
    # Concaténer tous les flags
    all_flags = pd.concat(all_flags_dfs, ignore_index=True)
    
    # Filtrer uniquement les flags critiques
    critical_keywords = ['CRITIQUE', 'FLAG_LISTE_VIDE', 'FLAG_ENSEMBLE_DIFFERENT']
    
    critical_flags = all_flags[
        all_flags['flag'].str.contains('|'.join(critical_keywords), case=False, na=False)
    ].copy()
    
    # Sélectionner les colonnes requises
    dataset_critique = critical_flags[['id', 'test', 'type_2']].copy()
    
    # Sauvegarder
    output_file = os.path.join(output_dir, 'dataset_flags_critique.csv')
    dataset_critique.to_csv(output_file, index=False, encoding='utf-8')
    
    print(f"Total flags critiques: {len(dataset_critique):,}")
    print(f"Recettes uniques avec flags critiques: {dataset_critique['id'].nunique():,}")
    print(f"\nDistribution par test:")

    for test in sorted(dataset_critique['test'].astype(str).unique()):
        count = len(dataset_critique[dataset_critique['test'] == test])
        print(f"  Test {test}: {count:,} flags")
    
    print(f"\n✅ Dataset sauvegardé : {output_file}")
    
    return output_file



import pandas as pd
import matplotlib.pyplot as plt
import os
from typing import Dict
from datetime import datetime


# ==============================================================================
# SECTION 9 :VISUALISATION DES PROPORTIONS DE FLAGS PAR TEST 
# ==============================================================================

def plot_flag_proportions(
    all_stats: Dict,
    output_dir: str
) -> str:
    """
    Crée une visualisation des proportions (%) de TOUS LES FLAGS pour chaque test
    
    Pour chaque test, affiche un bar plot avec TOUS ses flags (y compris CONFORME).
    
    Test 6 = Cohérence globale par recette:
    - CONFORME
    - FLAG_NO_PRINCIPALE
    - FLAG_MULTIPLE_PRINCIPALE
    
    Args:
        all_stats: Dictionnaire contenant les statistiques de tous les tests
        output_dir: Répertoire de sortie
    
    Returns:
        Chemin du fichier de visualisation
    """
    print(f"\n{'='*80}")
    print(f"GÉNÉRATION DE LA VISUALISATION DES PROPORTIONS")
    print(f"{'='*80}\n")
    
    # Préparer les données pour la visualisation
    data_for_plot = []
    
    for test_name, stats in all_stats.items():
        if test_name == 'test_1':
            continue  # Test 1 n'a pas de flags
        
        # Test 6 a une structure différente (flags dans un sous-dict)
        if test_name == 'test_6':
            if 'flags' in stats:
                # Calculer les pourcentages pour Test 6
                total_recettes = stats.get('total_recettes', 0)
                if total_recettes > 0:
                    # FLAG_NO_PRINCIPALE
                    flag_no_principale = stats['flags'].get('FLAG_NO_PRINCIPALE', 0)
                    pct_no_principale = (flag_no_principale / total_recettes) * 100
                    
                    # FLAG_MULTIPLE_PRINCIPALE
                    flag_multiple = stats['flags'].get('FLAG_MULTIPLE_PRINCIPALE', 0)
                    pct_multiple = (flag_multiple / total_recettes) * 100
                    
                    # CONFORME
                    pct_conforme = stats['flags'].get('pct_conformes', 0)
                    
                    # Ajouter les données
                    data_for_plot.append({
                        'test': 'Test 6',
                        'flag': 'CONFORME',
                        'percentage': pct_conforme,
                        'color': '#2ECC71'
                    })
                    
                    if pct_no_principale > 0:
                        data_for_plot.append({
                            'test': 'Test 6',
                            'flag': 'FLAG_NO_PRINCIPALE',
                            'percentage': pct_no_principale,
                            'color': '#E74C3C'
                        })
                    
                    if pct_multiple > 0:
                        data_for_plot.append({
                            'test': 'Test 6',
                            'flag': 'FLAG_MULTIPLE_PRINCIPALE',
                            'percentage': pct_multiple,
                            'color': '#E74C3C'
                        })
        else:
            # Pour les autres tests, extraire normalement
            for key, value in stats.items():
                if key.endswith('_pct'):
                    flag_name = key.replace('_pct', '')
                    
                    # Déterminer la couleur selon le type de flag
                    if 'CRITIQUE' in flag_name.upper() or 'VIDE' in flag_name.upper() or 'DIFFERENT' in flag_name.upper() or 'NEGATIF' in flag_name.upper():
                        color = '#E74C3C'  # Rouge critique
                    elif 'CONFORME' in flag_name.upper():
                        color = '#2ECC71'  # Vert conforme
                    else:
                        color = '#F39C12'  # Orange avertissement
                    
                    data_for_plot.append({
                        'test': test_name.replace('test_', 'Test '),
                        'flag': flag_name,
                        'percentage': value,
                        'color': color
                    })
    
    df_plot = pd.DataFrame(data_for_plot)
    
    # Créer la visualisation avec un subplot par test
    tests = sorted(df_plot['test'].unique())
    n_tests = len(tests)
    
    # Calculer la disposition optimale des subplots
    n_cols = 3
    n_rows = (n_tests + n_cols - 1) // n_cols
    
    fig, axes = plt.subplots(n_rows, n_cols, figsize=(20, 5 * n_rows))
    if n_tests == 1:
        axes = [axes]
    else:
        axes = axes.flatten()
    
    for idx, test in enumerate(tests):
        ax = axes[idx]
        test_data = df_plot[df_plot['test'] == test].copy()
        
        # Trier: CONFORME en premier, puis le reste par pourcentage décroissant
        conforme_rows = test_data[test_data['flag'].str.contains('CONFORME', case=False)]
        other_rows = test_data[~test_data['flag'].str.contains('CONFORME', case=False)].sort_values('percentage', ascending=False)
        test_data = pd.concat([conforme_rows, other_rows])
        
        # Graphique en barres avec couleurs individuelles
        bars = ax.bar(
            range(len(test_data)), 
            test_data['percentage'], 
            color=test_data['color'].values, 
            edgecolor='black', 
            linewidth=1.5
        )
        
        # Étiquettes des flags
        ax.set_xticks(range(len(test_data)))
        ax.set_xticklabels(test_data['flag'].values, rotation=45, ha='right', fontsize=9)
        ax.set_ylabel('Pourcentage (%)', fontsize=10, weight='bold')
        ax.set_title(test, fontsize=12, weight='bold', pad=10)
        ax.set_ylim(0, 105)
        ax.grid(axis='y', alpha=0.3, linestyle='--')
        
        # Ajouter les valeurs sur les barres
        for i, (bar, pct) in enumerate(zip(bars, test_data['percentage'])):
            if pct > 0:
                ax.text(
                    bar.get_x() + bar.get_width()/2, 
                    pct + 1, 
                    f'{pct:.1f}%', 
                    ha='center', 
                    va='bottom',
                    fontsize=9, 
                    weight='bold'
                )
    
    # Supprimer les axes inutilisés
    for idx in range(n_tests, len(axes)):
        fig.delaxes(axes[idx])
    
    # Légende globale
    from matplotlib.patches import Patch
    legend_elements = [
        Patch(facecolor='#2ECC71', edgecolor='black', label='✅ Conforme'),
        Patch(facecolor='#F39C12', edgecolor='black', label='⚠️ Avertissement'),
        Patch(facecolor='#E74C3C', edgecolor='black', label='❌ Critique')
    ]
    fig.legend(handles=legend_elements, loc='lower center', ncol=3, fontsize=11, frameon=True, fancybox=True, shadow=True)
    
    plt.suptitle('Distribution des Flags par Test - Stratégie 2', fontsize=16, weight='bold', y=0.995)
    plt.tight_layout(rect=[0, 0.02, 1, 0.98])
    
    # Sauvegarder
    output_file = os.path.join(output_dir, 'visualizations', 'flag_proportions_par_test.png')
    plt.savefig(output_file, dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"✅ Visualisation sauvegardée : {output_file}")
    print(f"   Nombre de tests visualisés : {n_tests}")
    print(f"   Nombre total de flags affichés : {len(df_plot)}")
    
    # Afficher le détail par test
    for test in tests:
        test_flags = df_plot[df_plot['test'] == test]
        print(f"\n   {test}:")
        for _, row in test_flags.iterrows():
            print(f"      - {row['flag']}: {row['percentage']:.1f}%")
    
    return output_file


# ==============================================================================
# SECTION 10 : GÉNÉRATION DU RAPPORT MARKDOWN 
# ==============================================================================

def generate_validation_report_strat_2(
    all_stats: Dict,
    critical_flags_file: str,
    visualization_file: str,
    output_dir: str
) -> str:
    """
    Génère le rapport Markdown de validation structurelle avec TOUTES les statistiques
    
    Test 6 = Cohérence globale par recette (vérification variante principale)
    
    Args:
        all_stats: Dictionnaire des statistiques de tous les tests
        critical_flags_file: Chemin du fichier des flags critiques
        visualization_file: Chemin de la visualisation
        output_dir: Répertoire de sortie
    
    Returns:
        Chemin du rapport généré
    """
    print(f"\n{'='*80}")
    print(f"GÉNÉRATION DU RAPPORT MARKDOWN")
    print(f"{'='*80}\n")
    
    report_lines = []
    
    # En-tête
    report_lines.append("# Rapport de Validation Structurelle - Stratégie 2")
    report_lines.append("")
    report_lines.append(f"**Date de génération:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report_lines.append("")
    report_lines.append("---")
    report_lines.append("")
    
    # Résumé exécutif
    report_lines.append("## Résumé Exécutif")
    report_lines.append("")
    report_lines.append("Ce rapport présente les résultats de la validation structurelle des graphes de gestes culinaires.")
    report_lines.append("La Stratégie 2 évalue la cohérence structurelle des variantes de graphes à travers 6 tests complémentaires.")
    report_lines.append("")
    report_lines.append("---")
    report_lines.append("")
    
    # Test 1 - STATISTIQUES DÉTAILLÉES
    if 'test_1' in all_stats:
        report_lines.append("## Test 1 : Statistiques de Taille des Listes d'Actions")
        report_lines.append("")
        
        stats_t1 = all_stats['test_1']
        
        # Statistiques globales
        if 'globales' in stats_t1:
            report_lines.append("### 📊 Statistiques Globales")
            report_lines.append("")
            glob = stats_t1['globales']
            report_lines.append(f"- **Total graphes analysés:** {glob['count']:,}")
            report_lines.append(f"- **Longueur moyenne:** {glob['longueur_moyenne']:.2f} ± {glob['ecart_type']:.2f}")
            report_lines.append(f"- **Médiane:** {glob['longueur_mediane']:.1f}")
            report_lines.append(f"- **Min - Max:** [{glob['longueur_min']}, {glob['longueur_max']}]")
            report_lines.append(f"- **Quartiles:** Q1 = {glob['q1']:.1f}, Q3 = {glob['q3']:.1f}")
            report_lines.append(f"- **Outliers détectés:** {glob['outliers_count']:,} ({glob['outliers_percentage']:.2f}%)")
            report_lines.append(f"- **Bornes:** [{glob['lower_bound']:.1f}, {glob['upper_bound']:.1f}]")
            report_lines.append("")
        
        # Statistiques par variante
        if 'par_variante' in stats_t1 and stats_t1['par_variante']:
            report_lines.append("### 📊 Statistiques par Variante")
            report_lines.append("")
            report_lines.append("| Variante | Count | Moyenne ± σ | Médiane | Min | Max | Q1 | Q3 | Outliers | Outliers % |")
            report_lines.append("|----------|-------|-------------|---------|-----|-----|----|----|----------|------------|")
            
            for variant, vstats in sorted(stats_t1['par_variante'].items()):
                report_lines.append(
                    f"| {variant} | {vstats['count']:,} | {vstats['longueur_moyenne']:.2f} ± {vstats['ecart_type']:.2f} | "
                    f"{vstats['longueur_mediane']:.1f} | {vstats['longueur_min']} | {vstats['longueur_max']} | "
                    f"{vstats['q1']:.1f} | {vstats['q3']:.1f} | {vstats['outliers_count']:,} | {vstats['outliers_percentage']:.2f}% |"
                )
            report_lines.append("")
        
        # Statistiques par catégorie de cuisine
        if 'par_category' in stats_t1 and stats_t1['par_category']:
            report_lines.append("### 📊 Statistiques par Catégorie de Cuisine")
            report_lines.append("")
            report_lines.append("| Catégorie | Count | Moyenne ± σ | Médiane | Min | Max | Q1 | Q3 | Outliers | Outliers % |")
            report_lines.append("|-----------|-------|-------------|---------|-----|-----|----|----|----------|------------|")
            
            for category, cstats in sorted(stats_t1['par_category'].items()):
                report_lines.append(
                    f"| {category} | {cstats['count']:,} | {cstats['longueur_moyenne']:.2f} ± {cstats['ecart_type']:.2f} | "
                    f"{cstats['longueur_mediane']:.1f} | {cstats['longueur_min']} | {cstats['longueur_max']} | "
                    f"{cstats['q1']:.1f} | {cstats['q3']:.1f} | {cstats['outliers_count']:,} | {cstats['outliers_percentage']:.2f}% |"
                )
            report_lines.append("")
        
        # Statistiques par niveau de complexité
        if 'par_complexity' in stats_t1 and stats_t1['par_complexity']:
            report_lines.append("### 📊 Statistiques par Niveau de Complexité")
            report_lines.append("")
            report_lines.append("| Complexité | Count | Moyenne ± σ | Médiane | Min | Max | Q1 | Q3 | Outliers | Outliers % |")
            report_lines.append("|------------|-------|-------------|---------|-----|-----|----|----|----------|------------|")
            
            complexity_order = ['simple', 'moyenne', 'elevee', 'unknown']
            for complexity in complexity_order:
                if complexity in stats_t1['par_complexity']:
                    cxstats = stats_t1['par_complexity'][complexity]
                    report_lines.append(
                        f"| {complexity} | {cxstats['count']:,} | {cxstats['longueur_moyenne']:.2f} ± {cxstats['ecart_type']:.2f} | "
                        f"{cxstats['longueur_mediane']:.1f} | {cxstats['longueur_min']} | {cxstats['longueur_max']} | "
                        f"{cxstats['q1']:.1f} | {cxstats['q3']:.1f} | {cxstats['outliers_count']:,} | {cxstats['outliers_percentage']:.2f}% |"
                    )
            report_lines.append("")
        
        report_lines.append("---")
        report_lines.append("")
    
    # Test 2
    if 'test_2' in all_stats:
        report_lines.append("## Test 2 : Variante Principale vs Nombre d'Instructions")
        report_lines.append("")
        stats = all_stats['test_2']
        report_lines.append(f"- **Total variantes principales:** {stats['total_principales']:,}")
        report_lines.append(f"- **Ratio moyen (actions/instructions):** {stats['ratio_moyen']:.2f}")
        report_lines.append(f"- **Ratio médian:** {stats['ratio_median']:.2f}")
        report_lines.append("")
        report_lines.append("### Distribution des Flags")
        report_lines.append("")
        report_lines.append(f"| Flag | Count | Pourcentage |")
        report_lines.append(f"|------|-------|-------------|")
        report_lines.append(f"| ✅ CONFORME (0.5 ≤ ratio ≤ 2.5) | {stats['CONFORME_count']:,} | {stats['CONFORME_pct']:.2f}% |")
        report_lines.append(f"| ⚠️ FLAG_RATIO_FAIBLE (0.3 ≤ ratio < 0.5) | {stats['FLAG_RATIO_FAIBLE_count']:,} | {stats['FLAG_RATIO_FAIBLE_pct']:.2f}% |")
        report_lines.append(f"| ⚠️ FLAG_RATIO_ELEVE (2.5 < ratio ≤ 5.0) | {stats['FLAG_RATIO_ELEVE_count']:,} | {stats['FLAG_RATIO_ELEVE_pct']:.2f}% |")
        report_lines.append(f"| ❌ FLAG_CRITIQUE (ratio < 0.3 ou > 5.0) | {stats['FLAG_CRITIQUE_count']:,} | {stats['FLAG_CRITIQUE_pct']:.2f}% |")
        report_lines.append("")
        report_lines.append("---")
        report_lines.append("")
    
    # Test 3
    if 'test_3' in all_stats:
        report_lines.append("## Test 3 : Variante Ingrédients")
        report_lines.append("")
        stats = all_stats['test_3']
        report_lines.append(f"- **Total paires principale-ingrédients:** {stats['total_paires']:,}")
        report_lines.append(f"- **Delta moyen:** {stats['delta_moyen']:.2f}")
        report_lines.append(f"- **Delta médian:** {stats['delta_median']:.1f}")
        report_lines.append(f"- **Delta [min, max]:** [{stats['delta_min']}, {stats['delta_max']}]")
        report_lines.append("")
        report_lines.append("### Distribution des Flags")
        report_lines.append("")
        report_lines.append(f"| Flag | Count | Pourcentage |")
        report_lines.append(f"|------|-------|-------------|")
        report_lines.append(f"| ✅ CONFORME | {stats['CONFORME_count']:,} | {stats['CONFORME_pct']:.2f}% |")
        report_lines.append(f"| ❌ FLAG_CRITIQUE_NEGATIF (delta < 0) | {stats['FLAG_CRITIQUE_NEGATIF_count']:,} | {stats['FLAG_CRITIQUE_NEGATIF_pct']:.2f}% |")
        report_lines.append(f"| ⚠️ FLAG_AUCUN_AJOUT (delta == 0) | {stats['FLAG_AUCUN_AJOUT_count']:,} | {stats['FLAG_AUCUN_AJOUT_pct']:.2f}% |")
        report_lines.append(f"| ⚠️ FLAG_TROP_AJOUTS (delta > nb_ingredients × 2) | {stats['FLAG_TROP_AJOUTS_count']:,} | {stats['FLAG_TROP_AJOUTS_pct']:.2f}% |")
        if 'FLAG_DEPASSEMENT_BORNE_STRICTE_count' in stats:
            report_lines.append(f"| ⚠️ FLAG_DEPASSEMENT_BORNE_STRICTE | {stats['FLAG_DEPASSEMENT_BORNE_STRICTE_count']:,} | {stats['FLAG_DEPASSEMENT_BORNE_STRICTE_pct']:.2f}% |")
        report_lines.append("")
        report_lines.append("---")
        report_lines.append("")
    
    # Test 4A
    if 'test_4a' in all_stats:
        report_lines.append("## Test 4A : Variante Permutation")
        report_lines.append("")
        stats = all_stats['test_4a']
        report_lines.append(f"- **Total paires principale-permutation:** {stats['total_paires']:,}")
        report_lines.append("")
        
        report_lines.append("### Métriques de Similarité et d'Ordre")
        report_lines.append("")
        report_lines.append(f"- **Jaccard (contenu):** Moyenne = {stats['jaccard_mean']:.3f}, Min = {stats['jaccard_min']:.3f}, Max = {stats['jaccard_max']:.3f}")
        report_lines.append(f"- **Levenshtein (ordre):** Moyenne = {stats['levenshtein_mean']:.1f}, Min = {stats['levenshtein_min']}, Max = {stats['levenshtein_max']}")
        report_lines.append("")
        
        report_lines.append("### Distribution des Flags")
        report_lines.append("")
        report_lines.append(f"| Flag | Count | Pourcentage |")
        report_lines.append(f"|------|-------|-------------|")
        report_lines.append(f"| ✅ CONFORME (overlap ∈ [0.6, 1.0] ET levenshtein > 0) | {stats['CONFORME_count']:,} | {stats['CONFORME_pct']:.2f}% |")
        report_lines.append(f"| ❌ FLAG_CRITIQUE_SIMILARITE_BASSE (overlap < 0.6) | {stats['FLAG_CRITIQUE_SIMILARITE_BASSE_count']:,} | {stats['FLAG_CRITIQUE_SIMILARITE_BASSE_pct']:.2f}% |")
        report_lines.append(f"| ⚠️ FLAG_IDENTIQUE (levenshtein = 0) | {stats['FLAG_IDENTIQUE_count']:,} | {stats['FLAG_IDENTIQUE_pct']:.2f}% |")
        report_lines.append("")
        report_lines.append("---")
        report_lines.append("")


    # Test 4B
    if 'test_4b' in all_stats:
        report_lines.append("## Test 4B : Similarité Variante Ingrédients")
        report_lines.append("")
        stats = all_stats['test_4b']
        report_lines.append(f"- **Total paires principale-ingrédients:** {stats['total_paires']:,}")
        report_lines.append(f"- **Overlap moyen:** {stats['overlap_moyen']:.3f}")
        report_lines.append(f"- **Overlap médian:** {stats['overlap_median']:.3f}")
        report_lines.append(f"- **Overlap [min, max]:** [{stats['overlap_min']:.3f}, {stats['overlap_max']:.3f}]")
        report_lines.append("")
        report_lines.append("### Distribution des Flags")
        report_lines.append("")
        report_lines.append(f"| Flag | Count | Pourcentage |")
        report_lines.append(f"|------|-------|-------------|")
        report_lines.append(f"| ✅ CONFORME (overlap ≥ 0.7) | {stats['CONFORME_count']:,} | {stats['CONFORME_pct']:.2f}% |")
        report_lines.append(f"| ⚠️ FLAG_SIMILARITE_MOYENNE (0.5 ≤ overlap < 0.7) | {stats['FLAG_SIMILARITE_MOYENNE_count']:,} | {stats['FLAG_SIMILARITE_MOYENNE_pct']:.2f}% |")
        report_lines.append(f"| ❌ FLAG_CRITIQUE_SIMILARITE_BASSE (overlap < 0.5) | {stats['FLAG_CRITIQUE_SIMILARITE_BASSE_count']:,} | {stats['FLAG_CRITIQUE_SIMILARITE_BASSE_pct']:.2f}% |")
        report_lines.append("")
        report_lines.append("---")
        report_lines.append("")
    
    # Test 6 - COHÉRENCE GLOBALE PAR RECETTE
    if 'test_6' in all_stats:
        report_lines.append("## Test 6 : Cohérence Globale par Recette")
        report_lines.append("")
        stats = all_stats['test_6']
        report_lines.append(f"- **Total recettes analysées:** {stats['total_recettes']:,}")
        report_lines.append("")
        report_lines.append("### Distribution des Flags")
        report_lines.append("")
        
        if 'flags' in stats:
            total = stats['total_recettes']
            flag_no_principale = stats['flags'].get('FLAG_NO_PRINCIPALE', 0)
            flag_multiple = stats['flags'].get('FLAG_MULTIPLE_PRINCIPALE', 0)
            pct_conforme = stats['flags'].get('pct_conformes', 0)
            pct_no_principale = (flag_no_principale / total * 100) if total > 0 else 0
            pct_multiple = (flag_multiple / total * 100) if total > 0 else 0
            
            report_lines.append(f"| Flag | Count | Pourcentage |")
            report_lines.append(f"|------|-------|-------------|")
            report_lines.append(f"| ✅ CONFORME (1 variante principale) | {int(total * pct_conforme / 100):,} | {pct_conforme:.2f}% |")
            report_lines.append(f"| ❌ FLAG_CRITIQUE_NO_PRINCIPALE (aucune variante principale) | {flag_no_principale:,} | {pct_no_principale:.2f}% |")
            report_lines.append(f"| ❌ FLAG_CRITIQUE_MULTIPLE_PRINCIPALE (>1 variante principale) | {flag_multiple:,} | {pct_multiple:.2f}% |")
        
        report_lines.append("")
        report_lines.append("---")
        report_lines.append("")
    
    # Visualisation
    report_lines.append("## Visualisation des Flags")
    report_lines.append("")
    report_lines.append("![Distribution Détaillée des Flags par Test](./visualizations/flag_proportions_par_test.png)")
    report_lines.append("")
    report_lines.append("---")
    report_lines.append("")
    
    # Fichiers générés
    report_lines.append("## Fichiers Générés")
    report_lines.append("")
    report_lines.append(f"- **Dataset des flags critiques:** `{os.path.basename(critical_flags_file)}`")
    report_lines.append(f"- **Visualisation:** `{os.path.basename(visualization_file)}`")
    report_lines.append(f"- **Rapport complet:** `rapport_test_validation_structurelle.md`")
    report_lines.append("")
    report_lines.append("---")
    report_lines.append("")
    
    # Conclusion
    report_lines.append("## Conclusion et Recommandations")
    report_lines.append("")
    report_lines.append("### Synthèse de la Validation")
    report_lines.append("")
    report_lines.append("La Stratégie 2 a permis d'identifier les recettes avec des problèmes structurels critiques à travers 6 tests complémentaires:")
    report_lines.append("")
    report_lines.append("1. **Test 1** : Statistiques descriptives complètes (par variante, catégorie, complexité)")
    report_lines.append("2. **Test 2** : Validation du ratio actions/instructions")
    report_lines.append("3. **Test 3** : Validation de l'ajout de gestes dans la variante ingrédients")
    report_lines.append("4. **Test 4A** : Validation de la permutation (même ensemble)")
    report_lines.append("5. **Test 4B** : Validation de la similarité variante ingrédients")
    report_lines.append("6. **Test 6** : Cohérence globale par recette (présence variante principale)")
    report_lines.append("")
    report_lines.append("### Actions Recommandées")
    report_lines.append("")
    report_lines.append("1. **Recettes avec flags critiques** → Mise de côté pour retraitement ultérieur")
    report_lines.append("2. **Recettes conformes** → Passage à la Stratégie 3 (Détection des Successions Illogiques)")
    report_lines.append("3. **Recettes avec avertissements** → Examen manuel si temps disponible")
    report_lines.append("")
    report_lines.append("---")
    report_lines.append("")
    report_lines.append(f"*Rapport généré automatiquement le {datetime.now().strftime('%Y-%m-%d à %H:%M:%S')}*")
    report_lines.append("")
    
    # Sauvegarder
    report_text = '\n'.join(report_lines)
    report_file = os.path.join(output_dir, 'rapport_test_validation_structurelle.md')
    
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write(report_text)
    
    print(f"✅ Rapport Markdown sauvegardé : {report_file}")
    print(f"   Nombre de sections : {len([l for l in report_lines if l.startswith('##')])}")
    print(f"   Taille du rapport : {len(report_text):,} caractères")
    
    return report_file

# ==============================================================================
# SECTION 11 : PIPELINE PRINCIPAL
# ==============================================================================

def run_strategy_2_pipeline(
    data_csv: str,
    recipes_csv: str,
    output_dir: str = "strategy_2_results"
):
    """
    Exécute le pipeline complet de la Stratégie 2 (VERSION MODIFIÉE)
    
    Args:
        data_csv: Chemin vers le dataset nettoyé avec toutes les actions
        recipes_csv: Chemin vers recipes.csv
        output_dir: Répertoire de sortie
    """
    print(f"\n{'#'*80}")
    print(f"# PIPELINE STRATÉGIE 2 - VALIDATION STRUCTURELLE ")
    print(f"{'#'*80}\n")
    
    # Créer les répertoires de sortie
    create_output_directory(output_dir)
    
    # Chargement des données
    print("ÉTAPE 1/8 : Chargement des données...")
    data_df = pd.read_csv(data_csv)
    recipes_df = pd.read_csv(recipes_csv)
    print(f"  ✅ Data chargé : {len(data_df):,} graphes")
    print(f"  ✅ Recipes chargé : {len(recipes_df):,} recettes")

    print(f"\n🏷️  CLASSIFICATION DES RECETTES...")
    recipes_df['category'] = recipes_df['title'].apply(classify_cuisine_type)
    recipes_df['complexity'] = recipes_df['number_of_steps'].apply(classify_complexity)
    
    print(f"  ✅ Colonne 'category' ajoutée")
    print(f"     Distribution : {dict(recipes_df['category'].value_counts())}")
    print(f"  ✅ Colonne 'complexity' ajoutée")
    print(f"     Distribution : {dict(recipes_df['complexity'].value_counts())}")

    # Stocker toutes les statistiques
    all_stats = {}
    all_flags_dfs = []
    
    # Test 1
    print("\nÉTAPE 2/8 : Test 1 - Statistiques de taille...")
    stats_1, resume_1 = test_1_calculate_action_lengths(data_df, recipes_df)
    all_stats['test_1'] = stats_1
    
    # Test 2
    print("\nÉTAPE 3/8 : Test 2 - Variante principale vs instructions...")
    stats_2, flags_2 = test_2_validate_principale_vs_steps(recipes_df, data_df)
    all_stats['test_2'] = stats_2
    all_flags_dfs.append(flags_2)
    
    # Test 3
    print("\nÉTAPE 4/8 : Test 3 - Variante ingrédients...")
    stats_3, flags_3 = test_3_validate_ingredients_variant(recipes_df, data_df)
    all_stats['test_3'] = stats_3
    all_flags_dfs.append(flags_3)
    
    # Test 4A
    print("\nÉTAPE 5/8 : Test 4A - Variante permutation...")
    stats_4a, flags_4a = test_4a_validate_permutation_variant(data_df)
    all_stats['test_4a'] = stats_4a
    all_flags_dfs.append(flags_4a)
    
    # Test 4B
    print("\nÉTAPE 6/8 : Test 4B - Similarité variante ingrédients...")
    stats_4b, flags_4b = test_4b_validate_ingredients_similarity(data_df)
    all_stats['test_4b'] = stats_4b
    all_flags_dfs.append(flags_4b)
    
    # Test 6
    print("\nÉTAPE 7/8 : Test 6 - Détection des listes vides...")
    stats_6, flags_6 = test_6_validate_recipe_coherence(data_df)
    all_stats['test_6'] = stats_6
    all_flags_dfs.append(flags_6)
    
    # Export du dataset des flags critiques
    print("\nÉTAPE 8/8 : Export et génération des rapports...")
    critical_flags_file = export_critical_flags_dataset(all_flags_dfs, output_dir)
    
    # Visualisation
    visualization_file = plot_flag_proportions(all_stats, output_dir)
    
    # Génération du rapport Markdown
    report_file = generate_validation_report_strat_2(
        all_stats,
        critical_flags_file,
        visualization_file,
        output_dir
    )
    
    print(f"\n{'#'*80}")
    print(f"# PIPELINE STRATÉGIE 2 TERMINÉ AVEC SUCCÈS")
    print(f"{'#'*80}\n")
    print(f"📁 Tous les résultats sont dans : {output_dir}/")
    print(f"   - Dataset flags critiques : {critical_flags_file}")
    print(f"   - Visualisation : {visualization_file}")
    print(f"   - Rapport Markdown : {report_file}")
    print(f"\n✅ Stratégie 2 complétée !\n")





""" *********************************************************************************
Pipeline Stratégie 3 : Validation Sémantique 
Détection des successions illogiques dans les graphes de gestes

    **********************************************************************************
"""

import pandas as pd
import numpy as np
import ast
import os
import json
from typing import List, Dict, Tuple, Set
from datetime import datetime
import matplotlib.pyplot as plt
import seaborn as sns


# ==============================================================================
# SECTION 1 : TAXONOMIE DES VERBES CULINAIRES
# ==============================================================================

class VerbTaxonomy:
    """
    Taxonomie complète des verbes culinaires en 11 catégories fonctionnelles
    """
  
    
    # ========================================================================
    # CATÉGORIE 1: PRÉPARATION INITIALE
    # Actions de nettoyage et préparation de base des ingrédients bruts
    # ========================================================================
    PREPARATION_INITIALE = {
        'wash', 'rinse', 'clean', 'cleanse', 'scrub', 'peel', 'pare', 'trim', 
        'core', 'pit', 'stone', 'bone', 'debone', 'devein', 'vein', 'scale', 
        'hull', 'husk', 'shell', 'unshell', 'shuck', 'gut', 'eviscerate', 
        'beard', 'behead', 'head', 'skin', 'flay', 'stem', 'deseed', 'seed', 
        'gill', 'fish', 'fillet', 'filet', 'butcher', 'cavity', 'disjoint',
        'supreme', 'spatchcock', 'butterfly', 'clip', 'harvest', 'pick',
        'bathe', 'soak', 'freshen', 'sterilize', 'cleanse'
    }
    
    # ========================================================================
    # CATÉGORIE 2: TRANSFORMATION MÉCANIQUE
    # Actions de découpe, broyage et modification physique de la structure
    # ========================================================================
    TRANSFORMATION_MECANIQUE = {
        'chop', 'prechop', 'dice', 'slice', 'cut', 'mince', 'julienne', 
        'cube', 'halve', 'quarter', 'trisect', 'grate', 'shred', 'crush', 
        'grind', 'mill', 'pound', 'hammer', 'tenderize', 'mash', 'remash',
        'puree', 'purée', 'shave', 'zest', 'chiffonade', 'brunoise', 
        'concasse', 'carve', 'tournee', 'chip', 'chunk', 'crumb', 'crumble', 
        'flake', 'smash', 'pulverise', 'splitter', 'hacken', 'wedge', 
        'crosscut', 'crosshatch', 'pieces', 'piece', 'break', 'crack', 
        'snap', 'tear', 'split', 'cleave', 'slash', 'slit', 'sliver',
        'gash', 'nick', 'nip', 'score', 'incision', 'process', 'blitz', 
        'blenderize', 'whiz', 'whizz', 'liquidise', 'texturize', 'fragment', 
        'powder', 'powderize', 'crunch', 'splinter', 'shatter', 'burst',
        'bruise', 'chisel', 'plane', 'shear', 'strip', 'finely',
        'moosh', 'mush', 'smoosh', 'smush', 'squish', 'compress',
        'diagonal', 'slather', 'scrape', 'pique', 'dock', 'finely'
    }
    
    # ========================================================================
    # CATÉGORIE 3: MÉLANGE ET COMBINAISON
    # Actions d'incorporation et homogénéisation d'ingrédients
    # ========================================================================
    MELANGE_COMBINAISON = {
        'mix', 'premix', 'remix', 'overmix', 'stir', 'combine', 'recombine',
        'whisk', 'rewhisk', 'beat', 'whip', 'fold', 'blend', 'reblend',
        'toss', 'incorporate', 'reincorporate', 'emulsify', 'cream', 
        'knead', 'work', 'rework', 'amalgamate', 'meld', 'mingle', 'marry', 
        'muddle', 'agitate', 'churn', 'swirl', 'swish', 'swizzle', 
        'scramble', 'massage', 'pulsate', 'pulse', 'claw', 'froth', 
        'foam', 'bubble', 'aerate', 'fluff', 'lighten', 'crank',
        'alternate', 'interlace', 'interleave', 'weave', 'marbleize'
    }
    
    # ========================================================================
    # CATÉGORIE 4: TRANSFERT ET MANIPULATION
    # Actions de placement, assemblage et application de matières
    # ========================================================================
    TRANSFERT_MANIPULATION = {
        'pour', 'pour off', 'pour out', 'add', 'place', 'put', 'transfer', 
        'spread', 'layer', 'arrange', 'rearrange', 'fill', 'refill', 
        'stuff', 'unstuff', 'wrap', 'unwrap', 'rewrap', 'coat', 'recoat',
        'brush', 'drizzle', 'sprinkle', 'dust', 'top', 'spoon', 'scoop', 
        'ladle', 'dab', 'dollop', 'blob', 'dot', 'drop', 'splash', 'splat',
        'dish', 'portion', 'stack', 'unstack', 'pile', 'mound', 'mount',
        'nest', 'nestle', 'cradle', 'tuck', 'cover', 'uncover', 'blanket', 
        'tent', 'foil', 'unfoil', 'apply', 'smear', 'grease', 'ungrease',
        'butter', 'flour', 'egg', 'egg wash', 'batter', 'bread', 
        'moisten', 'moisturize', 'wet', 'damp', 'dampen', 'drench', 'douse',
        'bathe', 'soak', 'spritz', 'spray', 'mist', 'squirt', 'drape', 
        'glaze', 'reglaze', 'unglaze', 'frost', 'ice', 'sugar', 'season',
        'reseason', 'funnel', 'decant', 'dredge', 'anoint', 'dress',
        'enrobe', 'encrust', 'encase', 'encapsulate', 'enclose', 'enfold',
        'film', 'cap', 'cork', 'bag', 'baggie', 'package', 'bundle',
        'bung', 'throw away', 'discard', 'dispose', 'ditch', 'remove',
        'retrieve', 'return', 'move', 'slide', 'tilt', 'incline', 'lean',
        'tip', 'tip out', 'tipping', 'pitch', 'chuck', 'dump', 'plonk',
        'plop', 'plunk', 'cram', 'pack', 'tamp', 'pad', 'feed',
        'dispense', 'dribble', 'drip', 'trickle', 'stream', 'rain',
        'puddle', 'float', 'submerge', 'drown', 'dunk', 'dip',
        'smother', 'stud', 'spike', 'spear', 'skewer', 'unskewer',
        'poke', 'pierce', 'needle', 'spit', 'unspit', 'spice',
        'flavor', 'enrich', 'color', 'darken', 'whiten', 'candy',
        'acidulate', 'sweeten', 'dash', 'spill', 'reintroduce'
    }
    
    # ========================================================================
    # CATÉGORIE 5: CUISSON ACTIVE
    # Actions de cuisson nécessitant manipulation et surveillance constante
    # ========================================================================
    CUISSON_ACTIVE = {
        'sauté', 'sautee', 'saute', 'stir-fry', 'stir fry', 'pan-fry', 
        'pan fry', 'shallow fry', 'sear', 'brown', 'caramelize', 'reduce',
        'flip', 'turn', 'baste', 'fry', 'refry', 'deep-fry', 'deep fry',
        'flash fry', 'air fry', 'griddle', 'sweat', 'char', 'blacken', 
        'singe', 'scorch', 'blister', 'crisp', 'crackle', 'sizzle', 
        'deglaze', 'toss', 'rotate', 'shake', 'fricassee', 'devil',
        'scramble', 'frizzle'
    }
    
    # ========================================================================
    # CATÉGORIE 6: CUISSON PASSIVE
    # Méthodes de cuisson où la chaleur/temps fait le travail
    # ========================================================================
    CUISSON_PASSIVE = {
        'bake', 'prebake', 'parbake', 'overbake', 'underbake', 'roast', 
        'droast', 'slow roast', 'spit roast', 'dry roast', 'boil', 'reboil',
        'hard boil', 'parboil', 'simmer', 'steam', 'broil', 'grill', 
        'poach', 'braise', 'slow-cook', 'slow cook', 'pressure-cook', 
        'pressure cook', 'pressurize', 'smoke', 'barbecue', 'toast', 
        'stew', 'coddle', 'parcook', 'precook', 'overcook', 'undercook',
        'blanch', 'scald', 'flambe', 'flambé', 'brûlée', 'flame', 'fire', 
        'blast', 'blaze', 'burn', 'cook', 'microwave', 'nuke', 'torch',
        'gratinee', 'roil', 'bubble', 'shimmer', 'confit', 'pasteurize'
    }
    
    # ========================================================================
    # CATÉGORIE 7: TRANSFORMATION THERMIQUE
    # Changements de température et d'état physique
    # ========================================================================
    TRANSFORMATION_THERMIQUE = {
        'cool', 'chill', 'freeze', 'refreeze', 'flash freeze', 'churn freeze',
        'freezer', 'refrigerate', 'thaw', 'defrost', 'warm', 'rewarm',
        'heat', 'reheat', 'preheat', 'melt', 'remelt', 'temper', 'room',
        'refresh', 'ice', 'congeal', 'solidify', 'crystallize', 'set',
        'reset', 'firm', 'harden', 'soften', 'liquefy', 'liquidise',
        'gel', 'coagulate', 'curdle', 'condense', 'evaporate', 'reduce',
        'thicken', 'thin', 'dissolve', 'dehydrate', 'dry', 'air dry',
        'parch', 'wilt', 'shrink', 'swell', 'expand', 'bloom', 'hydrate',
        'rehydrate', 'restore', 'flash', 'quench', 'scald', 'carbonate',
        'degas'
    }
    
    # ========================================================================
    # CATÉGORIE 8: ATTENTE ET REPOS
    # Processus passifs dépendant du temps (fermentation, maturation)
    # ========================================================================
    ATTENTE_REPOS = {
        'rest', 'set', 'settle', 'rise', 'proof', 'ferment', 'marinate', 
        'pickle', 'cure', 'age', 'soak', 'presoak', 'steep', 'mature', 
        'ripen', 'autolyse', 'leaven', 'inoculate', 'culture', 'preserve', 
        'can', 'rebottle', 'confit', 'brine', 'souse', 'macerate', 'mull',
        'leave', 'hang', 'infuse', 'brew', 'percolate', 'perk', 'activate',
        'develop', 'mellow', 'season', 'idle', 'wait', 'stand', 'absorb',
        'reabsorb', 'adhere', 'bind', 'sprout', 'convert', 'transform'
    }
    
    # ========================================================================
    # CATÉGORIE 9: EXTRACTION ET SÉPARATION
    # Retrait de liquides, filtration et clarification
    # ========================================================================
    EXTRACTION_SEPARATION = {
        'drain', 'strain', 'press', 'pressing', 'squeeze', 'filter', 
        'sift', 'sieve', 'separate', 'skim', 'scum', 'extract', 'wring', 
        'decant', 'siphon', 'leach', 'degas', 'deglaze', 'degrease', 
        'discharge', 'clarify', 'declump', 'render', 'distill', 
        'percolate', 'perk', 'seep', 'ooze', 'drip', 'blot', 'wipe',
        'remove', 'discard', 'scrape', 'express', 'extrude', 'force',
        'juice', 'dilute', 'reconstitute', 'reprocess', 'take off',
        'pour off', 'sweep', 'mop', 'patting'
    }
    
    # ========================================================================
    # CATÉGORIE 10: FINITION ET SERVICE
    # Décoration, présentation et dernières touches
    # ========================================================================
    FINITION_SERVICE = {
        'garnish', 'plate', 'serve', 'decorate', 'present', 'unmold', 
        'display', 'portion', 'ornament', 'dress', 'glaze', 'unglaze',
        'frost', 'brûlée', 'finish', 'nap', 'rim', 'mark', 'embellish', 
        'adorn', 'burnish', 'polish', 'buff', 'shine', 'gloss', 'tip',
        'crown', 'top', 'drizzle', 'dust', 'fringe', 'feather',
        'quenelle', 'dome', 'ripple', 'swirl', 'marbleize'
    }
    
    # ========================================================================
    # CATÉGORIE 11: ACTIONS SPÉCIALES
    # Façonnage, structuration et techniques spécialisées
    # ========================================================================
    ACTIONS_SPECIALES = {
        'roll', 'reroll', 'unroll', 'shape', 'reshape', 'form', 'reform',
        'mold', 'mould', 'unmold', 'score', 'pierce', 'prick', 'skewer',
        'unskewer', 'thread', 'unthread', 'tie', 'untie', 'secure', 
        'truss', 'untruss', 'bind', 'braid', 'plait', 'twine', 'untwine',
        'weave', 'coil', 'spiral', 'spiralize', 'curl', 'uncurl', 'twist',
        'untwist', 'twirl', 'crimp', 'flute', 'pleat', 'fold', 'unfold',
        'pinwheel', 'snake', 'wrap', 'unwrap', 'open', 'reopen', 'close',
        'seal', 'reseal', 'unseal', 'attach', 'detach', 'fasten', 'unfasten',
        'flatten', 'smooth', 'level', 'level off', 'even', 'round', 
        'square', 'curve', 'bend', 'sandwich', 'layer', 'laminate',
        'stack', 'tier', 'build', 'assemble', 'reassemble', 'construct',
        'gather', 'bunch', 'cluster', 'clump', 'bundle', 'spool',
        'stretch', 'compress', 'taper', 'ridge', 'groove', 'channel',
        'seam', 'crease', 'crinkle', 'wrinkle', 'ruffle', 'crimp',
        'flute', 'scallop', 'notch', 'indent', 'dent', 'dimple',
        'pocket', 'pouch', 'bag', 'encase', 'cradle', 'nest',
        'tuck', 'snug', 'seat', 'position', 'reposition', 'center',
        'adjust', 'readjust', 'straighten', 'neaten', 'arrange',
        'fan', 'splay', 'overlap', 'cross', 'interlock', 'hook',
        'string', 'lace', 'stitch', 'sew', 'pin', 'clip', 'clamp',
        'frame', 'border', 'edge', 'fringe', 'trim', 'finish',
        'pip', 'pipe', 'fork', 'comb', 'rake', 'score', 'mark',
        'stamp', 'imprint', 'emboss', 'engrave', 'etch', 'carve',
        'mold', 'cast', 'press', 'punch', 'punch down', 'knock',
        'tap', 'pat', 'brush', 'stroke', 'caress', 'smooth',
        'rough', 'roughen', 'coarsen', 'texturize', 'stiffen',
        'firm', 'stabilize', 'reinforce', 'strengthen', 'support',
        'sharpen', 'blunt', 'dull', 'hone', 'whet', 'grind',
        'unwind', 'unfurl', 'uncoil', 'unravel', 'loosen', 'relax',
        'nudge', 'bump', 'jostle', 'shake', 'rattle', 'vibrate',
        'handle', 'manipulate', 'work', 'massage', 'knead', 'press',
        'finger', 'thumb', 'poke', 'prod', 'push', 'pull', 'draw',
        'blow', 'puff', 'inflate', 'deflate', 'expand', 'contract',
        'hop', 'bounce', 'spring', 'flip', 'turn', 'rotate', 'spin',
        'turn off', 'switch', 'toggle', 'control', 'regulate',
        'moderate', 'modulate', 'temper', 'balance', 'equilibrate',
        'measure', 'premeasure', 'weigh', 'gauge', 'estimate',
        'test', 'check', 'verify', 'confirm', 'validate',
        'taste', 'sample', 'try', 'evaluate', 'assess',
        'prep', 'prepare', 'ready', 'setup', 'organize',
        'decrease', 'increase', 'double', 'halve', 'reduce',
        'splodge', 'sploosh', 'splotch', 'smudge', 'blur',
        'crust', 'skin', 'film', 'coat', 'cover', 'varnish',
        'frappe', 'crush', 'shave', 'grate', 'zest',
        'wrapper', 'enfold', 'envelop', 'sheathe', 'sleeve',
        'batten', 'scrunch'
    }
    
    CATEGORY_NAMES = {
        'PREPARATION_INITIALE': 'Préparation Initiale',
        'TRANSFORMATION_MECANIQUE': 'Transformation Mécanique',
        'MELANGE_COMBINAISON': 'Mélange et Combinaison',
        'TRANSFERT_MANIPULATION': 'Transfert et Manipulation',
        'CUISSON_ACTIVE': 'Cuisson Active',
        'CUISSON_PASSIVE': 'Cuisson Passive',
        'TRANSFORMATION_THERMIQUE': 'Transformation Thermique',
        'ATTENTE_REPOS': 'Attente et Repos',
        'EXTRACTION_SEPARATION': 'Extraction et Séparation',
        'FINITION_SERVICE': 'Finition et Service',
        'ACTIONS_SPECIALES': 'Actions Spéciales'
    }
    
    @classmethod
    def get_category(cls, verb: str) -> str:
        """Retourne la catégorie d'un verbe"""
        verb_lower = verb.lower().strip()
        
        if verb_lower in cls.PREPARATION_INITIALE:
            return 'PREPARATION_INITIALE'
        elif verb_lower in cls.TRANSFORMATION_MECANIQUE:
            return 'TRANSFORMATION_MECANIQUE'
        elif verb_lower in cls.MELANGE_COMBINAISON:
            return 'MELANGE_COMBINAISON'
        elif verb_lower in cls.TRANSFERT_MANIPULATION:
            return 'TRANSFERT_MANIPULATION'
        elif verb_lower in cls.CUISSON_ACTIVE:
            return 'CUISSON_ACTIVE'
        elif verb_lower in cls.CUISSON_PASSIVE:
            return 'CUISSON_PASSIVE'
        elif verb_lower in cls.TRANSFORMATION_THERMIQUE:
            return 'TRANSFORMATION_THERMIQUE'
        elif verb_lower in cls.ATTENTE_REPOS:
            return 'ATTENTE_REPOS'
        elif verb_lower in cls.EXTRACTION_SEPARATION:
            return 'EXTRACTION_SEPARATION'
        elif verb_lower in cls.FINITION_SERVICE:
            return 'FINITION_SERVICE'
        elif verb_lower in cls.ACTIONS_SPECIALES:
            return 'ACTIONS_SPECIALES'
        else:
            return 'UNKNOWN'
    
    @classmethod
    def get_all_verbs(cls) -> Set[str]:
        """Retourne l'ensemble de tous les verbes connus"""
        all_verbs = set()
        for category_set in [
            cls.PREPARATION_INITIALE, cls.TRANSFORMATION_MECANIQUE,
            cls.MELANGE_COMBINAISON, cls.TRANSFERT_MANIPULATION,
            cls.CUISSON_ACTIVE, cls.CUISSON_PASSIVE,
            cls.TRANSFORMATION_THERMIQUE, cls.ATTENTE_REPOS,
            cls.EXTRACTION_SEPARATION, cls.FINITION_SERVICE,
            cls.ACTIONS_SPECIALES
        ]:
            all_verbs.update(category_set)
        return all_verbs
    
    @classmethod
    def export_taxonomy(cls, output_file: str):
        """Exporte la taxonomie complète en JSON"""
        taxonomy = {}
        for attr_name in dir(cls):
            if attr_name.isupper() and not attr_name.startswith('CATEGORY'):
                category_set = getattr(cls, attr_name)
                if isinstance(category_set, set):
                    taxonomy[attr_name] = {
                        'name': cls.CATEGORY_NAMES.get(attr_name, attr_name),
                        'verbs': sorted(list(category_set)),
                        'count': len(category_set)
                    }
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(taxonomy, f, indent=2, ensure_ascii=False)
        
        print(f"✅ Taxonomie exportée : {output_file}")
        print(f"📊 Total verbes : {len(cls.get_all_verbs())}")


# ==============================================================================
# SECTION 2 : RÈGLES DE SUCCESSION 
# ==============================================================================

class SuccessionRules:
    """
    Règles de succession impossibles et suspectes
    
    MODIFICATIONS:
    - Règle A2_IRREVERSIBILITE_PHYSIQUE supprimée
    - 3 règles Type A conservées (A1, A2=ancien A3, A3=ancien A4)
    - Toutes les règles Type B conservées
    """
    
    # Règles Type A : ERREURS CERTAINES (3 règles)
    TYPE_A_RULES = {
        'A1_IRREVERSIBILITE_TEMPORELLE': {
            'description': 'FINITION → PRÉPARATION INITIALE',
            'from_category': 'FINITION_SERVICE',
            'to_category': 'PREPARATION_INITIALE',
            'severity': 'CRITICAL',
            'examples': ['serve → wash', 'plate → peel', 'garnish → clean']
        },
        'A2_ILLOGISME_FINITION_MELANGE': {
            'description': 'FINITION → MÉLANGE',
            'from_category': 'FINITION_SERVICE',
            'to_category': 'MELANGE_COMBINAISON',
            'severity': 'CRITICAL',
            'examples': ['serve → mix', 'plate → stir', 'garnish → combine']
        },
        'A3_ILLOGISME_FINITION_CUISSON': {
            'description': 'FINITION → CUISSON',
            'from_category': 'FINITION_SERVICE',
            'to_categories': ['CUISSON_ACTIVE', 'CUISSON_PASSIVE'],
            'severity': 'CRITICAL',
            'examples': ['serve → bake', 'garnish → sauté', 'plate → roast']
        }
    }
    
    # Règles Type B : SUSPICIONS (toutes conservées)
    TYPE_B_RULES = {
        'B1_SAUT_DE_PHASE': {
            'description': 'PRÉPARATION → FINITION (sans étapes intermédiaires)',
            'from_categories': ['PREPARATION_INITIALE', 'TRANSFORMATION_MECANIQUE'],
            'to_category': 'FINITION_SERVICE',
            'severity': 'SUSPICIOUS',
            'examples': ['chop → serve', 'wash → plate']
        },
        'B2_ORDRE_INVERSE': {
            'description': 'TRANSFERT → TRANSFORMATION MÉCANIQUE',
            'from_category': 'TRANSFERT_MANIPULATION',
            'to_category': 'TRANSFORMATION_MECANIQUE',
            'severity': 'SUSPICIOUS',
            'examples': ['pour → chop', 'add → dice']
        },
        'B3_CUISSON_SANS_PREPARATION': {
            'description': 'CUISSON comme première action',
            'severity': 'SUSPICIOUS',
            'examples': ['bake (first)', 'boil (first)']
        },
        'B4_MELANGE_APRES_CUISSON_PASSIVE': {
            'description': 'CUISSON PASSIVE → MÉLANGE (sans EXTRACTION)',
            'from_category': 'CUISSON_PASSIVE',
            'to_category': 'MELANGE_COMBINAISON',
            'severity': 'SUSPICIOUS',
            'examples': ['bake → mix', 'roast → stir']
        },
        'B5_TRANSFORMATION_APRES_TRANSFERT_FINAL': {
            'description': 'TRANSFERT (plat) → TRANSFORMATION MÉCANIQUE',
            'from_category': 'TRANSFERT_MANIPULATION',
            'to_category': 'TRANSFORMATION_MECANIQUE',
            'severity': 'SUSPICIOUS',
            'examples': ['plate → slice', 'serve → cut']
        }
    }
    
    @classmethod
    def check_type_a_violation(
        cls, 
        cat_from: str, 
        cat_to: str
    ) -> Tuple[bool, str, str]:
        """
        Vérifie si une transition viole une règle Type A
        
        MODIFIÉ: Règle A2_IRREVERSIBILITE_PHYSIQUE supprimée
        
        Returns:
            (violation_detected, rule_id, description)
        """
        # A1: FINITION → PRÉPARATION
        if cat_from == 'FINITION_SERVICE' and cat_to == 'PREPARATION_INITIALE':
            return (True, 'A1_IRREVERSIBILITE_TEMPORELLE', 
                    cls.TYPE_A_RULES['A1_IRREVERSIBILITE_TEMPORELLE']['description'])
        
        # A2: FINITION → MÉLANGE (ancien A3)
        if cat_from == 'FINITION_SERVICE' and cat_to == 'MELANGE_COMBINAISON':
            return (True, 'A2_ILLOGISME_FINITION_MELANGE',
                    cls.TYPE_A_RULES['A2_ILLOGISME_FINITION_MELANGE']['description'])
        
        # A3: FINITION → CUISSON (ancien A4)
        if cat_from == 'FINITION_SERVICE' and cat_to in ['CUISSON_ACTIVE', 'CUISSON_PASSIVE']:
            return (True, 'A3_ILLOGISME_FINITION_CUISSON',
                    cls.TYPE_A_RULES['A3_ILLOGISME_FINITION_CUISSON']['description'])
        
        return (False, '', '')
    
    @classmethod
    def check_type_b_violation(
        cls,
        cat_from: str,
        cat_to: str,
        is_first_action: bool = False
    ) -> Tuple[bool, str, str]:
        """
        Vérifie si une transition viole une règle Type B
        
        Returns:
            (violation_detected, rule_id, description)
        """
        # B1: PRÉPARATION → FINITION (saut de phase)
        if cat_from in ['PREPARATION_INITIALE', 'TRANSFORMATION_MECANIQUE'] and cat_to == 'FINITION_SERVICE':
            return (True, 'B1_SAUT_DE_PHASE',
                    cls.TYPE_B_RULES['B1_SAUT_DE_PHASE']['description'])
        
        # B2: TRANSFERT → TRANSFORMATION
        if cat_from == 'TRANSFERT_MANIPULATION' and cat_to == 'TRANSFORMATION_MECANIQUE':
            return (True, 'B2_ORDRE_INVERSE',
                    cls.TYPE_B_RULES['B2_ORDRE_INVERSE']['description'])
        
        # B3: CUISSON en première action
        if is_first_action and cat_from in ['CUISSON_ACTIVE', 'CUISSON_PASSIVE']:
            return (True, 'B3_CUISSON_SANS_PREPARATION',
                    cls.TYPE_B_RULES['B3_CUISSON_SANS_PREPARATION']['description'])
        
        # B4: CUISSON PASSIVE → MÉLANGE
        if cat_from == 'CUISSON_PASSIVE' and cat_to == 'MELANGE_COMBINAISON':
            return (True, 'B4_MELANGE_APRES_CUISSON_PASSIVE',
                    cls.TYPE_B_RULES['B4_MELANGE_APRES_CUISSON_PASSIVE']['description'])
        
        # B5: TRANSFERT → TRANSFORMATION (après service)
        if cat_from == 'TRANSFERT_MANIPULATION' and cat_to == 'TRANSFORMATION_MECANIQUE':
            return (True, 'B5_TRANSFORMATION_APRES_TRANSFERT_FINAL',
                    cls.TYPE_B_RULES['B5_TRANSFORMATION_APRES_TRANSFERT_FINAL']['description'])
        
        return (False, '', '')


# ==============================================================================
# SECTION 3 : UTILITAIRES
# ==============================================================================

def filter_critical_recipes(
    graphs_df: pd.DataFrame,
    critical_flags_csv: str
) -> pd.DataFrame:
    """
    Filtre le dataset pour retirer les recettes flaggées comme critiques par la Stratégie 2
    
    Args:
        graphs_df: DataFrame complet des graphes
        critical_flags_csv: Chemin vers dataset_flags_critique.csv de la Stratégie 2
    
    Returns:
        DataFrame filtré (sans les recettes critiques de Stratégie 2)
    """
    print(f"\n{'='*80}")
    print(f"FILTRAGE DES RECETTES CRITIQUES (STRATÉGIE 2)")
    print(f"{'='*80}\n")
    
    # Charger le dataset des flags critiques de Stratégie 2
    if not os.path.exists(critical_flags_csv):
        print(f"⚠️  Fichier {critical_flags_csv} introuvable.")
        print(f"   Aucun filtrage appliqué - analyse de tout le dataset.")
        return graphs_df
    
    critical_flags_df = pd.read_csv(critical_flags_csv)
    
    # Extraire les IDs uniques des recettes critiques
    critical_recipe_ids = set(critical_flags_df['id'].unique())
    
    # Statistiques avant filtrage
    total_recipes_before = graphs_df['id'].nunique()
    total_graphs_before = len(graphs_df)
    
    # Filtrer le dataset
    filtered_df = graphs_df[~graphs_df['id'].isin(critical_recipe_ids)].copy()
    
    # Statistiques après filtrage
    total_recipes_after = filtered_df['id'].nunique()
    total_graphs_after = len(filtered_df)
    
    recipes_removed = total_recipes_before - total_recipes_after
    graphs_removed = total_graphs_before - total_graphs_after
    
    pct_recipes_removed = (recipes_removed / total_recipes_before * 100) if total_recipes_before > 0 else 0
    pct_graphs_removed = (graphs_removed / total_graphs_before * 100) if total_graphs_before > 0 else 0
    
    # Affichage des statistiques
    print(f"📊 Statistiques de filtrage:")
    print(f"\n   AVANT filtrage:")
    print(f"      - Recettes uniques: {total_recipes_before:,}")
    print(f"      - Graphes totaux: {total_graphs_before:,}")
    
    print(f"\n   RECETTES RETIRÉES (flags critiques Stratégie 2):")
    print(f"      - Recettes: {recipes_removed:,} ({pct_recipes_removed:.2f}%)")
    print(f"      - Graphes: {graphs_removed:,} ({pct_graphs_removed:.2f}%)")
    
    print(f"\n   APRÈS filtrage:")
    print(f"      - Recettes conservées: {total_recipes_after:,} ({100-pct_recipes_removed:.2f}%)")
    print(f"      - Graphes conservés: {total_graphs_after:,} ({100-pct_graphs_removed:.2f}%)")
    
    print(f"\n✅ Filtrage terminé - analyse sur {total_recipes_after:,} recettes conformes\n")
    
    return filtered_df


def parse_actions_column(actions_value):
    """Parse robustement la colonne 'actions'"""
    if pd.isna(actions_value):
        return []
    
    if isinstance(actions_value, list):
        return actions_value
    
    if isinstance(actions_value, str):
        try:
            parsed = ast.literal_eval(actions_value)
            if isinstance(parsed, list):
                return parsed
        except (ValueError, SyntaxError):
            if ',' in actions_value:
                return [a.strip() for a in actions_value.split(',') if a.strip()]
            return [actions_value.strip()] if actions_value.strip() else []
    
    return []


def create_output_directory(output_dir: str):
    """Crée le répertoire de sortie s'il n'existe pas"""
    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(os.path.join(output_dir, 'errors'), exist_ok=True)
    os.makedirs(os.path.join(output_dir, 'statistics'), exist_ok=True)
    os.makedirs(os.path.join(output_dir, 'visualizations'), exist_ok=True)


# ==============================================================================
# SECTION 4 : ANNOTATION DES SÉQUENCES
# ==============================================================================

def annotate_sequence_with_categories(
    actions: List[str]
) -> List[Tuple[str, str]]:
    """
    Annote chaque action avec sa catégorie
    
    Args:
        actions: Liste d'actions
    
    Returns:
        Liste de tuples (action, category)
    """
    annotated = []
    for action in actions:
        category = VerbTaxonomy.get_category(action)
        annotated.append((action, category))
    
    return annotated


# ==============================================================================
# SECTION 5 : DÉTECTION DES VIOLATIONS
# ==============================================================================

def detect_violations(
    recipe_id: int,
    title: str,
    actions: List[str],
    annotated_sequence: List[Tuple[str, str]]
) -> Dict:
    """
    Détecte toutes les violations dans une séquence
    
    Args:
        recipe_id: ID de la recette
        title: Titre de la recette
        actions: Liste d'actions
        annotated_sequence: Séquence annotée [(action, category), ...]
    
    Returns:
        Dictionnaire avec les violations détectées
    """
    violations = {
        'recipe_id': recipe_id,
        'title': title,
        'sequence_length': len(actions),
        'unknown_verbs': [],
        'type_a_violations': [],
        'type_b_violations': [],
        'repetitions': []
    }
    
    # Détecter les verbes inconnus
    for action, category in annotated_sequence:
        if category == 'UNKNOWN':
            violations['unknown_verbs'].append(action)
    
    # Détecter les répétitions suspectes (>3 fois consécutives)
    i = 0
    while i < len(actions):
        action = actions[i]
        count = 1
        while i + count < len(actions) and actions[i + count] == action:
            count += 1
        
        if count > 3:
            violations['repetitions'].append({
                'action': action,
                'count': count,
                'position': i
            })
        
        i += count
    
    # Détecter les violations Type A et Type B
    for i in range(len(annotated_sequence) - 1):
        action_i, cat_i = annotated_sequence[i]
        action_j, cat_j = annotated_sequence[i + 1]
        
        # Vérifier Type A
        is_violation_a, rule_id_a, desc_a = SuccessionRules.check_type_a_violation(cat_i, cat_j)
        if is_violation_a:
            violations['type_a_violations'].append({
                'rule_id': rule_id_a,
                'description': desc_a,
                'succession': f"{action_i} → {action_j}",
                'position': i
            })
        
        # Vérifier Type B
        is_first = (i == 0)
        is_violation_b, rule_id_b, desc_b = SuccessionRules.check_type_b_violation(cat_i, cat_j, is_first)
        if is_violation_b:
            violations['type_b_violations'].append({
                'rule_id': rule_id_b,
                'description': desc_b,
                'succession': f"{action_i} → {action_j}",
                'position': i
            })
    
    return violations


# ==============================================================================
# SECTION 6 : ANALYSE DU DATASET
# ==============================================================================

def analyze_dataset(
    graphs_df: pd.DataFrame,
    output_dir: str
) -> Tuple[pd.DataFrame, Dict]:
    """
    Analyse l'ensemble du dataset et détecte toutes les violations
    
    Args:
        graphs_df: DataFrame des graphes
        output_dir: Répertoire de sortie
    
    Returns:
        Tuple (violations_df, statistics_dict)
    """
    print(f"\n{'='*80}")
    print(f"ANALYSE DU DATASET - DÉTECTION DES VIOLATIONS")
    print(f"{'='*80}\n")
    
    all_violations = []
    
    # Parser les actions
    graphs_df = graphs_df.copy()
    graphs_df['actions_parsed'] = graphs_df['actions'].apply(parse_actions_column)
    
    total_graphs = len(graphs_df)
    print(f"Total graphes à analyser : {total_graphs:,}")
    
    for idx, row in graphs_df.iterrows():
        if (idx + 1) % 1000 == 0:
            print(f"  Progression : {idx+1:,}/{total_graphs:,} ({(idx+1)/total_graphs*100:.1f}%)")
        
        recipe_id = row['id']
        title = row.get('title', f'Recipe {recipe_id}')
        actions = row['actions_parsed']
        
        if len(actions) == 0:
            continue
        
        # Annoter la séquence
        annotated = annotate_sequence_with_categories(actions)
        
        # Détecter les violations
        violations = detect_violations(recipe_id, title, actions, annotated)
        all_violations.append(violations)
    
    print(f"\n✅ Analyse terminée : {len(all_violations):,} graphes analysés")
    
    # Créer le DataFrame des violations
    violations_df = pd.DataFrame(all_violations)
    
    # Calculer les statistiques
    statistics = calculate_statistics(violations_df)
    
    return violations_df, statistics


def calculate_statistics(violations_df: pd.DataFrame) -> Dict:
    """
    Calcule les statistiques globales sur les violations
    
    Args:
        violations_df: DataFrame des violations
    
    Returns:
        Dictionnaire de statistiques
    """
    print(f"\n{'='*80}")
    print(f"CALCUL DES STATISTIQUES")
    print(f"{'='*80}\n")
    
    total_graphs = len(violations_df)
    
    # Compter les graphes avec violations
    has_type_a = violations_df['type_a_violations'].apply(lambda x: len(x) > 0)
    has_type_b = violations_df['type_b_violations'].apply(lambda x: len(x) > 0)
    has_any_violation = has_type_a | has_type_b
    
    graphs_with_violations = has_any_violation.sum()
    graphs_conforme = total_graphs - graphs_with_violations
    graphs_with_type_a = has_type_a.sum()
    graphs_with_type_b = has_type_b.sum()
    
    # Compter les violations par règle
    type_a_by_rule = {}
    type_b_by_rule = {}
    total_type_a = 0
    total_type_b = 0
    
    for violations_list in violations_df['type_a_violations']:
        for v in violations_list:
            rule_id = v['rule_id']
            type_a_by_rule[rule_id] = type_a_by_rule.get(rule_id, 0) + 1
            total_type_a += 1
    
    for violations_list in violations_df['type_b_violations']:
        for v in violations_list:
            rule_id = v['rule_id']
            type_b_by_rule[rule_id] = type_b_by_rule.get(rule_id, 0) + 1
            total_type_b += 1
    
    # Statistiques finales
    statistics = {
        'total_graphs_analyzed': int(total_graphs),
        'graphs_conforme': int(graphs_conforme),
        'graphs_with_warnings': int(graphs_with_type_b - graphs_with_type_a),  # Seulement Type B
        'graphs_with_critical': int(graphs_with_type_a),  # Au moins une Type A
        'pct_conforme': float(graphs_conforme / total_graphs * 100) if total_graphs > 0 else 0.0,
        'pct_warnings': float((graphs_with_type_b - graphs_with_type_a) / total_graphs * 100) if total_graphs > 0 else 0.0,
        'pct_critical': float(graphs_with_type_a / total_graphs * 100) if total_graphs > 0 else 0.0,
        'total_type_a_violations': int(total_type_a),
        'total_type_b_violations': int(total_type_b),
        'type_a_by_rule': type_a_by_rule,
        'type_b_by_rule': type_b_by_rule,
        'total_unknown_verbs': int(violations_df['unknown_verbs'].apply(len).sum()),
        'total_repetitions': int(violations_df['repetitions'].apply(len).sum())
    }
    
    print(f"Graphes conformes : {statistics['graphs_conforme']:,} ({statistics['pct_conforme']:.2f}%)")
    print(f"Graphes avec warnings (Type B only) : {statistics['graphs_with_warnings']:,} ({statistics['pct_warnings']:.2f}%)")
    print(f"Graphes avec erreurs critiques (Type A) : {statistics['graphs_with_critical']:,} ({statistics['pct_critical']:.2f}%)")
    print(f"\nTotal violations Type A : {statistics['total_type_a_violations']:,}")
    print(f"Total violations Type B : {statistics['total_type_b_violations']:,}")
    
    return statistics


# ==============================================================================
# SECTION 7 : EXPORT DES VIOLATIONS
# ==============================================================================

def export_violations(
    violations_df: pd.DataFrame,
    output_dir: str
):
    """
    Exporte les violations dans des fichiers CSV
    
    Args:
        violations_df: DataFrame des violations
        output_dir: Répertoire de sortie
    """
    print(f"\n{'='*80}")
    print(f"EXPORT DES VIOLATIONS")
    print(f"{'='*80}\n")
    output_dir = DATA_DIR/output_dir
    # Export Type A
    type_a_rows = []
    for _, row in violations_df.iterrows():
        for v in row['type_a_violations']:
            type_a_rows.append({
                'recipe_id': row['recipe_id'],
                'title': row['title'],
                'rule_id': v['rule_id'],
                'description': v['description'],
                'succession': v['succession'],
                'position': v['position']
            })
    
    if type_a_rows:
        type_a_df = pd.DataFrame(type_a_rows)
        type_a_file = os.path.join(output_dir, 'errors', 'type_a_violations.csv')
        type_a_df.to_csv(type_a_file, index=False, encoding='utf-8')
        print(f"✅ Violations Type A exportées : {type_a_file} ({len(type_a_rows):,} violations)")
    
    # Export Type B
    type_b_rows = []
    for _, row in violations_df.iterrows():
        for v in row['type_b_violations']:
            type_b_rows.append({
                'recipe_id': row['recipe_id'],
                'title': row['title'],
                'rule_id': v['rule_id'],
                'description': v['description'],
                'succession': v['succession'],
                'position': v['position']
            })
    
    if type_b_rows:
        type_b_df = pd.DataFrame(type_b_rows)
        type_b_file = os.path.join(output_dir, 'errors', 'type_b_violations.csv')
        type_b_df.to_csv(type_b_file, index=False, encoding='utf-8')
        print(f"✅ Violations Type B exportées : {type_b_file} ({len(type_b_rows):,} violations)")
    
    print(f"\n✅ Export terminé")


def export_semantic_violations_dataset(
    violations_df: pd.DataFrame,
    output_dir: str
) -> str:
    """
    Crée le dataset des violations sémantiques au format requis
    
    Format: id, flag
    Une ligne par violation (une recette peut avoir plusieurs lignes)
    
    Args:
        violations_df: DataFrame des violations
        output_dir: Répertoire de sortie
    
    Returns:
        Chemin du fichier généré
    """
    print(f"\n{'='*80}")
    print(f"EXPORT DATASET VIOLATIONS SÉMANTIQUES")
    print(f"{'='*80}\n")
    
    output_dir = DATA_DIR/output_dir
    dataset_rows = []
    
    # Parcourir toutes les recettes avec violations
    for _, row in violations_df.iterrows():
        recipe_id = row['recipe_id']
        
        # Ajouter toutes les violations Type A
        for v in row['type_a_violations']:
            dataset_rows.append({
                'id': recipe_id,
                'flag': v['rule_id'],
                'violations_type': 'A'
            })
        
        # Ajouter toutes les violations Type B
        for v in row['type_b_violations']:
            dataset_rows.append({
                'id': recipe_id,
                'flag': v['rule_id'],
                'violations_type': 'B'
            })
    
    # Créer le DataFrame
    dataset_df = pd.DataFrame(dataset_rows)
    
    # Trier par id puis par flag
    dataset_df = dataset_df.sort_values(['id', 'flag']).reset_index(drop=True)
    
    # Exporter
    output_file = os.path.join(output_dir, 'dataset_violations_semantiques.csv')
    dataset_df.to_csv(output_file, index=False, encoding='utf-8')
    
    # Statistiques
    total_violations = len(dataset_df)
    unique_recipes = dataset_df['id'].nunique()
    unique_flags = dataset_df['flag'].nunique()
    
    print(f"✅ Dataset des violations sémantiques créé : {output_file}")
    print(f"   - Total violations enregistrées : {total_violations:,}")
    print(f"   - Recettes uniques concernées : {unique_recipes:,}")
    print(f"   - Flags uniques détectés : {unique_flags}")
    print(f"\nAperçu des flags:")
    flag_counts = dataset_df['flag'].value_counts()
    for flag, count in flag_counts.items():
        print(f"   - {flag}: {count:,} occurrences")
    
    return output_file


# ==============================================================================
# SECTION 8 : VISUALISATIONS (3 GRAPHIQUES)
# ==============================================================================

def plot_pie_conforme_warning_critical(
    statistics: Dict,
    output_dir: str
) -> str:
    """
    Graphique 1: Pie plot montrant les proportions CONFORME / WARNING / CRITIQUE
    
    Args:
        statistics: Dictionnaire des statistiques
        output_dir: Répertoire de sortie
    
    Returns:
        Chemin du fichier généré
    """
    print(f"\n{'='*80}")
    print(f"GÉNÉRATION GRAPHIQUE 1 : PIE PLOT CONFORME/WARNING/CRITIQUE")
    print(f"{'='*80}\n")
    
    # Données
    labels = ['Conforme', 'Warning (Type B)', 'Critique (Type A)']
    sizes = [
        statistics['graphs_conforme'],
        statistics['graphs_with_warnings'],
        statistics['graphs_with_critical']
    ]
    colors = ['#2ECC71', '#F39C12', '#E74C3C']
    explode = (0, 0.05, 0.1)  # Mettre en avant les violations
    
    # Créer le graphique
    fig, ax = plt.subplots(figsize=(10, 8))
    
    wedges, texts, autotexts = ax.pie(
        sizes, 
        explode=explode, 
        labels=labels, 
        colors=colors,
        autopct='%1.1f%%',
        startangle=90,
        textprops={'fontsize': 12, 'weight': 'bold'}
    )
    
    # Améliorer l'affichage des pourcentages
    for autotext in autotexts:
        autotext.set_color('white')
        autotext.set_fontsize(14)
        autotext.set_weight('bold')
    
    ax.set_title(
        'Distribution des Graphes par Statut de Validation\nStratégie 3 - Validation Sémantique',
        fontsize=16,
        weight='bold',
        pad=20
    )
    
    # Ajouter une légende avec les counts
    legend_labels = [
        f'{labels[0]}: {sizes[0]:,} graphes ({sizes[0]/sum(sizes)*100:.1f}%)',
        f'{labels[1]}: {sizes[1]:,} graphes ({sizes[1]/sum(sizes)*100:.1f}%)',
        f'{labels[2]}: {sizes[2]:,} graphes ({sizes[2]/sum(sizes)*100:.1f}%)'
    ]
    ax.legend(legend_labels, loc='upper left', bbox_to_anchor=(1, 1), fontsize=10)
    
    plt.tight_layout()
    
    # Sauvegarder
    output_file = os.path.join(output_dir, 'visualizations', 'pie_conforme_warning_critical.png')
    plt.savefig(output_file, dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"✅ Graphique 1 sauvegardé : {output_file}")
    
    return output_file


def plot_bar_type_a_distribution(
    statistics: Dict,
    output_dir: str
) -> str:
    """
    Graphique 2: Bar plot montrant la distribution des règles Type A
    
    Args:
        statistics: Dictionnaire des statistiques
        output_dir: Répertoire de sortie
    
    Returns:
        Chemin du fichier généré
    """
    print(f"\n{'='*80}")
    print(f"GÉNÉRATION GRAPHIQUE 2 : BAR PLOT RÈGLES TYPE A")
    print(f"{'='*80}\n")
    
    # Données
    type_a_by_rule = statistics['type_a_by_rule']
    
    if not type_a_by_rule:
        print("⚠️  Aucune violation Type A détectée, graphique non généré")
        return None
    
    rules = list(type_a_by_rule.keys())
    counts = list(type_a_by_rule.values())
    
    # Créer le graphique
    fig, ax = plt.subplots(figsize=(12, 6))
    
    bars = ax.bar(range(len(rules)), counts, color='#E74C3C', edgecolor='black', linewidth=1.5)
    
    # Étiquettes
    ax.set_xticks(range(len(rules)))
    ax.set_xticklabels(rules, rotation=0, ha='center', fontsize=10)
    ax.set_ylabel('Nombre d\'occurrences', fontsize=12, weight='bold')
    ax.set_title('Distribution des Violations Type A (Critiques)\nStratégie 3', fontsize=14, weight='bold', pad=15)
    ax.grid(axis='y', alpha=0.3, linestyle='--')
    
    # Ajouter les valeurs sur les barres
    for bar, count in zip(bars, counts):
        height = bar.get_height()
        ax.text(
            bar.get_x() + bar.get_width()/2,
            height + max(counts)*0.01,
            f'{count:,}',
            ha='center',
            va='bottom',
            fontsize=11,
            weight='bold'
        )
    
    plt.tight_layout()
    
    # Sauvegarder
    output_file = os.path.join(output_dir, 'visualizations', 'bar_type_a_distribution.png')
    plt.savefig(output_file, dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"✅ Graphique 2 sauvegardé : {output_file}")
    
    return output_file


def plot_bar_type_b_distribution(
    statistics: Dict,
    output_dir: str
) -> str:
    """
    Graphique 3: Bar plot montrant la distribution des règles Type B
    
    Args:
        statistics: Dictionnaire des statistiques
        output_dir: Répertoire de sortie
    
    Returns:
        Chemin du fichier généré
    """
    print(f"\n{'='*80}")
    print(f"GÉNÉRATION GRAPHIQUE 3 : BAR PLOT RÈGLES TYPE B")
    print(f"{'='*80}\n")
    
    # Données
    type_b_by_rule = statistics['type_b_by_rule']
    
    if not type_b_by_rule:
        print("⚠️  Aucune violation Type B détectée, graphique non généré")
        return None
    
    rules = list(type_b_by_rule.keys())
    counts = list(type_b_by_rule.values())
    
    # Créer le graphique
    fig, ax = plt.subplots(figsize=(14, 6))
    
    bars = ax.bar(range(len(rules)), counts, color='#F39C12', edgecolor='black', linewidth=1.5)
    
    # Étiquettes
    ax.set_xticks(range(len(rules)))
    ax.set_xticklabels(rules, rotation=45, ha='right', fontsize=10)
    ax.set_ylabel('Nombre d\'occurrences', fontsize=12, weight='bold')
    ax.set_title('Distribution des Violations Type B (Warnings)\nStratégie 3', fontsize=14, weight='bold', pad=15)
    ax.grid(axis='y', alpha=0.3, linestyle='--')
    
    # Ajouter les valeurs sur les barres
    for bar, count in zip(bars, counts):
        height = bar.get_height()
        ax.text(
            bar.get_x() + bar.get_width()/2,
            height + max(counts)*0.01,
            f'{count:,}',
            ha='center',
            va='bottom',
            fontsize=10,
            weight='bold'
        )
    
    plt.tight_layout()
    
    # Sauvegarder
    output_file = os.path.join(output_dir, 'visualizations', 'bar_type_b_distribution.png')
    plt.savefig(output_file, dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"✅ Graphique 3 sauvegardé : {output_file}")
    
    return output_file


# ==============================================================================
# SECTION 9 : GÉNÉRATION DU RAPPORT MARKDOWN
# ==============================================================================

def generate_semantic_validation_report(
    statistics: Dict,
    output_dir: str,
    viz_files: Dict
) -> str:
    """
    Génère le rapport Markdown complet de la Stratégie 3
    
    Args:
        statistics: Dictionnaire des statistiques
        output_dir: Répertoire de sortie
        viz_files: Dictionnaire des chemins de visualisations
    
    Returns:
        Chemin du rapport généré
    """
    print(f"\n{'='*80}")
    print(f"GÉNÉRATION DU RAPPORT MARKDOWN")
    print(f"{'='*80}\n")
    
    report_lines = []
    
    # En-tête
    report_lines.append("# Rapport de Validation Sémantique - Stratégie 3")
    report_lines.append("")
    report_lines.append(f"**Date de génération:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report_lines.append("")
    report_lines.append("---")
    report_lines.append("")
    
    # Introduction
    report_lines.append("## Introduction")
    report_lines.append("")
    report_lines.append("Ce rapport présente les résultats de la validation sémantique des graphes de gestes culinaires.")
    report_lines.append("La Stratégie 3 détecte les successions illogiques d'actions en utilisant une taxonomie de verbes et des règles de succession.")
    report_lines.append("")
    report_lines.append("### Règles de Validation")
    report_lines.append("")
    report_lines.append("**Règles Type A (Erreurs Critiques) - 3 règles:**")
    report_lines.append("")
    report_lines.append("1. **A1_IRREVERSIBILITE_TEMPORELLE**: FINITION → PRÉPARATION INITIALE")
    report_lines.append("2. **A2_ILLOGISME_FINITION_MELANGE**: FINITION → MÉLANGE")
    report_lines.append("3. **A3_ILLOGISME_FINITION_CUISSON**: FINITION → CUISSON")
    report_lines.append("")
    report_lines.append("**Règles Type B (Warnings) - 5 règles:**")
    report_lines.append("")
    report_lines.append("1. **B1_SAUT_DE_PHASE**: PRÉPARATION → FINITION (sans étapes intermédiaires)")
    report_lines.append("2. **B2_ORDRE_INVERSE**: TRANSFERT → TRANSFORMATION MÉCANIQUE")
    report_lines.append("3. **B3_CUISSON_SANS_PREPARATION**: CUISSON comme première action")
    report_lines.append("4. **B4_MELANGE_APRES_CUISSON_PASSIVE**: CUISSON PASSIVE → MÉLANGE")
    report_lines.append("5. **B5_TRANSFORMATION_APRES_TRANSFERT_FINAL**: TRANSFERT → TRANSFORMATION")
    report_lines.append("")
    report_lines.append("---")
    report_lines.append("")
    
    # Résultats globaux
    report_lines.append("## Résultats Globaux")
    report_lines.append("")
    report_lines.append(f"- **Total graphes analysés:** {statistics['total_graphs_analyzed']:,}")
    report_lines.append(f"- **Graphes conformes:** {statistics['graphs_conforme']:,} ({statistics['pct_conforme']:.2f}%)")
    report_lines.append(f"- **Graphes avec warnings (Type B uniquement):** {statistics['graphs_with_warnings']:,} ({statistics['pct_warnings']:.2f}%)")
    report_lines.append(f"- **Graphes avec erreurs critiques (Type A):** {statistics['graphs_with_critical']:,} ({statistics['pct_critical']:.2f}%)")
    report_lines.append("")
    report_lines.append(f"- **Total violations Type A:** {statistics['total_type_a_violations']:,}")
    report_lines.append(f"- **Total violations Type B:** {statistics['total_type_b_violations']:,}")
    report_lines.append(f"- **Total verbes inconnus:** {statistics['total_unknown_verbs']:,}")
    report_lines.append(f"- **Total répétitions suspectes:** {statistics['total_repetitions']:,}")
    report_lines.append("")
    report_lines.append("---")
    report_lines.append("")
    
    # Visualisations
    report_lines.append("## Visualisations")
    report_lines.append("")
    
    if viz_files.get('pie'):
        report_lines.append("### Distribution Globale (Conforme / Warning / Critique)")
        report_lines.append("")
        report_lines.append("![Distribution Globale](./visualizations/pie_conforme_warning_critical.png)")
        report_lines.append("")
    
    if viz_files.get('bar_a'):
        report_lines.append("### Distribution des Violations Type A (Critiques)")
        report_lines.append("")
        report_lines.append("![Violations Type A](./visualizations/bar_type_a_distribution.png)")
        report_lines.append("")
    
    if viz_files.get('bar_b'):
        report_lines.append("### Distribution des Violations Type B (Warnings)")
        report_lines.append("")
        report_lines.append("![Violations Type B](./visualizations/bar_type_b_distribution.png)")
        report_lines.append("")
    
    report_lines.append("---")
    report_lines.append("")
    
    # Détails des violations Type A
    report_lines.append("## Détails des Violations Type A (Critiques)")
    report_lines.append("")
    
    if statistics['type_a_by_rule']:
        report_lines.append("| Règle | Occurrences |")
        report_lines.append("|-------|-------------|")
        for rule_id, count in sorted(statistics['type_a_by_rule'].items()):
            report_lines.append(f"| {rule_id} | {count:,} |")
        report_lines.append("")
    else:
        report_lines.append("✅ Aucune violation Type A détectée.")
        report_lines.append("")
    
    report_lines.append("---")
    report_lines.append("")
    
    # Détails des violations Type B
    report_lines.append("## Détails des Violations Type B (Warnings)")
    report_lines.append("")
    
    if statistics['type_b_by_rule']:
        report_lines.append("| Règle | Occurrences |")
        report_lines.append("|-------|-------------|")
        for rule_id, count in sorted(statistics['type_b_by_rule'].items()):
            report_lines.append(f"| {rule_id} | {count:,} |")
        report_lines.append("")
    else:
        report_lines.append("✅ Aucune violation Type B détectée.")
        report_lines.append("")
    
    report_lines.append("---")
    report_lines.append("")
    
    # Fichiers générés
    report_lines.append("## Fichiers Générés")
    report_lines.append("")
    report_lines.append("- `dataset_violations_semantiques.csv` - **Dataset principal des violations (id, flag)**")
    report_lines.append("- `errors/type_a_violations.csv` - Violations Type A détaillées")
    report_lines.append("- `errors/type_b_violations.csv` - Violations Type B détaillées")
    report_lines.append("- `visualizations/pie_conforme_warning_critical.png` - Distribution globale")
    report_lines.append("- `visualizations/bar_type_a_distribution.png` - Distribution Type A")
    report_lines.append("- `visualizations/bar_type_b_distribution.png` - Distribution Type B")
    report_lines.append("- `verb_taxonomy.json` - Taxonomie complète des verbes")
    report_lines.append("")
    report_lines.append("---")
    report_lines.append("")
    
    # Conclusion
    report_lines.append("## Conclusion et Recommandations")
    report_lines.append("")
    
    if statistics['pct_critical'] < 5:
        report_lines.append("✅ **Excellente qualité sémantique** (<5% de violations critiques).")
    elif statistics['pct_critical'] < 10:
        report_lines.append("⚠️ **Qualité acceptable** (5-10% de violations critiques).")
    else:
        report_lines.append("❌ **Problèmes significatifs** (>10% de violations critiques).")
    
    report_lines.append("")
    report_lines.append("### Actions Recommandées")
    report_lines.append("")
    report_lines.append("1. **Corriger toutes les violations Type A** (erreurs certaines)")
    report_lines.append("2. **Examiner les violations Type B les plus fréquentes**")
    report_lines.append("3. **Enrichir la taxonomie** avec les verbes inconnus fréquents")
    report_lines.append("4. **Recettes avec violations critiques** → Mise de côté pour retraitement")
    report_lines.append("5. **Recettes conformes** → Passage à la validation finale")
    report_lines.append("")
    report_lines.append("---")
    report_lines.append("")
    report_lines.append(f"*Rapport généré automatiquement le {datetime.now().strftime('%Y-%m-%d à %H:%M:%S')}*")
    report_lines.append("")
    
    # Sauvegarder
    report_text = '\n'.join(report_lines)
    report_file = os.path.join(output_dir, 'rapport_test_validation_semantique.md')
    
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write(report_text)
    
    print(f"✅ Rapport Markdown sauvegardé : {report_file}")
    print(f"   Nombre de sections : {len([l for l in report_lines if l.startswith('##')])}")
    print(f"   Taille du rapport : {len(report_text):,} caractères")
    
    return report_file


# ==============================================================================
# SECTION 10 : PIPELINE PRINCIPAL
# ==============================================================================

def run_strategy_3_pipeline(
    graphs_csv: str,
    critical_flags_csv: str = None,
    output_dir: str = "strategy_3_results"
):
    """
    Exécute le pipeline complet de la Stratégie 3 
    
    Args:
        graphs_csv: Chemin vers le fichier CSV des graphes
        critical_flags_csv: Chemin vers dataset_flags_critique.csv de Stratégie 2 (optionnel)
        output_dir: Répertoire de sortie
    """
    print(f"\n{'#'*80}")
    print(f"# PIPELINE STRATÉGIE 3 - VALIDATION SÉMANTIQUE ")
    print(f"{'#'*80}\n")
    
    # Créer les répertoires de sortie
    create_output_directory(output_dir)
    
    # Étape 1 : Export de la taxonomie
    print("ÉTAPE 1/8 : Export de la taxonomie des verbes...")
    taxonomy_file = os.path.join(output_dir, 'verb_taxonomy.json')
    VerbTaxonomy.export_taxonomy(taxonomy_file)
    
    # Étape 2 : Chargement des données
    print("\nÉTAPE 2/8 : Chargement des données...")
    graphs_df = pd.read_csv(graphs_csv)
    print(f"  ✅ Graphes chargés : {len(graphs_df):,}")
    
    # Étape 2b : Filtrage des recettes critiques (Stratégie 2)
    print("\nÉTAPE 2b/8 : Filtrage des recettes critiques (Stratégie 2)...")
    if critical_flags_csv:
        graphs_df = filter_critical_recipes(graphs_df, critical_flags_csv)
    else:
        print("  ⚠️  Aucun fichier de flags critiques fourni - analyse complète du dataset")
    
    # Étape 3 : Analyse et détection
    print("\nÉTAPE 3/8 : Analyse et détection des violations...")
    violations_df, statistics = analyze_dataset(graphs_df, output_dir)
    
    # Étape 4 : Export des violations détaillées
    print("\nÉTAPE 4/8 : Export des violations détaillées...")
    export_violations(violations_df, output_dir)
    
    # Étape 5 : Export du dataset des violations sémantiques
    print("\nÉTAPE 5/8 : Export du dataset des violations sémantiques...")
    semantic_violations_file = export_semantic_violations_dataset(violations_df, output_dir)
    
    # Étape 6 : Génération des visualisations
    print("\nÉTAPE 6/8 : Génération des visualisations...")
    viz_files = {}
    viz_files['pie'] = plot_pie_conforme_warning_critical(statistics, output_dir)
    viz_files['bar_a'] = plot_bar_type_a_distribution(statistics, output_dir)
    viz_files['bar_b'] = plot_bar_type_b_distribution(statistics, output_dir)
    
    # Étape 7 : Génération du rapport
    print("\nÉTAPE 7/8 : Génération du rapport final...")
    generate_semantic_validation_report(statistics, output_dir, viz_files)
    
    # Étape 8 : Résumé final
    print(f"\n{'#'*80}")
    print(f"# PIPELINE TERMINÉ AVEC SUCCÈS")
    print(f"{'#'*80}\n")
    print(f"📁 Tous les résultats sont dans : {output_dir}/")
    print(f"   - Taxonomie : {output_dir}/verb_taxonomy.json")
    print(f"   - Violations détaillées : {output_dir}/errors/")
    print(f"   - Dataset violations : {DATA_DIR/output_dir}/dataset_violations_semantiques.csv")
    print(f"   - Visualisations : {output_dir}/visualizations/")
    print(f"   - Rapport : {output_dir}/rapport_test_validation_semantique.md")
    print(f"\n✅ Stratégie 3 complétée !\n")


"""
Pipeline Stratégie 1 : Validation par Annotations Manuelles (VERSION MODIFIÉE)
Comparaison tripartite avec 2 variantes (Principale + Ingrédients)

Modifications:
- Paramètre nombre_recettes flexible
- Annotation de 2 variantes (principale + ingrédients)
- Fichier Excel avec 4 colonnes d'annotation (2 anciennes + 2 manuelles)
- Ré-annotation LLM pour les 2 variantes
- Comparaisons tripartites pour chaque variante
- Rapport Markdown avec visualisations
"""

import pandas as pd
import json
import os
from typing import List, Dict, Tuple
from datetime import datetime
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
import seaborn as sns
import ast


# ==============================================================================
# SECTION 1 : ÉCHANTILLONNAGE STRATIFIÉ
# ==============================================================================

def classify_recipe_complexity(row: pd.Series) -> str:
    """
    Classifie la complexité d'une recette selon le nombre d'instructions et d'ingrédients
    
    Args:
        row: Ligne du DataFrame contenant 'number_of_steps' et 'number_of_ingredients'
    
    Returns:
        'simple', 'moyenne', ou 'elevee'
    """
    steps = row['number_of_steps']
    ingredients = row['number_of_ingredients']
    
    if steps <= 5 and ingredients <= 8:
        return 'simple'
    elif steps <= 10 and ingredients <= 15:
        return 'moyenne'
    else:
        return 'elevee'


def detect_pretransformed_ingredients(ingredients_list: List[str]) -> bool:
    """
    Détecte si la liste d'ingrédients contient des ingrédients pré-transformés
    
    Args:
        ingredients_list: Liste des ingrédients de la recette
    
    Returns:
        True si au moins un ingrédient pré-transformé est détecté
    """
    transformation_keywords = [
        'sliced', 'minced', 'crushed', 'ground', 'peeled', 'cored', 'julienned',
        'chopped', 'diced', 'cubed', 'quartered', 'halved', 'cut', 'chunked',
        'grated', 'shredded', 'zested', 'mashed', 'pounded', 'pulverized',
        'smashed', 'crumbled', 'pressed', 'pureed', 'blended'
    ]
    
    ingredients_list = [str(x) for x in ingredients_list if pd.notnull(x)]
    ingredients_text = ' '.join(ingredients_list).lower()
    
    return any(keyword in ingredients_text for keyword in transformation_keywords)


def classify_cuisine_type_2(title: str, instructions: List[str]) -> str:
    """
    Classifie le type de cuisine d'une recette (heuristique simple)
    
    Args:
        title: Titre de la recette
        instructions: Liste des instructions
    
    Returns:
        Type de cuisine détecté
    """
    title_lower = title.lower()
    instructions_text = ' '.join(instructions).lower() if isinstance(instructions, list) else ''
    
    # Mots-clés par catégorie
    patisserie_keywords = ['cake', 'cookie', 'bread', 'pie', 'tart', 'pastry']
    mijote_keywords = ['stew', 'soup', 'braise', 'slow', 'simmer']
    rapide_keywords = ['salad', 'sandwich', 'smoothie', 'wrap', 'bowl']
    
    full_text = title_lower + ' ' + instructions_text
    
    if any(keyword in full_text for keyword in patisserie_keywords):
        return 'patisserie'
    elif any(keyword in full_text for keyword in mijote_keywords):
        return 'plat_mijote'
    elif any(keyword in full_text for keyword in rapide_keywords):
        return 'preparation_rapide'
    else:
        return 'autre'


def stratified_sampling(
    recipes_df: pd.DataFrame,
    recipe_instructions_df: pd.DataFrame,
    recipe_ingredients_df: pd.DataFrame,
    graphs_df: pd.DataFrame,
    nombre_recettes: int = 50,
    random_state: int = 42
) -> pd.DataFrame:
    """
    Effectue un échantillonnage stratifié des recettes
    
    Args:
        recipes_df: DataFrame recipes.csv
        recipe_instructions_df: DataFrame recipe_instructions.csv
        recipe_ingredients_df: DataFrame recipe_ingredients.csv
        graphs_df: DataFrame des graphes avec variantes
        nombre_recettes: Nombre de recettes à échantillonner
        random_state: Seed pour reproductibilité
    
    Returns:
        DataFrame avec les recettes échantillonnées
    """
    print(f"\n{'='*80}")
    print(f"ÉCHANTILLONNAGE STRATIFIÉ - {nombre_recettes} recettes")
    print(f"{'='*80}\n")
    
    # Récupérer les variantes principale et ingrédients pour chaque recette
    graphs_principale = graphs_df[graphs_df['type_2'] == 'variante_principale'][['id', 'actions']].copy()
    graphs_principale.rename(columns={'actions': 'actions_principale'}, inplace=True)
    
    graphs_ingredients = graphs_df[graphs_df['type_2'] == 'variante_ingredients'][['id', 'actions']].copy()
    graphs_ingredients.rename(columns={'actions': 'actions_ingredients'}, inplace=True)
    
    # Joindre les données
    recipes_with_graphs = recipes_df.merge(graphs_principale, on='id', how='inner')
    recipes_with_graphs = recipes_with_graphs.merge(graphs_ingredients, on='id', how='inner')
    
    print(f"Recettes disponibles avec les 2 variantes : {len(recipes_with_graphs):,}")
    
    # Classifier la complexité
    recipes_with_graphs['complexity_category'] = recipes_with_graphs.apply(
        classify_recipe_complexity, axis=1
    )
    
    # Obtenir instructions et ingrédients
    instructions_grouped = recipe_instructions_df.groupby('id')['instruction'].apply(list).reset_index()
    ingredients_grouped = recipe_ingredients_df.groupby('id')['ingredient'].apply(list).reset_index()
    
    recipes_with_graphs = recipes_with_graphs.merge(instructions_grouped, on='id', how='left')
    recipes_with_graphs = recipes_with_graphs.merge(ingredients_grouped, on='id', how='left')
    
    # Détecter ingrédients pré-transformés
    recipes_with_graphs['has_pretransformed_ingredients'] = recipes_with_graphs['ingredient'].apply(
        lambda x: detect_pretransformed_ingredients(x) if isinstance(x, list) else False
    )
    
    # Classifier type de cuisine
    recipes_with_graphs['cuisine_type'] = recipes_with_graphs.apply(
        lambda row: classify_cuisine_type_2(row['title'], row['instruction']) 
        if isinstance(row['instruction'], list) else 'autre',
        axis=1
    )
    
    # Distribution cible par complexité
    complexity_distribution = {
        'simple': 0.30,
        'moyenne': 0.40,
        'elevee': 0.30
    }
    
    # Calculer nombre de recettes par strate
    samples_per_complexity = {
        cat: max(1, int(nombre_recettes * pct)) 
        for cat, pct in complexity_distribution.items()
    }
    
    # Ajuster pour atteindre exactement nombre_recettes
    total_assigned = sum(samples_per_complexity.values())
    if total_assigned != nombre_recettes:
        samples_per_complexity['moyenne'] += (nombre_recettes - total_assigned)
    
    print("\nDistribution cible par complexité :")
    for cat, count in samples_per_complexity.items():
        print(f"  {cat}: {count} recettes ({count/nombre_recettes*100:.1f}%)")
    
    # Échantillonner par complexité
    sampled_recipes = []
    
    for complexity, n_samples in samples_per_complexity.items():
        complexity_recipes = recipes_with_graphs[
            recipes_with_graphs['complexity_category'] == complexity
        ]
        
        if len(complexity_recipes) < n_samples:
            print(f"\n⚠️  Attention: Seulement {len(complexity_recipes)} recettes {complexity} disponibles")
            n_samples = len(complexity_recipes)
        
        sample = complexity_recipes.sample(n=n_samples, random_state=random_state)
        sampled_recipes.append(sample)
    
    # Combiner tous les échantillons
    final_sample = pd.concat(sampled_recipes, ignore_index=True)
    
    print(f"\n✅ Échantillonnage terminé : {len(final_sample)} recettes sélectionnées")
    
    # Afficher statistiques
    print("\nRépartition par complexité :")
    print(final_sample['complexity_category'].value_counts().sort_index())
    
    print("\nRépartition par type de cuisine :")
    print(final_sample['cuisine_type'].value_counts())
    
    print(f"\nRecettes avec ingrédients pré-transformés : {final_sample['has_pretransformed_ingredients'].sum()}")
    
    return final_sample


# ==============================================================================
# SECTION 2 : GÉNÉRATION DU FICHIER EXCEL D'ANNOTATION (2 VARIANTES)
# ==============================================================================

def create_annotation_excel(
    sampled_recipes: pd.DataFrame,
    output_path: str = 'echantillon_annotation_manuelle.xlsx'
) -> str:
    """
    Crée un fichier Excel structuré pour l'annotation manuelle des 2 variantes
    
    Args:
        sampled_recipes: DataFrame des recettes échantillonnées
        output_path: Chemin du fichier Excel de sortie
    
    Returns:
        Chemin du fichier créé
    """
    print(f"\n{'='*80}")
    print(f"CRÉATION DU FICHIER EXCEL D'ANNOTATION")
    print(f"{'='*80}\n")
    
    # Préparer le DataFrame pour l'export
    export_df = sampled_recipes[[
        'id', 'title', 'number_of_steps', 'number_of_ingredients',
        'complexity_category', 'cuisine_type', 'has_pretransformed_ingredients',
        'instruction', 'ingredient', 'actions_principale', 'actions_ingredients'
    ]].copy()
    
    # Convertir listes en strings formatées
    export_df['instruction'] = export_df['instruction'].apply(
        lambda x: '\n'.join([f"{i+1}. {instr}" for i, instr in enumerate(x)]) 
        if isinstance(x, list) else ''
    )
    export_df['ingredient'] = export_df['ingredient'].apply(
        lambda x: '\n'.join([f"- {ing}" for ing in x]) 
        if isinstance(x, list) else ''
    )
    
    # Renommer pour clarté
    export_df.rename(columns={
        'actions_principale': 'actions_llm_ancien_PRINCIPALE',
        'actions_ingredients': 'actions_llm_ancien_INGREDIENTS'
    }, inplace=True)
    
    # Ajouter colonnes vides pour annotations manuelles
    export_df['annotations_manuelles_PRINCIPALE'] = ''
    export_df['annotations_manuelles_INGREDIENTS'] = ''
    
    # Colonnes finales
    final_columns = [
        'id', 'title', 'number_of_steps', 'number_of_ingredients',
        'complexity_category', 'cuisine_type', 'has_pretransformed_ingredients',
        'instruction', 'ingredient',
        'actions_llm_ancien_PRINCIPALE', 'annotations_manuelles_PRINCIPALE',
        'actions_llm_ancien_INGREDIENTS', 'annotations_manuelles_INGREDIENTS'
    ]
    export_df = export_df[final_columns]
    
    # Créer le workbook Excel avec mise en forme
    wb = Workbook()
    ws = wb.active
    ws.title = "Annotations"
    
    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    annotation_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Ajouter les données
    for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Style header
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Highlighting colonnes à remplir
            if r_idx > 1 and c_idx in [11, 13]:  # Colonnes annotations manuelles
                cell.fill = annotation_fill
            
            # Wrap text pour instructions et ingredients
            if c_idx in [8, 9]:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Ajuster largeurs
    column_widths = {
        'A': 15, 'B': 30, 'C': 8, 'D': 8, 'E': 12, 'F': 15, 'G': 10,
        'H': 40, 'I': 30, 'J': 35, 'K': 35, 'L': 35, 'M': 35
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Sauvegarder
    wb.save(output_path)
    
    print(f"✅ Fichier Excel créé : {output_path}")
    print(f"   Nombre de recettes : {len(export_df)}")
    print(f"   Colonnes à remplir manuellement :")
    print(f"      - Colonne K : annotations_manuelles_PRINCIPALE")
    print(f"      - Colonne M : annotations_manuelles_INGREDIENTS")
    
    return output_path


# ==============================================================================
# SECTION 3 : RÉ-ANNOTATION LLM (2 VARIANTES)
# ==============================================================================

def extract_json_from_response(response_text: str):
    """Extrait le JSON d'une réponse LLM"""
    import re
    
    # Nettoyer la réponse
    response_text = response_text.strip()
    
    # Chercher le JSON entre ```json et ```
    json_match = re.search(r'```json\s*(\{.*?\})\s*```', response_text, re.DOTALL)
    if json_match:
        json_str = json_match.group(1)
    else:
        # Chercher juste entre ```
        json_match = re.search(r'```\s*(\{.*?\})\s*```', response_text, re.DOTALL)
        if json_match:
            json_str = json_match.group(1)
        else:
            # Prendre toute la réponse si pas de backticks
            json_str = response_text
    
    # Nettoyer les commentaires
    json_str = re.sub(r'/\*.*?\*/', '', json_str, flags=re.DOTALL)
    json_str = re.sub(r'//.*?$', '', json_str, flags=re.MULTILINE)
    
    try:
        return json.loads(json_str)
    except json.JSONDecodeError as e:
        print(f"Erreur de parsing JSON: {e}")
        print(f"Texte problématique: {json_str[:200]}...")
        return None


def reannotate_with_llm(
    sampled_recipes: pd.DataFrame,
    api_key: str,
    model_name: str = "mistralai/mistral-7b-instruct",
    output_file: str = 'reannotations_llm.csv',
    batch_size: int = 10,
    max_retries: int = 3
) -> pd.DataFrame:
    """
    Ré-annote les recettes avec un nouveau système LLM pour les 2 variantes
    
    Args:
        sampled_recipes: DataFrame des recettes échantillonnées
        api_key: Clé API OpenRouter
        model_name: Nom du modèle OpenRouter à utiliser
        output_file: Fichier de sortie
        batch_size: Nombre de recettes par batch
        max_retries: Nombre maximum de tentatives en cas d'erreur 429
    
    Returns:
        DataFrame avec nouvelles annotations LLM
    """
    import time
    from openai import OpenAI
    
    print(f"\n{'='*80}")
    print(f"RÉ-ANNOTATION LLM DES 2 VARIANTES")
    print(f"{'='*80}\n")
    print(f"Modèle utilisé: {model_name}")
    print(f"Nombre de recettes: {len(sampled_recipes)}")
    print(f"Taille des batches: {batch_size}")
    
    # Préparer le résultat
    reannotated_df = sampled_recipes[['id', 'title', 'instruction', 'ingredient']].copy()
    reannotated_df['actions_llm_nouveau_PRINCIPALE'] = None
    reannotated_df['actions_llm_nouveau_INGREDIENTS'] = None
    
    # Parser les instructions et ingrédients
    def format_instructions(inst_list):
        if isinstance(inst_list, list):
            return " ".join([f"{i+1}. {instr}" for i, instr in enumerate(inst_list)])
        return str(inst_list)
    
    def format_ingredients(ing_list):
        if isinstance(ing_list, list):
            return ", ".join(ing_list)
        return str(ing_list)
    
    # Traiter par batches
    total_batches = (len(sampled_recipes) + batch_size - 1) // batch_size
    
    for batch_idx in range(total_batches):
        start_idx = batch_idx * batch_size
        end_idx = min((batch_idx + 1) * batch_size, len(sampled_recipes))
        batch = sampled_recipes.iloc[start_idx:end_idx]
        
        print(f"\n--- Traitement batch {batch_idx + 1}/{total_batches} (recettes {start_idx+1}-{end_idx}) ---")
        
        # Préparer le prompt pour ce batch
        recipes_data = []
        for idx, row in batch.iterrows():
            instructions_str = format_instructions(row['instruction'])
            ingredients_str = format_ingredients(row['ingredient'])
            
            recipe_info = f"""
RECIPE {len(recipes_data) + 1}:
ID: {row['id']}
TITLE: {row['title']}
COMPLETE INSTRUCTIONS: {instructions_str}
INGREDIENTS: {ingredients_str}
"""
            recipes_data.append(recipe_info)
        
        recipes_text = "\n".join(recipes_data)
        
        prompt = f"""
You are a culinary expert specialized in extracting cooking action sequences.

Here are {len(batch)} recipes to analyze:

{recipes_text}

TASK: GENERATE 2 ACTION VARIANTS FOR EACH RECIPE

For EACH recipe, create exactly 2 variants:

1. **VARIANTE PRINCIPALE**: Main action sequence from instructions
   - Extract ALL action verbs from instructions (both physical gestures AND non-gesture actions)
   - Include physical gestures: add, mix, stir, cut, chop, slice, dice, pour, season, blend, whip, fold, spread, grate, peel, wash, clean, drain, strain, cube, mince, crush, beat, core, seed, flip, turn, toss, combine, knead, roll, press, squeeze, scrape, sprinkle, garnish, arrange, layer, wrap
   - Include non-gesture actions: preheat, cool, refrigerate, boil, simmer, bake, broil, freeze, thaw, chill, warm, rest, set, marinate, steep, proof, rise, cook, fry, sauté, roast, grill, steam, poach, braise, sear, brown, reduce, bring, let
   - Maintain the logical order as they appear in instructions
   - Do NOT invent actions not mentioned

2. **VARIANTE INGRÉDIENTS**: Ingredient-based variant
   - Analyze ingredients for pre-processed items (diced, sliced, chopped, minced, grated, etc.)
   - Extract preparation gestures from ingredient descriptions
   - Pre-processing keywords to detect:
     * cubed → cube
     * diced → dice
     * chopped → chop
     * sliced → slice
     * minced → mince
     * grated → grate
     * shredded → shred
     * crushed → crush
     * beaten → beat
     * peeled → peel
     * cored → core
     * seeded → seed
   - Place ingredient preparation actions AT THE BEGINNING
   - Then include all actions from variante principale (ONLY PHYSICAL GESTURES, exclude non-gestures)
   - If no pre-processed ingredients detected, return the same as principale but filter out non-gestures

EXAMPLES:

Example 1: Fig Jam
- Instructions: "clean figs, take off remainder of the stems. cut in halfs in a large pot add sugar, jello, can of strawberries and juice, and figs, bring to a hard boil, stirring mix well. let it boil hard for 10 minutes, reduce heat to simmer boil, stir constantly for 15 minutes."
- Ingredients: "fresh figs, sugar, strawberry jello, can of strawberries"
- Variante principale: ["clean", "cut", "add", "bring", "boil", "stir", "mix", "reduce", "simmer", "stir"]
- Variante ingrédients: ["clean", "cut", "add", "stir", "mix", "stir"] (no pre-processed ingredients, only gestures)

Example 2: Tomato Salad
- Instructions: "combine all ingredients in a bowl, mix well, season to taste"
- Ingredients: "diced tomatoes, sliced cucumbers, olive oil, salt, pepper"
- Variante principale: ["combine", "mix", "season"]
- Variante ingrédients: ["dice", "slice", "combine", "mix", "season"]

Example 3: Roasted Chicken
- Instructions: "preheat oven to 375°F, season chicken with salt and pepper, place in roasting pan, roast for 1 hour, let rest before serving"
- Ingredients: "whole chicken, salt, pepper, olive oil"
- Variante principale: ["preheat", "season", "place", "roast", "rest"]
- Variante ingrédients: ["season", "place"] (no pre-processed ingredients, only gestures)

RESPONSE FORMAT:
Return ONLY a JSON with this exact structure:
{{
  "recipes": [
    {{
      "id": "recipe_id_1",
      "variante_principale": ["action1", "action2", "action3"],
      "variante_ingredients": ["prep1", "prep2", "action1", "action2"]
    }},
    {{
      "id": "recipe_id_2",
      "variante_principale": ["action1", "action2"],
      "variante_ingredients": ["action1", "action2"]
    }},
    ...
  ]
}}

CRITICAL RULES:
- Process ALL {len(batch)} recipes
- Return ONLY valid JSON with NO comments (no /* */ or //)
- Each recipe MUST have both variants
- Variante principale: ALL actions (gestures + non-gestures)
- Variante ingrédients: Ingredient preps first + ONLY GESTURES from principale
- Do NOT invent actions
- Maintain logical action order
- If uncertain, keep the sequence simple and accurate
"""

        # Appel API avec retry
        for retry in range(max_retries):
            try:
                client = OpenAI(
                    base_url="https://openrouter.ai/api/v1",
                    api_key=api_key,
                )
                
                extra_headers = {
                    "HTTP-Referer": "strategy-1-reannotation",
                    "X-Title": "Strategy 1: Recipe Reannotation"
                }
                
                completion = client.chat.completions.create(
                    extra_headers=extra_headers,
                    model=model_name,
                    messages=[
                        {
                            "role": "system",
                            "content": "You are a culinary expert who extracts action sequences from recipes. You MUST return only valid JSON without any comments."
                        },
                        {
                            "role": "user",
                            "content": prompt
                        }
                    ],
                    max_tokens=4000,
                    temperature=0.1,
                )
                
                llm_response = completion.choices[0].message.content.strip()
                print(f"  Réponse reçue (100 premiers caractères): {llm_response[:100]}...")
                
                # Parser la réponse
                result = extract_json_from_response(llm_response)
                
                if result is None:
                    print(f"  ⚠️  Erreur de parsing JSON pour batch {batch_idx + 1}")
                    if retry < max_retries - 1:
                        print(f"  Nouvelle tentative ({retry + 2}/{max_retries})...")
                        time.sleep(5)
                        continue
                    else:
                        print(f"  ❌ Échec après {max_retries} tentatives")
                        break
                
                # Extraire les résultats
                recipes_results = result.get("recipes", [])
                
                if len(recipes_results) != len(batch):
                    print(f"  ⚠️  Nombre de résultats ({len(recipes_results)}) != nombre de recettes ({len(batch)})")
                
                # Mettre à jour le DataFrame
                for recipe_result in recipes_results:
                    recipe_id = recipe_result.get("id")
                    variante_principale = recipe_result.get("variante_principale", [])
                    variante_ingredients = recipe_result.get("variante_ingredients", [])
                    
                     # Trouver l'index dans le DataFrame
                    mask = reannotated_df['id'] == recipe_id
                    if mask.any():
                        idx = reannotated_df[mask].index[0]  # Obtenir l'index numérique
                        reannotated_df.at[idx, 'actions_llm_nouveau_PRINCIPALE'] = variante_principale  
                        reannotated_df.at[idx, 'actions_llm_nouveau_INGREDIENTS'] = variante_ingredients 
                                                
                print(f"  ✅ Batch {batch_idx + 1}/{total_batches} terminé avec succès")
                break  # Succès, sortir de la boucle de retry
                
            except Exception as e:
                if "429" in str(e):
                    if retry < max_retries - 1:
                        wait_time = (retry + 1) * 3 * 60  # 3, 6, 9 minutes
                        print(f"  ⚠️  Erreur 429, attente de {wait_time//60} minutes (tentative {retry+1}/{max_retries})...")
                        time.sleep(wait_time)
                    else:
                        print(f"  ❌ Erreur 429 persistante après {max_retries} tentatives")
                        break
                else:
                    print(f"  ❌ Erreur: {str(e)}")
                    if retry < max_retries - 1:
                        print(f"  Nouvelle tentative ({retry + 2}/{max_retries})...")
                        time.sleep(10)
                    else:
                        break
        
        # Petite pause entre les batches
        if batch_idx < total_batches - 1:
            time.sleep(2)
    
    # Vérifier les résultats
    principale_filled = reannotated_df['actions_llm_nouveau_PRINCIPALE'].notna().sum()
    ingredients_filled = reannotated_df['actions_llm_nouveau_INGREDIENTS'].notna().sum()
    
    print(f"\n{'='*80}")
    print(f"RÉSUMÉ DE LA RÉ-ANNOTATION")
    print(f"{'='*80}")
    print(f"Total recettes: {len(reannotated_df)}")
    print(f"Variante principale remplie: {principale_filled} / {len(reannotated_df)}")
    print(f"Variante ingrédients remplie: {ingredients_filled} / {len(reannotated_df)}")
    
    if principale_filled < len(reannotated_df) or ingredients_filled < len(reannotated_df):
        print(f"\n⚠️  ATTENTION : Certaines annotations sont manquantes!")
        print(f"   Recettes sans variante principale: {len(reannotated_df) - principale_filled}")
        print(f"   Recettes sans variante ingrédients: {len(reannotated_df) - ingredients_filled}")
    
    # Sauvegarder
    reannotated_df.to_csv(output_file, index=False, encoding='utf-8')
    print(f"\n✅ Fichier de ré-annotation sauvegardé : {output_file}")
    
    return reannotated_df


# ==============================================================================
# SECTION 4 : CHARGEMENT DES ANNOTATIONS MANUELLES
# ==============================================================================

def parse_action_list(actions_str):
    """Parse une liste d'actions depuis string ou liste"""
    if pd.isna(actions_str) or actions_str == '':
        return []
    
    if isinstance(actions_str, list):
        return actions_str
    
    if isinstance(actions_str, str):
        try:
            parsed = ast.literal_eval(actions_str)
            if isinstance(parsed, list):
                return parsed
        except:
            # Si échec, split par virgule
            if ',' in actions_str:
                return [a.strip() for a in actions_str.split(',') if a.strip()]
            return [actions_str.strip()] if actions_str.strip() else []
    
    return []


def load_manual_annotations(excel_file: str) -> pd.DataFrame:
    """
    Charge les annotations manuelles depuis le fichier Excel rempli
    
    Args:
        excel_file: Chemin du fichier Excel avec annotations manuelles
    
    Returns:
        DataFrame avec annotations parsées
    """
    print(f"\n{'='*80}")
    print(f"CHARGEMENT DES ANNOTATIONS MANUELLES")
    print(f"{'='*80}\n")
    
    df = pd.read_excel(excel_file, sheet_name='Annotations')
    
    print(f"Recettes chargées : {len(df)}")
    
    # Parser les colonnes d'actions
    df['actions_llm_ancien_PRINCIPALE_parsed'] = df['actions_llm_ancien_PRINCIPALE'].apply(parse_action_list)
    df['actions_llm_ancien_INGREDIENTS_parsed'] = df['actions_llm_ancien_INGREDIENTS'].apply(parse_action_list)
    df['annotations_manuelles_PRINCIPALE_parsed'] = df['annotations_manuelles_PRINCIPALE'].apply(parse_action_list)
    df['annotations_manuelles_INGREDIENTS_parsed'] = df['annotations_manuelles_INGREDIENTS'].apply(parse_action_list)
    
    # Vérifier que les annotations manuelles sont remplies
    principale_filled = df['annotations_manuelles_PRINCIPALE_parsed'].apply(len).sum()
    ingredients_filled = df['annotations_manuelles_INGREDIENTS_parsed'].apply(len).sum()
    
    print(f"Annotations manuelles PRINCIPALE remplies : {(df['annotations_manuelles_PRINCIPALE_parsed'].apply(len) > 0).sum()} / {len(df)}")
    print(f"Annotations manuelles INGREDIENTS remplies : {(df['annotations_manuelles_INGREDIENTS_parsed'].apply(len) > 0).sum()} / {len(df)}")
    
    if principale_filled == 0 or ingredients_filled == 0:
        print("\n⚠️  ATTENTION : Certaines annotations manuelles sont vides!")
    
    return df


# ==============================================================================
# SECTION 5 : MÉTRIQUES DE COMPARAISON
# ==============================================================================

def calculate_exact_match(seq1: List[str], seq2: List[str]) -> bool:
    """Calcule si deux séquences sont exactement identiques"""
    return seq1 == seq2


def calculate_accuracy(seq1: List[str], seq2: List[str]) -> float:
    """Calcule la précision (proportion de verbes identiques en position identique)"""
    if len(seq1) == 0 and len(seq2) == 0:
        return 1.0
    
    max_len = max(len(seq1), len(seq2))
    if max_len == 0:
        return 1.0
    
    matches = sum(1 for i in range(min(len(seq1), len(seq2))) if seq1[i] == seq2[i])
    return matches / max_len


def calculate_jaccard_similarity(seq1: List[str], seq2: List[str]) -> float:
    """Calcule la similarité de Jaccard (overlap sans ordre)"""
    set1 = set(seq1)
    set2 = set(seq2)
    
    if len(set1) == 0 and len(set2) == 0:
        return 1.0
    
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    
    return intersection / union if union > 0 else 0.0


def calculate_levenshtein_distance(seq1: List[str], seq2: List[str]) -> int:
    """Calcule la distance de Levenshtein (distance d'édition)"""
    if len(seq1) == 0:
        return len(seq2)
    if len(seq2) == 0:
        return len(seq1)
    
    dp = [[0] * (len(seq2) + 1) for _ in range(len(seq1) + 1)]
    
    for i in range(len(seq1) + 1):
        dp[i][0] = i
    for j in range(len(seq2) + 1):
        dp[0][j] = j
    
    for i in range(1, len(seq1) + 1):
        for j in range(1, len(seq2) + 1):
            if seq1[i-1] == seq2[j-1]:
                dp[i][j] = dp[i-1][j-1]
            else:
                dp[i][j] = 1 + min(dp[i-1][j], dp[i][j-1], dp[i-1][j-1])
    
    return dp[len(seq1)][len(seq2)]


def calculate_lcs_ratio(seq1: List[str], seq2: List[str]) -> float:
    """Calcule le ratio de la plus longue sous-séquence commune (LCS)"""
    if len(seq1) == 0 or len(seq2) == 0:
        return 0.0
    
    dp = [[0] * (len(seq2) + 1) for _ in range(len(seq1) + 1)]
    
    for i in range(1, len(seq1) + 1):
        for j in range(1, len(seq2) + 1):
            if seq1[i-1] == seq2[j-1]:
                dp[i][j] = dp[i-1][j-1] + 1
            else:
                dp[i][j] = max(dp[i-1][j], dp[i][j-1])
    
    lcs_length = dp[len(seq1)][len(seq2)]
    max_len = max(len(seq1), len(seq2))
    
    return lcs_length / max_len if max_len > 0 else 0.0


def compare_sequences(
    seq_manual: List[str],
    seq_ancien: List[str],
    seq_nouveau: List[str]
) -> Dict:
    """
    Compare trois séquences et calcule toutes les métriques
    
    Args:
        seq_manual: Séquence annotée manuellement
        seq_ancien: Séquence LLM ancien
        seq_nouveau: Séquence LLM nouveau
    
    Returns:
        Dictionnaire contenant toutes les métriques
    """
    metrics = {}
    
    # Comparaison 1: Manuel vs Ancien
    metrics['manuel_vs_ancien'] = {
        'exact_match': calculate_exact_match(seq_manual, seq_ancien),
        'accuracy': calculate_accuracy(seq_manual, seq_ancien),
        'jaccard': calculate_jaccard_similarity(seq_manual, seq_ancien),
        'levenshtein': calculate_levenshtein_distance(seq_manual, seq_ancien),
        'lcs_ratio': calculate_lcs_ratio(seq_manual, seq_ancien),
        'length_diff': abs(len(seq_manual) - len(seq_ancien))
    }
    
    # Comparaison 2: Manuel vs Nouveau
    metrics['manuel_vs_nouveau'] = {
        'exact_match': calculate_exact_match(seq_manual, seq_nouveau),
        'accuracy': calculate_accuracy(seq_manual, seq_nouveau),
        'jaccard': calculate_jaccard_similarity(seq_manual, seq_nouveau),
        'levenshtein': calculate_levenshtein_distance(seq_manual, seq_nouveau),
        'lcs_ratio': calculate_lcs_ratio(seq_manual, seq_nouveau),
        'length_diff': abs(len(seq_manual) - len(seq_nouveau))
    }
    
    # Comparaison 3: Ancien vs Nouveau
    metrics['ancien_vs_nouveau'] = {
        'exact_match': calculate_exact_match(seq_ancien, seq_nouveau),
        'accuracy': calculate_accuracy(seq_ancien, seq_nouveau),
        'jaccard': calculate_jaccard_similarity(seq_ancien, seq_nouveau),
        'levenshtein': calculate_levenshtein_distance(seq_ancien, seq_nouveau),
        'lcs_ratio': calculate_lcs_ratio(seq_ancien, seq_nouveau),
        'length_diff': abs(len(seq_ancien) - len(seq_nouveau))
    }
    
    # Catégorisation des changements
    ancien_correct = calculate_exact_match(seq_ancien, seq_manual)
    nouveau_correct = calculate_exact_match(seq_nouveau, seq_manual)
    ancien_vs_nouveau_same = calculate_exact_match(seq_ancien, seq_nouveau)
    
    if ancien_correct and nouveau_correct:
        change_category = 'stable_correct'
    elif not ancien_correct and nouveau_correct:
        change_category = 'correction'
    elif ancien_correct and not nouveau_correct:
        change_category = 'regression'
    elif not ancien_correct and not nouveau_correct:
        if ancien_vs_nouveau_same:
            change_category = 'stable_incorrect'
        else:
            change_category = 'changement_lateral'
    else:
        change_category = 'unknown'
    
    metrics['change_category'] = change_category
    
    return metrics


# ==============================================================================
# SECTION 6 : COMPARAISON TRIPARTITE COMPLÈTE (2 VARIANTES)
# ==============================================================================

def perform_tripartite_comparison(
    manual_annotations_df: pd.DataFrame,
    reannotated_df: pd.DataFrame,
    output_dir: str
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Effectue la comparaison tripartite pour les 2 variantes
    
    Args:
        manual_annotations_df: DataFrame avec annotations manuelles
        reannotated_df: DataFrame avec ré-annotations LLM
        output_dir: Répertoire de sortie
    
    Returns:
        Tuple (metrics_principale_df, metrics_ingredients_df)
    """
    print(f"\n{'='*80}")
    print(f"COMPARAISON TRIPARTITE - 2 VARIANTES")
    print(f"{'='*80}\n")
    
    # Fusionner les DataFrames
    comparison_df = manual_annotations_df.merge(
        reannotated_df[['id', 'actions_llm_nouveau_PRINCIPALE', 'actions_llm_nouveau_INGREDIENTS']],
        on='id',
        how='inner'
    )
    
    print(f"Recettes à comparer : {len(comparison_df)}")
    
    # Parser les nouvelles annotations LLM
    comparison_df['actions_llm_nouveau_PRINCIPALE_parsed'] = comparison_df['actions_llm_nouveau_PRINCIPALE'].apply(parse_action_list)
    comparison_df['actions_llm_nouveau_INGREDIENTS_parsed'] = comparison_df['actions_llm_nouveau_INGREDIENTS'].apply(parse_action_list)
    
    # ========== VARIANTE PRINCIPALE ==========
    print("\n--- Comparaison VARIANTE PRINCIPALE ---")
    
    metrics_principale = []
    
    for idx, row in comparison_df.iterrows():
        seq_manual = row['annotations_manuelles_PRINCIPALE_parsed']
        seq_ancien = row['actions_llm_ancien_PRINCIPALE_parsed']
        seq_nouveau = row['actions_llm_nouveau_PRINCIPALE_parsed']
        
        if len(seq_manual) == 0:
            continue
        
        metrics = compare_sequences(seq_manual, seq_ancien, seq_nouveau)
        
        metrics_principale.append({
            'id': row['id'],
            'title': row['title'],
            'complexity': row.get('complexity_category', 'unknown'),
            'cuisine_type': row.get('cuisine_type', 'unknown'),
            'variante': 'principale',
            
            'len_manual': len(seq_manual),
            'len_ancien': len(seq_ancien),
            'len_nouveau': len(seq_nouveau),
            
            'man_anc_exact_match': metrics['manuel_vs_ancien']['exact_match'],
            'man_anc_accuracy': metrics['manuel_vs_ancien']['accuracy'],
            'man_anc_jaccard': metrics['manuel_vs_ancien']['jaccard'],
            'man_anc_levenshtein': metrics['manuel_vs_ancien']['levenshtein'],
            'man_anc_lcs_ratio': metrics['manuel_vs_ancien']['lcs_ratio'],
            
            'man_nouv_exact_match': metrics['manuel_vs_nouveau']['exact_match'],
            'man_nouv_accuracy': metrics['manuel_vs_nouveau']['accuracy'],
            'man_nouv_jaccard': metrics['manuel_vs_nouveau']['jaccard'],
            'man_nouv_levenshtein': metrics['manuel_vs_nouveau']['levenshtein'],
            'man_nouv_lcs_ratio': metrics['manuel_vs_nouveau']['lcs_ratio'],
            
            'anc_nouv_exact_match': metrics['ancien_vs_nouveau']['exact_match'],
            'anc_nouv_accuracy': metrics['ancien_vs_nouveau']['accuracy'],
            'anc_nouv_jaccard': metrics['ancien_vs_nouveau']['jaccard'],
            'anc_nouv_levenshtein': metrics['ancien_vs_nouveau']['levenshtein'],
            'anc_nouv_lcs_ratio': metrics['ancien_vs_nouveau']['lcs_ratio'],
            
            'change_category': metrics['change_category']
        })
    
    metrics_principale_df = pd.DataFrame(metrics_principale)
    
    # ========== VARIANTE INGRÉDIENTS ==========
    print("--- Comparaison VARIANTE INGRÉDIENTS ---")
    
    metrics_ingredients = []
    
    for idx, row in comparison_df.iterrows():
        seq_manual = row['annotations_manuelles_INGREDIENTS_parsed']
        seq_ancien = row['actions_llm_ancien_INGREDIENTS_parsed']
        seq_nouveau = row['actions_llm_nouveau_INGREDIENTS_parsed']
        
        if len(seq_manual) == 0:
            continue
        
        metrics = compare_sequences(seq_manual, seq_ancien, seq_nouveau)
        
        metrics_ingredients.append({
            'id': row['id'],
            'title': row['title'],
            'complexity': row.get('complexity_category', 'unknown'),
            'cuisine_type': row.get('cuisine_type', 'unknown'),
            'variante': 'ingredients',
            
            'len_manual': len(seq_manual),
            'len_ancien': len(seq_ancien),
            'len_nouveau': len(seq_nouveau),
            
            'man_anc_exact_match': metrics['manuel_vs_ancien']['exact_match'],
            'man_anc_accuracy': metrics['manuel_vs_ancien']['accuracy'],
            'man_anc_jaccard': metrics['manuel_vs_ancien']['jaccard'],
            'man_anc_levenshtein': metrics['manuel_vs_ancien']['levenshtein'],
            'man_anc_lcs_ratio': metrics['manuel_vs_ancien']['lcs_ratio'],
            
            'man_nouv_exact_match': metrics['manuel_vs_nouveau']['exact_match'],
            'man_nouv_accuracy': metrics['manuel_vs_nouveau']['accuracy'],
            'man_nouv_jaccard': metrics['manuel_vs_nouveau']['jaccard'],
            'man_nouv_levenshtein': metrics['manuel_vs_nouveau']['levenshtein'],
            'man_nouv_lcs_ratio': metrics['manuel_vs_nouveau']['lcs_ratio'],
            
            'anc_nouv_exact_match': metrics['ancien_vs_nouveau']['exact_match'],
            'anc_nouv_accuracy': metrics['ancien_vs_nouveau']['accuracy'],
            'anc_nouv_jaccard': metrics['ancien_vs_nouveau']['jaccard'],
            'anc_nouv_levenshtein': metrics['ancien_vs_nouveau']['levenshtein'],
            'anc_nouv_lcs_ratio': metrics['ancien_vs_nouveau']['lcs_ratio'],
            
            'change_category': metrics['change_category']
        })
    
    metrics_ingredients_df = pd.DataFrame(metrics_ingredients)
    
    # Sauvegarder
    os.makedirs(output_dir, exist_ok=True)
    
    file_principale = os.path.join(output_dir, 'comparison_metrics_principale.csv')
    metrics_principale_df.to_csv(file_principale, index=False, encoding='utf-8')
    print(f"\n✅ Métriques PRINCIPALE sauvegardées : {file_principale}")
    
    file_ingredients = os.path.join(output_dir, 'comparison_metrics_ingredients.csv')
    metrics_ingredients_df.to_csv(file_ingredients, index=False, encoding='utf-8')
    print(f"✅ Métriques INGRÉDIENTS sauvegardées : {file_ingredients}")
    
    return metrics_principale_df, metrics_ingredients_df


# ==============================================================================
# SECTION 7 : VISUALISATIONS
# ==============================================================================

def plot_comparison_summary(
    metrics_principale_df: pd.DataFrame,
    metrics_ingredients_df: pd.DataFrame,
    output_dir: str
) -> str:
    """
    Crée le graphique résumé des comparaisons
    
    Args:
        metrics_principale_df: Métriques variante principale
        metrics_ingredients_df: Métriques variante ingrédients
        output_dir: Répertoire de sortie
    
    Returns:
        Chemin du fichier généré
    """
    print(f"\n{'='*80}")
    print(f"GÉNÉRATION DU GRAPHIQUE RÉSUMÉ")
    print(f"{'='*80}\n")
    
    # Préparer les données
    data = []
    
    # Variante Principale
    if len(metrics_principale_df) > 0:
        data.append({
            'Variante': 'Principale',
            'Comparaison': 'Manuel vs Ancien',
            'Accuracy': metrics_principale_df['man_anc_accuracy'].mean()
        })
        data.append({
            'Variante': 'Principale',
            'Comparaison': 'Manuel vs Nouveau',
            'Accuracy': metrics_principale_df['man_nouv_accuracy'].mean()
        })
        data.append({
            'Variante': 'Principale',
            'Comparaison': 'Ancien vs Nouveau',
            'Accuracy': metrics_principale_df['anc_nouv_accuracy'].mean()
        })
    
    # Variante Ingrédients
    if len(metrics_ingredients_df) > 0:
        data.append({
            'Variante': 'Ingrédients',
            'Comparaison': 'Manuel vs Ancien',
            'Accuracy': metrics_ingredients_df['man_anc_accuracy'].mean()
        })
        data.append({
            'Variante': 'Ingrédients',
            'Comparaison': 'Manuel vs Nouveau',
            'Accuracy': metrics_ingredients_df['man_nouv_accuracy'].mean()
        })
        data.append({
            'Variante': 'Ingrédients',
            'Comparaison': 'Ancien vs Nouveau',
            'Accuracy': metrics_ingredients_df['anc_nouv_accuracy'].mean()
        })
    
    df_plot = pd.DataFrame(data)
    
    # Créer le graphique
    fig, ax = plt.subplots(figsize=(12, 6))
    
    # Bar plot groupé
    x = np.arange(len(df_plot['Comparaison'].unique()))
    width = 0.35
    
    principale_data = df_plot[df_plot['Variante'] == 'Principale']['Accuracy'].values
    ingredients_data = df_plot[df_plot['Variante'] == 'Ingrédients']['Accuracy'].values
    
    bars1 = ax.bar(x - width/2, principale_data, width, label='Variante Principale', color='#4472C4', edgecolor='black', linewidth=1.5)
    bars2 = ax.bar(x + width/2, ingredients_data, width, label='Variante Ingrédients', color='#ED7D31', edgecolor='black', linewidth=1.5)
    
    # Étiquettes
    ax.set_xlabel('Type de Comparaison', fontsize=12, weight='bold')
    ax.set_ylabel('Accuracy Moyenne', fontsize=12, weight='bold')
    ax.set_title('Comparaison des Performances - 2 Variantes\nStratégie 1 - Validation par Annotations Manuelles', fontsize=14, weight='bold', pad=20)
    ax.set_xticks(x)
    ax.set_xticklabels(df_plot['Comparaison'].unique(), fontsize=10)
    ax.set_ylim(0, 1.1)
    ax.legend(fontsize=11, loc='upper right')
    ax.grid(axis='y', alpha=0.3, linestyle='--')
    
    # Ajouter valeurs sur barres
    for bars in [bars1, bars2]:
        for bar in bars:
            height = bar.get_height()
            ax.text(
                bar.get_x() + bar.get_width()/2,
                height + 0.02,
                f'{height:.3f}',
                ha='center',
                va='bottom',
                fontsize=9,
                weight='bold'
            )
    
    plt.tight_layout()
    
    # Sauvegarder
    os.makedirs(os.path.join(output_dir, 'visualizations'), exist_ok=True)
    output_file = os.path.join(output_dir, 'visualizations', 'comparison_summary.png')
    plt.savefig(output_file, dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"✅ Graphique sauvegardé : {output_file}")
    
    return output_file


# ==============================================================================
# SECTION 8 : GÉNÉRATION DU RAPPORT MARKDOWN
# ==============================================================================

def generate_validation_report_strat_1(
    metrics_principale_df: pd.DataFrame,
    metrics_ingredients_df: pd.DataFrame,
    output_dir: str,
    viz_file: str
) -> str:
    """
    Génère le rapport Markdown complet
    
    Args:
        metrics_principale_df: Métriques variante principale
        metrics_ingredients_df: Métriques variante ingrédients
        output_dir: Répertoire de sortie
        viz_file: Chemin du graphique
    
    Returns:
        Chemin du rapport généré
    """
    print(f"\n{'='*80}")
    print(f"GÉNÉRATION DU RAPPORT MARKDOWN")
    print(f"{'='*80}\n")
    
    report_lines = []
    
    # En-tête
    report_lines.append("# Rapport de Validation par Annotations Manuelles - Stratégie 1")
    report_lines.append("")
    report_lines.append(f"**Date de génération:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report_lines.append(f"**Nombre de recettes analysées:** {len(metrics_principale_df)}")
    report_lines.append("")
    report_lines.append("---")
    report_lines.append("")
    
    # Introduction
    report_lines.append("## Introduction")
    report_lines.append("")
    report_lines.append("Ce rapport présente les résultats de la validation de la qualité des annotations LLM")
    report_lines.append("par comparaison avec des annotations manuelles (référence gold standard).")
    report_lines.append("")
    report_lines.append("**Comparaisons effectuées:**")
    report_lines.append("- **Manuel vs Ancien**: Performance du système LLM initial")
    report_lines.append("- **Manuel vs Nouveau**: Performance du système LLM amélioré")
    report_lines.append("- **Ancien vs Nouveau**: Évolution entre les deux systèmes")
    report_lines.append("")
    report_lines.append("**Variantes analysées:**")
    report_lines.append("- Variante Principale")
    report_lines.append("- Variante Ingrédients")
    report_lines.append("")
    report_lines.append("---")
    report_lines.append("")
    
    # Visualisation
    report_lines.append("## Visualisation Globale")
    report_lines.append("")
    report_lines.append("![Graphique Résumé](./visualizations/comparison_summary.png)")
    report_lines.append("")
    report_lines.append("---")
    report_lines.append("")
    
    # Métriques PRINCIPALE
    if len(metrics_principale_df) > 0:
        report_lines.append("## Métriques - Variante Principale")
        report_lines.append("")
        
        report_lines.append("### Métriques Moyennes")
        report_lines.append("")
        report_lines.append("| Comparaison | Exact Match | Accuracy | Jaccard | Levenshtein | LCS Ratio |")
        report_lines.append("|-------------|-------------|----------|---------|-------------|-----------|")
        
        # Manuel vs Ancien
        exact_anc = (metrics_principale_df['man_anc_exact_match'].sum() / len(metrics_principale_df) * 100)
        acc_anc = metrics_principale_df['man_anc_accuracy'].mean()
        jac_anc = metrics_principale_df['man_anc_jaccard'].mean()
        lev_anc = metrics_principale_df['man_anc_levenshtein'].mean()
        lcs_anc = metrics_principale_df['man_anc_lcs_ratio'].mean()
        
        report_lines.append(f"| Manuel vs Ancien | {exact_anc:.1f}% | {acc_anc:.3f} | {jac_anc:.3f} | {lev_anc:.1f} | {lcs_anc:.3f} |")
        
        # Manuel vs Nouveau
        exact_nouv = (metrics_principale_df['man_nouv_exact_match'].sum() / len(metrics_principale_df) * 100)
        acc_nouv = metrics_principale_df['man_nouv_accuracy'].mean()
        jac_nouv = metrics_principale_df['man_nouv_jaccard'].mean()
        lev_nouv = metrics_principale_df['man_nouv_levenshtein'].mean()
        lcs_nouv = metrics_principale_df['man_nouv_lcs_ratio'].mean()
        
        report_lines.append(f"| Manuel vs Nouveau | {exact_nouv:.1f}% | {acc_nouv:.3f} | {jac_nouv:.3f} | {lev_nouv:.1f} | {lcs_nouv:.3f} |")
        
        # Ancien vs Nouveau
        exact_comp = (metrics_principale_df['anc_nouv_exact_match'].sum() / len(metrics_principale_df) * 100)
        acc_comp = metrics_principale_df['anc_nouv_accuracy'].mean()
        jac_comp = metrics_principale_df['anc_nouv_jaccard'].mean()
        lev_comp = metrics_principale_df['anc_nouv_levenshtein'].mean()
        lcs_comp = metrics_principale_df['anc_nouv_lcs_ratio'].mean()
        
        report_lines.append(f"| Ancien vs Nouveau | {exact_comp:.1f}% | {acc_comp:.3f} | {jac_comp:.3f} | {lev_comp:.1f} | {lcs_comp:.3f} |")
        report_lines.append("")
        
        # Distribution changements
        report_lines.append("### Distribution des Changements (Ancien → Nouveau)")
        report_lines.append("")
        
        change_counts = metrics_principale_df['change_category'].value_counts()
        total = len(metrics_principale_df)
        
        report_lines.append("| Catégorie | Count | Pourcentage |")
        report_lines.append("|-----------|-------|-------------|")
        
        for cat in ['correction', 'regression', 'stable_correct', 'stable_incorrect', 'changement_lateral']:
            count = change_counts.get(cat, 0)
            pct = (count / total * 100) if total > 0 else 0
            report_lines.append(f"| {cat} | {count} | {pct:.1f}% |")
        
        report_lines.append("")
        report_lines.append("---")
        report_lines.append("")
    
    # Métriques INGRÉDIENTS
    if len(metrics_ingredients_df) > 0:
        report_lines.append("## Métriques - Variante Ingrédients")
        report_lines.append("")
        
        report_lines.append("### Métriques Moyennes")
        report_lines.append("")
        report_lines.append("| Comparaison | Exact Match | Accuracy | Jaccard | Levenshtein | LCS Ratio |")
        report_lines.append("|-------------|-------------|----------|---------|-------------|-----------|")
        
        # Manuel vs Ancien
        exact_anc = (metrics_ingredients_df['man_anc_exact_match'].sum() / len(metrics_ingredients_df) * 100)
        acc_anc = metrics_ingredients_df['man_anc_accuracy'].mean()
        jac_anc = metrics_ingredients_df['man_anc_jaccard'].mean()
        lev_anc = metrics_ingredients_df['man_anc_levenshtein'].mean()
        lcs_anc = metrics_ingredients_df['man_anc_lcs_ratio'].mean()
        
        report_lines.append(f"| Manuel vs Ancien | {exact_anc:.1f}% | {acc_anc:.3f} | {jac_anc:.3f} | {lev_anc:.1f} | {lcs_anc:.3f} |")
        
        # Manuel vs Nouveau
        exact_nouv = (metrics_ingredients_df['man_nouv_exact_match'].sum() / len(metrics_ingredients_df) * 100)
        acc_nouv = metrics_ingredients_df['man_nouv_accuracy'].mean()
        jac_nouv = metrics_ingredients_df['man_nouv_jaccard'].mean()
        lev_nouv = metrics_ingredients_df['man_nouv_levenshtein'].mean()
        lcs_nouv = metrics_ingredients_df['man_nouv_lcs_ratio'].mean()
        
        report_lines.append(f"| Manuel vs Nouveau | {exact_nouv:.1f}% | {acc_nouv:.3f} | {jac_nouv:.3f} | {lev_nouv:.1f} | {lcs_nouv:.3f} |")
        
        # Ancien vs Nouveau
        exact_comp = (metrics_ingredients_df['anc_nouv_exact_match'].sum() / len(metrics_ingredients_df) * 100)
        acc_comp = metrics_ingredients_df['anc_nouv_accuracy'].mean()
        jac_comp = metrics_ingredients_df['anc_nouv_jaccard'].mean()
        lev_comp = metrics_ingredients_df['anc_nouv_levenshtein'].mean()
        lcs_comp = metrics_ingredients_df['anc_nouv_lcs_ratio'].mean()
        
        report_lines.append(f"| Ancien vs Nouveau | {exact_comp:.1f}% | {acc_comp:.3f} | {jac_comp:.3f} | {lev_comp:.1f} | {lcs_comp:.3f} |")
        report_lines.append("")
        
        # Distribution changements
        report_lines.append("### Distribution des Changements (Ancien → Nouveau)")
        report_lines.append("")
        
        change_counts = metrics_ingredients_df['change_category'].value_counts()
        total = len(metrics_ingredients_df)
        
        report_lines.append("| Catégorie | Count | Pourcentage |")
        report_lines.append("|-----------|-------|-------------|")
        
        for cat in ['correction', 'regression', 'stable_correct', 'stable_incorrect', 'changement_lateral']:
            count = change_counts.get(cat, 0)
            pct = (count / total * 100) if total > 0 else 0
            report_lines.append(f"| {cat} | {count} | {pct:.1f}% |")
        
        report_lines.append("")
        report_lines.append("---")
        report_lines.append("")
    
    # Fichiers générés
    report_lines.append("## Fichiers Générés")
    report_lines.append("")
    report_lines.append("- `echantillon_annotation_manuelle.xlsx` - Fichier Excel pour annotations manuelles")
    report_lines.append("- `comparison_metrics_principale.csv` - Métriques détaillées variante principale")
    report_lines.append("- `comparison_metrics_ingredients.csv` - Métriques détaillées variante ingrédients")
    report_lines.append("- `visualizations/comparison_summary.png` - Graphique résumé des comparaisons")
    report_lines.append("- `rapport_validation_annotations.md` - Ce rapport")
    report_lines.append("")
    report_lines.append("---")
    report_lines.append("")
    
    # Conclusion
    report_lines.append("## Conclusion")
    report_lines.append("")
    
    if len(metrics_principale_df) > 0:
        acc_diff_p = metrics_principale_df['man_nouv_accuracy'].mean() - metrics_principale_df['man_anc_accuracy'].mean()
        if acc_diff_p > 0.05:
            report_lines.append(f"✅ **Variante Principale**: Amélioration significative (+{acc_diff_p:.3f} accuracy)")
        elif acc_diff_p < -0.05:
            report_lines.append(f"⚠️ **Variante Principale**: Régression détectée ({acc_diff_p:.3f} accuracy)")
        else:
            report_lines.append(f"🔵 **Variante Principale**: Performance stable ({acc_diff_p:+.3f} accuracy)")
    
    if len(metrics_ingredients_df) > 0:
        acc_diff_i = metrics_ingredients_df['man_nouv_accuracy'].mean() - metrics_ingredients_df['man_anc_accuracy'].mean()
        if acc_diff_i > 0.05:
            report_lines.append(f"✅ **Variante Ingrédients**: Amélioration significative (+{acc_diff_i:.3f} accuracy)")
        elif acc_diff_i < -0.05:
            report_lines.append(f"⚠️ **Variante Ingrédients**: Régression détectée ({acc_diff_i:.3f} accuracy)")
        else:
            report_lines.append(f"🔵 **Variante Ingrédients**: Performance stable ({acc_diff_i:+.3f} accuracy)")
    
    report_lines.append("")
    report_lines.append("---")
    report_lines.append("")
    report_lines.append(f"*Rapport généré automatiquement le {datetime.now().strftime('%Y-%m-%d à %H:%M:%S')}*")
    report_lines.append("")
    
    # Sauvegarder
    report_text = '\n'.join(report_lines)
    report_file = os.path.join(output_dir, 'rapport_validation_annotations.md')
    
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write(report_text)
    
    print(f"✅ Rapport Markdown sauvegardé : {report_file}")
    print(f"   Taille : {len(report_text):,} caractères")
    
    return report_file


# ==============================================================================
# SECTION 9 : PIPELINE PRINCIPAL
# ==============================================================================

def run_strategy_1_pipeline(
    recipes_csv: str = None,
    recipe_instructions_csv: str = None,
    recipe_ingredients_csv: str = None,
    graphs_csv: str = None,
    nombre_recettes: int = 50,
    random_state: int = None ,
    output_dir: str = "strategy_1_results",
    openrouter_api_key: str = None,
    model_name: str = "mistralai/mistral-7b-instruct",
    manual_annotations_excel: str = None,
    reannotations_llm_csv: str = None,
    critical_flags_csv: str = None,
):
    """
    Exécute le pipeline complet de la Stratégie 1 (VERSION MODIFIÉE)
    
    2 modes d'exécution:
    - MODE COMPLET: Si manual_annotations_excel=None → Échantillonnage + Excel + Ré-annotation
    - MODE COMPARAISON: Si manual_annotations_excel fourni → Skip au chargement et comparaison
    
    Args:
        recipes_csv: Chemin vers recipes.csv (requis en mode complet)
        recipe_instructions_csv: Chemin vers recipe_instructions.csv (requis en mode complet)
        recipe_ingredients_csv: Chemin vers recipe_ingredients.csv (requis en mode complet)
        graphs_csv: Chemin vers le fichier des graphes (requis en mode complet)
        nombre_recettes: Nombre de recettes à échantillonner (mode complet uniquement)
        output_dir: Répertoire de sortie
        openrouter_api_key: Clé API OpenRouter pour ré-annotation LLM
        model_name: Nom du modèle OpenRouter à utiliser
        manual_annotations_excel: Chemin du fichier Excel rempli (active mode comparaison)
        reannotations_llm_csv: Chemin du CSV de ré-annotations LLM (optionnel en mode comparaison)
    """
    print(f"\n{'#'*80}")
    print(f"# PIPELINE STRATÉGIE 1 - VALIDATION PAR ANNOTATIONS MANUELLES")
    
    
    os.makedirs(output_dir, exist_ok=True)
    
    # ========== MODE COMPARAISON : Annotations déjà faites ==========
    if manual_annotations_excel is not None:
    
        print(f"# PARTIE 2:  COMPARAISON TRIPARTITE")
        print(f"{'#'*80}\n")
        
        # Étape 1 : Chargement des annotations manuelles
        print("ÉTAPE 1/3 : Chargement des annotations manuelles...")
        manual_df = load_manual_annotations(manual_annotations_excel)
        
        # Étape 2 : Chargement des ré-annotations LLM
        print("\nÉTAPE 2/3 : Chargement des ré-annotations LLM...")
        
        if reannotations_llm_csv and os.path.exists(reannotations_llm_csv):
            print(f"  Chargement depuis : {reannotations_llm_csv}")
            reannotated_df = pd.read_csv(reannotations_llm_csv)
        else:
            # Chercher dans output_dir
            default_reannotation_file = os.path.join(output_dir, 'reannotations_llm.csv')
            if os.path.exists(default_reannotation_file):
                print(f"  Chargement depuis : {default_reannotation_file}")
                reannotated_df = pd.read_csv(default_reannotation_file)
            else:
                print(f"  ❌ Fichier de ré-annotations non trouvé!")
                print(f"     Chemins cherchés:")
                if reannotations_llm_csv:
                    print(f"       - {reannotations_llm_csv}")
                print(f"       - {default_reannotation_file}")
                return
        
        # Vérifier que les colonnes nécessaires existent
        required_cols = ['id', 'actions_llm_nouveau_PRINCIPALE', 'actions_llm_nouveau_INGREDIENTS']
        missing_cols = [col for col in required_cols if col not in reannotated_df.columns]
        if missing_cols:
            print(f"  ❌ Colonnes manquantes dans le fichier de ré-annotations: {missing_cols}")
            return
        
        print(f"  ✅ Ré-annotations chargées : {len(reannotated_df)} recettes")
        
        # Étape 3 : Comparaison tripartite
        print("\nÉTAPE 3/3 : Comparaison tripartite...")
        metrics_p, metrics_i = perform_tripartite_comparison(manual_df, reannotated_df, output_dir)
        
        # Étape 4 : Visualisations et rapport
        print("\nÉTAPE 4/3 : Génération visualisations et rapport...")
        viz_file = plot_comparison_summary(metrics_p, metrics_i, output_dir)
        generate_validation_report_strat_1(metrics_p, metrics_i, output_dir, viz_file)
        
        print(f"\n{'#'*80}")
        print(f"# PIPELINE TERMINÉ AVEC SUCCÈS (MODE COMPARAISON)")
        print(f"{'#'*80}\n")
        print(f"📁 Tous les résultats sont dans : {output_dir}/")
        print(f"   - Métriques : comparison_metrics_principale.csv & comparison_metrics_ingredients.csv")
        print(f"   - Visualisation : visualizations/comparison_summary.png")
        print(f"   - Rapport : rapport_validation_annotations.md")
        
    # ========== MODE COMPLET : Tout le pipeline ==========
    else:
        print(f"\n{'#'*80}")
        print(f"# PIPELINE STRATÉGIE 1 - VALIDATION PAR ANNOTATIONS MANUELLES")
        print(f"# PARTIE 1:  ÉCHANTILLONNAGE + EXCEL + RÉ-ANNOTATION")
        print(f"{'#'*80}\n")
        
        # Vérifier que les fichiers requis sont fournis
        if not all([recipes_csv, recipe_instructions_csv, recipe_ingredients_csv, graphs_csv]):
            print("❌ ERREUR : En mode complet, tous les fichiers CSV sont requis:")
            print("   - recipes_csv")
            print("   - recipe_instructions_csv")
            print("   - recipe_ingredients_csv")
            print("   - graphs_csv")
            return
        
        # Étape 1 : Chargement des données
        print("ÉTAPE 1/7 : Chargement des données...")
        recipes_df = pd.read_csv(recipes_csv)
        recipe_instructions_df = pd.read_csv(recipe_instructions_csv)
        recipe_ingredients_df = pd.read_csv(recipe_ingredients_csv)
        graphs_df = pd.read_csv(graphs_csv)
        
        print(f"  ✅ Recettes chargées : {len(recipes_df):,}")
        print(f"  ✅ Graphes chargés : {len(graphs_df):,}")
        
        # Étape 2a : Filtrage des recettes critiques (Stratégie 2)
        print("\nÉTAPE 2a/7 : Filtrage des recettes critiques (Stratégie 2)...")
        if critical_flags_csv:
            graphs_df = filter_critical_recipes(graphs_df, critical_flags_csv)
        else:
            print("  ⚠️  Aucun fichier de flags critiques fourni - analyse complète du dataset")
        
        
        # Étape 2b: Échantillonnage stratifié
        print(f"\nÉTAPE 2b/7 : Échantillonnage stratifié ({nombre_recettes} recettes)...")
        sampled_recipes = stratified_sampling(
            recipes_df,
            recipe_instructions_df,
            recipe_ingredients_df,
            graphs_df,
            nombre_recettes=nombre_recettes,
            random_state=random_state
        )
        
        # Étape 3 : Génération fichier Excel
        print("\nÉTAPE 3/7 : Génération du fichier Excel...")
        excel_file = os.path.join(output_dir, 'echantillon_annotation_manuelle.xlsx')
        create_annotation_excel(sampled_recipes, excel_file)
        
        # Étape 4 : Ré-annotation LLM
        print("\nÉTAPE 4/7 : Ré-annotation LLM...")
        reannotation_file = os.path.join(output_dir, 'reannotations_llm.csv')
        
        if openrouter_api_key:
            reannotated_df = reannotate_with_llm(
                sampled_recipes, 
                api_key=openrouter_api_key,
                model_name=model_name,
                output_file=reannotation_file
            )
        else:
            print("⚠️  Pas de clé API OpenRouter fournie - skip de la ré-annotation")
            reannotated_df = sampled_recipes[['id', 'title']].copy()
            reannotated_df['actions_llm_nouveau_PRINCIPALE'] = None
            reannotated_df['actions_llm_nouveau_INGREDIENTS'] = None
            reannotated_df.to_csv(reannotation_file, index=False, encoding='utf-8')
        
        # Fin du mode complet - attendre les annotations manuelles
        print(f"\n{'='*80}")
        print(f"⚠️  PIPELINE EN PAUSE - ANNOTATIONS MANUELLES REQUISES")
        print(f"{'='*80}\n")
        print(f"Prochaines étapes :")
        print(f"1. Ouvrir le fichier : {excel_file}")
        print(f"2. Remplir les colonnes :")
        print(f"   - annotations_manuelles_PRINCIPALE (colonne K)")
        print(f"   - annotations_manuelles_INGREDIENTS (colonne M)")
        print(f"3. Sauvegarder le fichier")
        print(f"4. Relancer le pipeline avec :")
        print(f"   manual_annotations_excel='{excel_file}'")
        print(f"   reannotations_llm_csv='{reannotation_file}'  ")
    
    print(f"\n✅ Stratégie 1 complétée !\n")

