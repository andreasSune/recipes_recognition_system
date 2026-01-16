"""
Pipeline complet pour la Stratégie 1 : Annotations Manuelles + Ré-annotation LLM
Validation de la qualité des données par comparaison tripartite

"""

import pandas as pd
import json
import re
import time
import os
from typing import List, Dict, Tuple
from datetime import datetime
from openai import OpenAI
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

import numpy as np
import ast
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.patches import Rectangle



# ==============================================================================
# SECTION 1 : ÉCHANTILLONNAGE STRATIFIÉ
# ==============================================================================

def classify_recipe_complexity(row: pd.Series) -> str:
    """
    Classifie la complexité d'une recette selon le nombre d'instructions et d'ingrédients
    
    Args:
        row: Ligne du DataFrame contenant 'number_of_steps' et 'number_of_ingredients'
    
    Returns:
        'SIMPLE', 'MOYENNE', ou 'COMPLEXE'
    """
    steps = row['number_of_steps']
    ingredients = row['number_of_ingredients']
    
    if steps <= 5 and ingredients <= 8:
        return 'SIMPLE'
    elif steps <= 10 and ingredients <= 15:
        return 'MOYENNE'
    else:
        return 'COMPLEXE'


def detect_pretransformed_ingredients(ingredients_list: List[str]) -> bool:
    """
    Détecte si la liste d'ingrédients contient des ingrédients pré-transformés
    
    Args:
        ingredients_list: Liste des ingrédients de la recette
    
    Returns:
        True si au moins un ingrédient pré-transformé est détecté
    """
    # Mots-clés indiquant une transformation
    transformation_keywords = [

    'sliced', 'minced',
    'crushed', 'ground', 'peeled', 'cored', 'julienned',
   
    
    'chopped', 'diced', 'sliced', 'minced', 'julienned', 'cubed', 'quartered',
    'halved', 'cut', 'chunked', 'wedged', 'bias-cut', 'fine-chopped', 
    'coarsely chopped', 'finely chopped', 'rough-chopped',

    'grated', 'shredded', 'zested', 'fine-grated', 'coarse-grated',
   
    'ground', 'crushed', 'mashed', 'pounded', 'pulverized', 'smashed',
    'crumbled', 'pressed', 'pureed', 'blended',
   
]
    ingredients_list = [str(x) for x in ingredients_list if pd.notnull(x)]
    ingredients_text = ' '.join(ingredients_list).lower()
    
    return any(keyword in ingredients_text for keyword in transformation_keywords)


def classify_cuisine_type(title: str, instructions: List[str]) -> str:
    """
    Classifie le type de cuisine d'une recette (heuristique simple)
    
    Args:
        title: Titre de la recette
        instructions: Liste des instructions
    
    Returns:
        'pâtisserie', 'plat_mijoté', 'préparation_rapide', ou 'autre'
    """
    title_lower = title.lower()
    instructions_text = ' '.join(instructions).lower()
    
    # Mots-clés pour la pâtisserie
    patisserie_keywords = ['cake', 'cookie', 'bread', 'pie', 'tart', 'pastry', 
                           'gâteau', 'biscuit', 'pain', 'tarte', 'pâtisserie']
    
    # Mots-clés pour plats mijotés
    mijote_keywords = ['stew', 'soup', 'braise', 'slow', 'simmer', 
                       'ragoût', 'soupe', 'mijoté']
    
    # Mots-clés pour préparations rapides
    rapide_keywords = ['salad', 'sandwich', 'smoothie', 'wrap', 'bowl',
                       'salade', 'cru']
    
    # Vérifier dans le titre et les instructions
    full_text = title_lower + ' ' + instructions_text
    
    if any(keyword in full_text for keyword in patisserie_keywords):
        return 'pâtisserie'
    elif any(keyword in full_text for keyword in mijote_keywords):
        return 'plat_mijoté'
    elif any(keyword in full_text for keyword in rapide_keywords):
        return 'préparation_rapide'
    else:
        return 'autre'


def stratified_sampling(
    recipes_df: pd.DataFrame,
    recipe_instructions_df: pd.DataFrame,
    recipe_ingredients_df: pd.DataFrame,
    graphs_recipes_df: pd.DataFrame,
    nb_recettes: int = 50,
    random_state: int = 42
) -> pd.DataFrame:
    """
    Effectue un échantillonnage stratifié des recettes selon la complexité et le type culinaire
    
    Args:
        recipes_df: DataFrame recipes.csv (colonnes: id, title, number_of_steps, number_of_ingredients)
        recipe_instructions_df: DataFrame recipe_instructions.csv (colonnes: id, instruction)
        recipe_ingredients_df: DataFrame recipe_ingredients.csv (colonnes: id, ingredients)
        graphs_recipes_df: DataFrame graphs_recipes (colonnes: id, title, actions, type, type_2)
        nb_recettes: Nombre total de recettes à échantillonner (défaut: 50)
        random_state: Seed pour la reproductibilité
    
    Returns:
        DataFrame avec les recettes échantillonnées et leurs métadonnées
    """
    print(f"\n{'='*80}")
    print(f"ÉCHANTILLONNAGE STRATIFIÉ - {nb_recettes} recettes")
    print(f"{'='*80}\n")
    
    # Joindre les données nécessaires
    recipes_with_graphs = recipes_df.merge(
        graphs_recipes_df[graphs_recipes_df['type_2'] == 'variante_principale'][['id', 'actions']],
        on='id',
        how='inner'
    )
    
    print(f"Recettes disponibles avec graphes : {len(recipes_with_graphs)}")
    
    # Classifier la complexité
    recipes_with_graphs['complexity_category'] = recipes_with_graphs.apply(
        classify_recipe_complexity, axis=1
    )
    
    # Obtenir les instructions et ingrédients pour chaque recette
    instructions_grouped = recipe_instructions_df.groupby('id')['instruction'].apply(list).reset_index()
    ingredients_grouped = recipe_ingredients_df.groupby('id')['ingredient'].apply(list).reset_index()
    
    recipes_with_graphs = recipes_with_graphs.merge(instructions_grouped, on='id', how='left')
    recipes_with_graphs = recipes_with_graphs.merge(ingredients_grouped, on='id', how='left')
    
    # Détecter les ingrédients pré-transformés
    recipes_with_graphs['has_pretransformed_ingredients'] = recipes_with_graphs['ingredient'].apply(
        lambda x: detect_pretransformed_ingredients(x) if isinstance(x, list) else False
    )
    
    # Classifier le type de cuisine
    recipes_with_graphs['cuisine_type'] = recipes_with_graphs.apply(
        lambda row: classify_cuisine_type(row['title'], row['instruction']) 
        if isinstance(row['instruction'], list) else 'autre',
        axis=1
    )
    
    # Distribution cible (pourcentages)
    complexity_distribution = {
        'SIMPLE': 0.30,
        'MOYENNE': 0.40,
        'COMPLEXE': 0.30
    }
    
    # Calculer le nombre de recettes par strate
    samples_per_complexity = {
        cat: max(1, int(nb_recettes * pct)) 
        for cat, pct in complexity_distribution.items()
    }
    
    # Ajuster pour atteindre exactement nb_recettes
    total_assigned = sum(samples_per_complexity.values())
    if total_assigned != nb_recettes:
        # Ajouter la différence à la catégorie MOYENNE
        samples_per_complexity['MOYENNE'] += (nb_recettes - total_assigned)
    
    print("\nDistribution cible par complexité :")
    for cat, count in samples_per_complexity.items():
        print(f"  {cat}: {count} recettes ({count/nb_recettes*100:.1f}%)")
    
    # Échantillonner par complexité
    sampled_recipes = []
    
    for complexity, n_samples in samples_per_complexity.items():
        complexity_recipes = recipes_with_graphs[
            recipes_with_graphs['complexity_category'] == complexity
        ]
        
        if len(complexity_recipes) < n_samples:
            print(f"\n⚠️  Attention: Seulement {len(complexity_recipes)} recettes {complexity} disponibles")
            n_samples = len(complexity_recipes)
        
        # Échantillonner avec diversité de type culinaire si possible
        sample = complexity_recipes.sample(n=n_samples, random_state=random_state)
        sampled_recipes.append(sample)
    
    # Combiner tous les échantillons
    final_sample = pd.concat(sampled_recipes, ignore_index=True)
    
    print(f"\n✅ Échantillonnage terminé : {len(final_sample)} recettes sélectionnées")
    
    # Afficher les statistiques de l'échantillon
    print("\nRépartition par complexité :")
    print(final_sample['complexity_category'].value_counts().sort_index())
    
    print("\nRépartition par type de cuisine :")
    print(final_sample['cuisine_type'].value_counts())
    
    print("\nRecettes avec ingrédients pré-transformés :")
    print(f"  Oui: {final_sample['has_pretransformed_ingredients'].sum()}")
    print(f"  Non: {(~final_sample['has_pretransformed_ingredients']).sum()}")
    
    return final_sample


# ==============================================================================
# SECTION 2 : GÉNÉRATION DU FICHIER EXCEL D'ANNOTATION
# ==============================================================================

def create_annotation_excel(
    sampled_recipes: pd.DataFrame,
    output_path: str = 'echantillon_annotation_manuelle.xlsx'
) -> str:
    """
    Crée un fichier Excel structuré pour l'annotation manuelle
    
    Args:
        sampled_recipes: DataFrame des recettes échantillonnées
        output_path: Chemin du fichier Excel de sortie
    
    Returns:
        Chemin du fichier créé
    """
    print(f"\n{'='*80}")
    print(f"CRÉATION DU FICHIER EXCEL D'ANNOTATION")
    print(f"{'='*80}\n")
    
    # Préparer le DataFrame pour l'export
    export_df = sampled_recipes[[
        'id', 'title', 'number_of_steps', 'number_of_ingredients',
        'complexity_category', 'cuisine_type', 'has_pretransformed_ingredients',
        'instruction', 'ingredient', 'actions'
    ]].copy()
    
    # Convertir les listes en strings formatées pour Excel
    export_df['instruction'] = export_df['instruction'].apply(
        lambda x: '\n'.join([f"{i+1}. {instr}" for i, instr in enumerate(x)]) 
        if isinstance(x, list) else ''
    )
    export_df['ingredient'] = export_df['ingredient'].apply(
        lambda x: '\n'.join([f"- {ing}" for ing in x]) 
        if isinstance(x, list) else ''
    )
    export_df['actions_llm_ancien'] = export_df['actions']
    
    # Ajouter la colonne vide pour annotations manuelles
    export_df['annotations_manuelles'] = ''
    
    # Sélectionner et réordonner les colonnes finales
    final_columns = [
        'id', 'title', 'number_of_steps', 'number_of_ingredients',
        'complexity_category', 'cuisine_type', 'has_pretransformed_ingredients',
        'instruction', 'ingredient', 'actions_llm_ancien', 'annotations_manuelles'
    ]
    export_df = export_df[final_columns]
    
    # Créer le workbook Excel avec mise en forme
    wb = Workbook()
    ws = wb.active
    ws.title = "Annotations"
    
    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    annotation_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Écrire les en-têtes
    headers = final_columns
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Écrire les données
    for row_idx, row_data in enumerate(export_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Mettre en évidence la colonne annotations_manuelles
            if col_idx == len(headers):  # Dernière colonne
                cell.fill = annotation_fill
    
    # Ajuster les largeurs de colonnes
    column_widths = {
        'id': 12,
        'title': 30,
        'number_of_steps': 10,
        'number_of_ingredients': 10,
        'complexity_category': 12,
        'cuisine_type': 15,
        'has_pretransformed_ingredients': 12,
        'instruction': 50,
        'ingredient': 40,
        'actions_llm_ancien': 50,
        'annotations_manuelles': 50
    }
    
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = column_widths[header]
    
    # Figer la première ligne
    ws.freeze_panes = 'A2'
    
    # Créer un second onglet pour les instructions
    ws_instructions = wb.create_sheet("Instructions")
    
    instructions_text = [
        ["INSTRUCTIONS POUR L'ANNOTATION MANUELLE"],
        [""],
        ["Objectif", "Générer manuellement la séquence d'actions (verbes) pour chaque recette"],
        [""],
        ["Consignes générales :"],
        ["1. Extraction des verbes", "Identifier le verbe d'action principal de chaque instruction"],
        ["", "Utiliser la liste de verbes culinaires de référence quand possible"],
        ["", "Si nécessaire, choisir un nouveau verbe approprié"],
        [""],
        ["2. Nombre d'actions par instruction", "Une instruction peut générer 0 à 3 actions :"],
        ["", "- 0 action : instructions non-culinaires ou purement informatives"],
        ["", "- 1 action : instruction simple (ex: 'Chop the onions' → 'chop')"],
        ["", "- 2-3 actions : instruction complexe (ex: 'Beat eggs and pour into pan' → 'beat, pour')"],
        [""],
        ["3. Gestes implicites à capturer", "Si un ingrédient est mentionné sous forme transformée (ex: 'fromage râpé')"],
        ["", "alors qu'il apparaît non-transformé dans la liste d'ingrédients,"],
        ["", "ajouter le geste de transformation (ex: 'grate')"],
        [""],
        ["4. Actions passives", "Pour 'laisser reposer', 'préchauffer', 'cuire 30 min', utiliser :"],
        ["", "- Les verbes de cuisson passive (bake, simmer, etc.)"],
        ["", "- Ou 'idle' si aucune action"],
        [""],
        ["5. Format de sortie", "Séquence d'actions séparées par des virgules dans la colonne 'annotations_manuelles'"],
        ["", "Format : action1, action2, action3, ..."],
        ["", "Exemple : wash, chop, heat, pour, stir, simmer, drain, serve"],
        [""],
        ["Liste de référence des verbes culinaires :"],
        ["bake, boil, fry, grill, roast, steam, sauté, simmer, broil, toast, heat, warm, cool, chill, freeze,"],
        ["melt, dissolve, mix, stir, combine, chop, dice, slice, cut, mince, grate, peel, core, bone, fillet,"],
        ["trim, wash, clean, marinate, season, salt, oil, grease, coat, stuff, fill, wrap, tie, arrange, prepare,"],
        ["braise, poach, blanch, sear, brown, caramelize, glaze, reduce, thicken, whip, beat, fold, knead, rise,"],
        ["proof, ferment, smoke, cure, pickle, preserve, drain, strain, press, squeeze, mash, puree, blend, whisk,"],
        ["cream, emulsify, separate, extract, filter, sift, dust, deep fry, stir fry, pan fry, barbecue,"],
        ["pressure cook, slow cook, baste, flip, turn, toss, skewer, pierce, prick, garnish, plate, serve,"],
        ["drizzle, sprinkle, brush, spread, layer, top, decorate, reheat, taste, adjust, finish"]
    ]
    
    for row_idx, row_content in enumerate(instructions_text, start=1):
        for col_idx, cell_value in enumerate(row_content, start=1):
            cell = ws_instructions.cell(row=row_idx, column=col_idx)
            cell.value = cell_value
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif len(row_content) > 1 and col_idx == 1:
                cell.font = Font(bold=True)
    
    ws_instructions.column_dimensions['A'].width = 30
    ws_instructions.column_dimensions['B'].width = 80
    
    # Sauvegarder le fichier
    wb.save(output_path)
    
    print(f"✅ Fichier Excel créé : {output_path}")
    print(f"   - Onglet 'Annotations' : {len(export_df)} recettes à annoter")
    print(f"   - Onglet 'Instructions' : Guide d'annotation")
    print(f"\n📝 Prochaine étape : Remplir la colonne 'annotations_manuelles' manuellement")
    
    return output_path


# ==============================================================================
# SECTION 3 : RÉ-ANNOTATION LLM (NOUVEAU)
# ==============================================================================

def create_recipe_reannotation_prompt(
    instructions: List[str]
) -> str:
    """
    Crée un prompt optimisé pour ré-annoter une recette complète
    Génère une séquence d'actions UNIQUEMENT à partir des instructions
    
    Args:
        instructions: Liste des instructions de la recette
    
    Returns:
        Prompt formaté pour le LLM
    """
    culinary_verbs = [
        "bake", "boil", "fry", "grill", "roast", "steam", "sauté", "simmer", "broil", "toast",
        "heat", "warm", "cool", "chill", "freeze", "melt", "dissolve", "mix", "stir", "combine",
        "chop", "dice", "slice", "cut", "mince", "grate", "peel", "core", "bone", "fillet",
        "trim", "wash", "clean", "marinate", "season", "salt", "oil", "grease", "coat", "stuff",
        "fill", "wrap", "tie", "arrange", "prepare", "braise", "poach", "blanch", "sear", "brown",
        "caramelize", "glaze", "reduce", "thicken", "whip", "beat", "fold", "knead", "rise", "proof",
        "ferment", "smoke", "cure", "pickle", "preserve", "drain", "strain", "press", "squeeze", "mash",
        "puree", "blend", "whisk", "cream", "emulsify", "separate", "extract", "filter", "sift", "dust",
        "deep fry", "stir fry", "pan fry", "barbecue", "pressure cook", "slow cook", "baste", "flip", 
        "turn", "toss", "skewer", "pierce", "prick", "garnish", "plate", "serve", "drizzle", "sprinkle", 
        "brush", "spread", "layer", "top", "decorate", "reheat", "taste", "adjust", "finish", "measure", 
        "weigh", "scale", "idle"
    ]
    
    # Formater les instructions avec numéros
    instructions_formatted = '\n'.join([f"{i+1}. {instr}" for i, instr in enumerate(instructions)])
    
    prompt = f"""You are an expert culinary analyst. Your task is to extract the sequence of cooking actions (verbs) from recipe instructions.

REFERENCE COOKING VERBS (choose from these when possible):
{', '.join(culinary_verbs)}

RECIPE INSTRUCTIONS TO ANALYZE:
{instructions_formatted}

TASK:
Generate the complete sequence of cooking actions (verbs) by analyzing ONLY the instructions above.

CRITICAL RULES:
1. Extract the main cooking verb from each instruction
2. Choose from the reference list when possible, or use an appropriate alternative if needed
3. Each instruction can generate 0 to 3 actions:
   - 0 actions: Non-cooking content (informative only)
   - 1 action: Simple instruction (e.g., "Chop onions" → "chop")
   - 2-3 actions: Complex instruction (e.g., "Beat eggs and pour into pan" → "beat", "pour")

4. IMPLICIT ACTIONS - Important:
   If an instruction mentions an ingredient in a transformed form (e.g., "add grated cheese", "chopped onions"),
   AND this transformation is mentioned IN THE INSTRUCTION ITSELF,
   you MUST ADD the transformation action before using that ingredient.
   
   Examples:
   - "Add grated cheese" → "grate, add" (grating is mentioned in instruction)
   - "Add chopped onions" → "chop, add" (chopping is mentioned in instruction)
   - "Add cheese" → "add" (no transformation mentioned, just add)

5. For passive cooking steps (bake, simmer, rest), use the appropriate verb or "idle"

6. Return ONLY the sequence of verbs, separated by commas, in chronological order

EXAMPLES:
- "Preheat oven to 350°F. Mix flour and sugar." → "heat, mix"
- "Chop onions and garlic. Sauté in oil until golden." → "chop, sauté"
- "Add shredded carrots and mix well." → "shred, add, mix"
- "Add carrots and mix well." → "add, mix"

OUTPUT FORMAT:
Return ONLY a comma-separated list of verbs (no JSON, no explanations):
verb1, verb2, verb3, ...

Begin:"""
    
    return prompt


def extract_action_sequence_from_response(content: str) -> List[str]:
    """
    Extrait la séquence d'actions de la réponse du LLM
    
    Args:
        content: Réponse du LLM
    
    Returns:
        Liste des actions extraites
    """
    # Nettoyer la réponse
    content = content.strip()
    
    # Retirer les balises ou texte superflu
    content = re.sub(r'```.*?```', '', content, flags=re.DOTALL)
    content = re.sub(r'\[.*?\]', '', content)
    content = re.sub(r'\{.*?\}', '', content)
    
    # Extraire les verbes séparés par des virgules
    # Chercher une ligne contenant des mots séparés par des virgules
    lines = content.split('\n')
    
    for line in lines:
        line = line.strip()
        if ',' in line and len(line.split(',')) > 1:
            # C'est probablement notre séquence
            actions = [action.strip().lower() for action in line.split(',')]
            # Filtrer les éléments vides
            actions = [a for a in actions if a and len(a) > 0]
            if len(actions) > 0:
                return actions
    
    # Si pas trouvé avec virgules, essayer de spliter sur espaces
    words = content.lower().split()
    # Prendre seulement les mots alphabétiques
    actions = [w.strip('.,;:!?') for w in words if w.strip('.,;:!?').isalpha()]
    
    return actions if actions else ['NA']


def reannotate_recipes_with_llm(
    sampled_recipes: pd.DataFrame,
    api_key: str,
    model: str = "mistralai/mistral-small-3.1-24b-instruct",
    output_dir: str = "strategy_1_results",
    save_interval: int = 10
) -> pd.DataFrame:
    """
    Ré-annote les recettes échantillonnées avec le LLM
    
    Args:
        sampled_recipes: DataFrame des recettes échantillonnées
        api_key: Clé API Anthropic
        model: Modèle à utiliser
        output_dir: Répertoire de sauvegarde
        save_interval: Intervalle de sauvegarde (tous les N recettes)
    
    Returns:
        DataFrame avec colonne 'actions_llm_nouveau' ajoutée
    """
    print(f"\n{'='*80}")
    print(f"RÉ-ANNOTATION LLM - {len(sampled_recipes)} recettes")
    print(f"{'='*80}\n")
    
    # Créer le répertoire de sortie
    os.makedirs(output_dir, exist_ok=True)
    
    # Initialiser le client OpenAI (compatible Anthropic)
    client = OpenAI(
        base_url="https://openrouter.ai/api/v1",
        api_key=api_key,
            )
    
    results = []
    failed_recipes = []
    
    total_recipes = len(sampled_recipes)
    
    for idx, row in sampled_recipes.iterrows():
        recipe_id = row['id']
        title = row['title']
        instructions = row['instruction'] if isinstance(row['instruction'], list) else []
        # ingredients = row['ingredients']  <- LIGNE À RETIRER
        
        print(f"\n[{idx+1}/{total_recipes}] Traitement : {title} (ID: {recipe_id})")
        
        try:
            # Créer le prompt
            prompt = create_recipe_reannotation_prompt(instructions)

            # Appeler le LLM
            response = client.chat.completions.create(
                model=model,
                messages=[
                   {
                        "role": "system",
                        "content": "Tu es un expert culinaire qui corrige les séquences d'actions de cuisine avec précision."
                    },
                    {
                        "role": "user",
                        "content": prompt
                    }
                ],
                max_tokens=3000,
                temperature=0.1
            )
            
            # Extraire la réponse
            content = response.choices[0].message.content
            print(f"Réponse brute (100 premiers caractères): {content[:100]}...")
            # Parser la séquence d'actions
            actions_sequence = extract_action_sequence_from_response(content)
            
            print(f"   ✅ Actions extraites : {', '.join(actions_sequence[:10])}{'...' if len(actions_sequence) > 10 else ''}")
            print(f"   Nombre d'actions : {len(actions_sequence)}")
            
            results.append({
                'id': recipe_id,
                'title': title,
                'actions_llm_nouveau': actions_sequence,
                'success': True
            })
            
            # Pause pour éviter rate limiting
            time.sleep(1)
            
        except Exception as e:
            print(f"   ❌ Erreur : {str(e)}")
            failed_recipes.append({
                'id': recipe_id,
                'title': title,
                'error': str(e)
            })
            
            results.append({
                'id': recipe_id,
                'title': title,
                'actions_llm_nouveau': ['ERROR'],
                'success': False
            })
        
        # Sauvegarde intermédiaire
        if (idx + 1) % save_interval == 0:
            temp_file = os.path.join(output_dir, f"temp_reannotation_progress_{idx+1}.json")
            with open(temp_file, 'w', encoding='utf-8') as f:
                json.dump({
                    'results': results,
                    'failed_recipes': failed_recipes,
                    'processed': idx + 1,
                    'total': total_recipes
                }, f, indent=2, ensure_ascii=False)
            print(f"\n💾 Sauvegarde intermédiaire : {temp_file}")
    
    # Convertir en DataFrame
    results_df = pd.DataFrame(results)
    
    # Fusionner avec sampled_recipes
    final_df = sampled_recipes.merge(
        results_df[['id', 'actions_llm_nouveau', 'success']],
        on='id',
        how='left'
    )
    
    # Sauvegarde finale
    final_file = os.path.join(output_dir, 'reannotation_llm_results.json')
    with open(final_file, 'w', encoding='utf-8') as f:
        json.dump({
            'results': results,
            'failed_recipes': failed_recipes,
            'total_processed': len(results),
            'total_success': sum(1 for r in results if r['success']),
            'total_failed': len(failed_recipes)
        }, f, indent=2, ensure_ascii=False)
    
    print(f"\n{'='*80}")
    print(f"RÉ-ANNOTATION TERMINÉE")
    print(f"{'='*80}")
    print(f"Total traité : {len(results)}")
    print(f"Succès : {sum(1 for r in results if r['success'])}")
    print(f"Échecs : {len(failed_recipes)}")
    print(f"Sauvegarde finale : {final_file}")
    
    return final_df


# ==============================================================================
# SECTION 4 : COMPARAISON TRIPARTITE ET MÉTRIQUES
# ==============================================================================

def calculate_exact_match(seq1: List[str], seq2: List[str]) -> bool:
    """Calcule si deux séquences sont exactement identiques"""
    return seq1 == seq2


def calculate_accuracy(seq1: List[str], seq2: List[str]) -> float:
    """
    Calcule la précision des verbes (proportion de verbes identiques en position identique)
    """
    if len(seq1) == 0 and len(seq2) == 0:
        return 1.0
    
    max_len = max(len(seq1), len(seq2))
    if max_len == 0:
        return 1.0
    
    matches = sum(1 for i in range(min(len(seq1), len(seq2))) if seq1[i] == seq2[i])
    return matches / max_len


def calculate_jaccard_similarity(seq1: List[str], seq2: List[str]) -> float:
    """
    Calcule la similarité de Jaccard (overlap sans tenir compte de l'ordre)
    """
    set1 = set(seq1)
    set2 = set(seq2)
    
    if len(set1) == 0 and len(set2) == 0:
        return 1.0
    
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    
    return intersection / union if union > 0 else 0.0


def calculate_levenshtein_distance(seq1: List[str], seq2: List[str]) -> int:
    """
    Calcule la distance de Levenshtein (distance d'édition) entre deux séquences
    """
    if len(seq1) == 0:
        return len(seq2)
    if len(seq2) == 0:
        return len(seq1)
    
    # Matrice de programmation dynamique
    dp = [[0] * (len(seq2) + 1) for _ in range(len(seq1) + 1)]
    
    # Initialisation
    for i in range(len(seq1) + 1):
        dp[i][0] = i
    for j in range(len(seq2) + 1):
        dp[0][j] = j
    
    # Calcul
    for i in range(1, len(seq1) + 1):
        for j in range(1, len(seq2) + 1):
            if seq1[i-1] == seq2[j-1]:
                dp[i][j] = dp[i-1][j-1]
            else:
                dp[i][j] = 1 + min(
                    dp[i-1][j],      # Suppression
                    dp[i][j-1],      # Insertion
                    dp[i-1][j-1]     # Substitution
                )
    
    return dp[len(seq1)][len(seq2)]


def calculate_lcs_ratio(seq1: List[str], seq2: List[str]) -> float:
    """
    Calcule le ratio de la plus longue sous-séquence commune (LCS)
    """
    if len(seq1) == 0 or len(seq2) == 0:
        return 0.0
    
    # Matrice LCS
    dp = [[0] * (len(seq2) + 1) for _ in range(len(seq1) + 1)]
    
    for i in range(1, len(seq1) + 1):
        for j in range(1, len(seq2) + 1):
            if seq1[i-1] == seq2[j-1]:
                dp[i][j] = dp[i-1][j-1] + 1
            else:
                dp[i][j] = max(dp[i-1][j], dp[i][j-1])
    
    lcs_length = dp[len(seq1)][len(seq2)]
    max_len = max(len(seq1), len(seq2))
    
    return lcs_length / max_len if max_len > 0 else 0.0


def compare_sequences(
    seq_manual: List[str],
    seq_ancien: List[str],
    seq_nouveau: List[str]
) -> Dict:
    """
    Compare trois séquences et calcule toutes les métriques
    
    Args:
        seq_manual: Séquence annotée manuellement
        seq_ancien: Séquence LLM ancien
        seq_nouveau: Séquence LLM nouveau
    
    Returns:
        Dictionnaire contenant toutes les métriques de comparaison
    """
    metrics = {}
    
    # Comparaison 1: Manuel vs Ancien
    metrics['manuel_vs_ancien'] = {
        'exact_match': calculate_exact_match(seq_manual, seq_ancien),
        'accuracy': calculate_accuracy(seq_manual, seq_ancien),
        'jaccard': calculate_jaccard_similarity(seq_manual, seq_ancien),
        'levenshtein': calculate_levenshtein_distance(seq_manual, seq_ancien),
        'lcs_ratio': calculate_lcs_ratio(seq_manual, seq_ancien),
        'length_diff': abs(len(seq_manual) - len(seq_ancien))
    }
    
    # Comparaison 2: Manuel vs Nouveau
    metrics['manuel_vs_nouveau'] = {
        'exact_match': calculate_exact_match(seq_manual, seq_nouveau),
        'accuracy': calculate_accuracy(seq_manual, seq_nouveau),
        'jaccard': calculate_jaccard_similarity(seq_manual, seq_nouveau),
        'levenshtein': calculate_levenshtein_distance(seq_manual, seq_nouveau),
        'lcs_ratio': calculate_lcs_ratio(seq_manual, seq_nouveau),
        'length_diff': abs(len(seq_manual) - len(seq_nouveau))
    }
    
    # Comparaison 3: Ancien vs Nouveau
    metrics['ancien_vs_nouveau'] = {
        'exact_match': calculate_exact_match(seq_ancien, seq_nouveau),
        'accuracy': calculate_accuracy(seq_ancien, seq_nouveau),
        'jaccard': calculate_jaccard_similarity(seq_ancien, seq_nouveau),
        'levenshtein': calculate_levenshtein_distance(seq_ancien, seq_nouveau),
        'lcs_ratio': calculate_lcs_ratio(seq_ancien, seq_nouveau),
        'length_diff': abs(len(seq_ancien) - len(seq_nouveau))
    }
    
    # Catégorisation des changements (Ancien vs Nouveau par rapport à Manuel)
    ancien_correct = calculate_exact_match(seq_ancien, seq_manual)
    nouveau_correct = calculate_exact_match(seq_nouveau, seq_manual)
    ancien_vs_nouveau_same = calculate_exact_match(seq_ancien, seq_nouveau)
    
    if ancien_correct and nouveau_correct:
        change_category = 'stable_correct'
    elif not ancien_correct and nouveau_correct:
        change_category = 'correction'
    elif ancien_correct and not nouveau_correct:
        change_category = 'regression'
    elif not ancien_correct and not nouveau_correct:
        if ancien_vs_nouveau_same:
            change_category = 'stable_incorrect'
        else:
            change_category = 'changement_lateral'
    else:
        change_category = 'unknown'
    
    metrics['change_category'] = change_category
    
    return metrics


def load_manual_annotations(excel_file: str) -> pd.DataFrame:
    """
    Charge les annotations manuelles depuis le fichier Excel rempli
    
    Args:
        excel_file: Chemin du fichier Excel avec annotations manuelles
    
    Returns:
        DataFrame avec les annotations manuelles parsées
    """
    print(f"\n{'='*80}")
    print(f"CHARGEMENT DES ANNOTATIONS MANUELLES")
    print(f"{'='*80}\n")
    
    df = pd.read_excel(excel_file, sheet_name='Annotations')
    
    print(f"Recettes chargées : {len(df)}")
    
    # Parser la colonne annotations_manuelles (format: "action1, action2, action3")
    def parse_manual_annotation(annotation_str):
        if pd.isna(annotation_str) or annotation_str == '':
            return []
        
        # Split par virgules et nettoyer
        actions = [action.strip().lower() for action in str(annotation_str).split(',')]
        # Filtrer les éléments vides
        actions = [a for a in actions if a and len(a) > 0]
        return actions
    
    df['annotations_manuelles_parsed'] = df['annotations_manuelles'].apply(parse_manual_annotation)
    
    # Parser actions_llm_ancien aussi (format: "action1, action2, action3")
    def parse_llm_ancien(actions_str):
        if pd.isna(actions_str) or actions_str == '':
            return []
        actions = [action.strip().lower() for action in str(actions_str).split(',')]
        return [a for a in actions if a and len(a) > 0]
    
    df['actions_llm_ancien_parsed'] = df['actions_llm_ancien'].apply(parse_llm_ancien)
    
    # Vérifier combien de recettes ont été annotées
    annotated_count = df['annotations_manuelles_parsed'].apply(lambda x: len(x) > 0).sum()
    print(f"Recettes annotées manuellement : {annotated_count}/{len(df)}")
    
    if annotated_count < len(df):
        print(f"⚠️  Attention : {len(df) - annotated_count} recettes n'ont pas été annotées")
    
    return df


def perform_tripartite_comparison(
    manual_annotations_df: pd.DataFrame,
    reannotated_df: pd.DataFrame,
    output_dir: str = "strategy_1_results"
) -> pd.DataFrame:
    """
    Effectue la comparaison tripartite complète et génère les métriques
    
    Args:
        manual_annotations_df: DataFrame avec annotations manuelles chargées
        reannotated_df: DataFrame avec ré-annotations LLM
        output_dir: Répertoire de sortie
    
    Returns:
        DataFrame avec toutes les métriques de comparaison
    """
    print(f"\n{'='*80}")
    print(f"COMPARAISON TRIPARTITE ET CALCUL DES MÉTRIQUES")
    print(f"{'='*80}\n")
    
    # Fusionner les DataFrames
    comparison_df = manual_annotations_df.merge(
        reannotated_df[['id', 'actions_llm_nouveau']],
        on='id',
        how='inner'
    )
    
    print(f"Recettes à comparer : {len(comparison_df)}")
    
    # Calculer les métriques pour chaque recette
    all_metrics = []
    
    for idx, row in comparison_df.iterrows():
        recipe_id = row['id']
        title = row['title']
        
        seq_manual = row['annotations_manuelles_parsed']
        seq_ancien = row['actions_llm_ancien_parsed']
        seq_nouveau = row['actions_llm_nouveau'] if isinstance(row['actions_llm_nouveau'], list) else []
        
        # Skip si pas d'annotation manuelle
        if len(seq_manual) == 0:
            continue
        
        metrics = compare_sequences(seq_manual, seq_ancien, seq_nouveau)
        
        all_metrics.append({
            'id': recipe_id,
            'title': title,
            'complexity': row.get('complexity_category', 'unknown'),
            'cuisine_type': row.get('cuisine_type', 'unknown'),
            
            # Longueurs des séquences
            'len_manual': len(seq_manual),
            'len_ancien': len(seq_ancien),
            'len_nouveau': len(seq_nouveau),
            
            # Métriques Manuel vs Ancien
            'man_anc_exact_match': metrics['manuel_vs_ancien']['exact_match'],
            'man_anc_accuracy': metrics['manuel_vs_ancien']['accuracy'],
            'man_anc_jaccard': metrics['manuel_vs_ancien']['jaccard'],
            'man_anc_levenshtein': metrics['manuel_vs_ancien']['levenshtein'],
            'man_anc_lcs_ratio': metrics['manuel_vs_ancien']['lcs_ratio'],
            'man_anc_length_diff': metrics['manuel_vs_ancien']['length_diff'],
            
            # Métriques Manuel vs Nouveau
            'man_nouv_exact_match': metrics['manuel_vs_nouveau']['exact_match'],
            'man_nouv_accuracy': metrics['manuel_vs_nouveau']['accuracy'],
            'man_nouv_jaccard': metrics['manuel_vs_nouveau']['jaccard'],
            'man_nouv_levenshtein': metrics['manuel_vs_nouveau']['levenshtein'],
            'man_nouv_lcs_ratio': metrics['manuel_vs_nouveau']['lcs_ratio'],
            'man_nouv_length_diff': metrics['manuel_vs_nouveau']['length_diff'],
            
            # Métriques Ancien vs Nouveau
            'anc_nouv_exact_match': metrics['ancien_vs_nouveau']['exact_match'],
            'anc_nouv_accuracy': metrics['ancien_vs_nouveau']['accuracy'],
            'anc_nouv_jaccard': metrics['ancien_vs_nouveau']['jaccard'],
            'anc_nouv_levenshtein': metrics['ancien_vs_nouveau']['levenshtein'],
            'anc_nouv_lcs_ratio': metrics['ancien_vs_nouveau']['lcs_ratio'],
            'anc_nouv_length_diff': metrics['ancien_vs_nouveau']['length_diff'],
            
            # Catégorie de changement
            'change_category': metrics['change_category']
        })
    
    metrics_df = pd.DataFrame(all_metrics)
    
    # Sauvegarder les métriques détaillées
    os.makedirs(output_dir, exist_ok=True)
    metrics_file = os.path.join(output_dir, 'comparison_metrics_detailed.csv')
    metrics_df.to_csv(metrics_file, index=False, encoding='utf-8')
    
    print(f"✅ Métriques détaillées sauvegardées : {metrics_file}")
    
    return metrics_df


def generate_comparison_report(
    metrics_df: pd.DataFrame,
    output_dir: str = "strategy_1_results"
) -> str:
    """
    Génère un rapport de synthèse de la comparaison tripartite
    
    Args:
        metrics_df: DataFrame des métriques de comparaison
        output_dir: Répertoire de sortie
    
    Returns:
        Chemin du fichier de rapport
    """
    print(f"\n{'='*80}")
    print(f"GÉNÉRATION DU RAPPORT DE SYNTHÈSE")
    print(f"{'='*80}\n")
    
    report_lines = []
    
    report_lines.append("=" * 80)
    report_lines.append("RAPPORT DE COMPARAISON TRIPARTITE - STRATÉGIE 1")
    report_lines.append("=" * 80)
    report_lines.append("")
    report_lines.append(f"Date de génération : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report_lines.append(f"Nombre de recettes analysées : {len(metrics_df)}")
    report_lines.append("")
    
    # ========== TABLEAU 1 : MÉTRIQUES MOYENNES ==========
    report_lines.append("-" * 80)
    report_lines.append("TABLEAU 1 : MÉTRIQUES MOYENNES")
    report_lines.append("-" * 80)
    report_lines.append("")
    
    report_lines.append(f"{'Comparaison':<25} {'Exact Match':<15} {'Accuracy':<15} {'Jaccard':<15} {'Levenshtein':<15} {'LCS Ratio':<15}")
    report_lines.append("-" * 100)
    
    # Manuel vs Ancien
    exact_match_pct_anc = (metrics_df['man_anc_exact_match'].sum() / len(metrics_df) * 100)
    accuracy_anc = metrics_df['man_anc_accuracy'].mean()
    jaccard_anc = metrics_df['man_anc_jaccard'].mean()
    levenshtein_anc = metrics_df['man_anc_levenshtein'].mean()
    lcs_anc = metrics_df['man_anc_lcs_ratio'].mean()
    
    report_lines.append(
        f"{'Manuel vs Ancien':<25} {exact_match_pct_anc:>13.1f}% "
        f"{accuracy_anc:>14.3f} {jaccard_anc:>14.3f} {levenshtein_anc:>14.1f} {lcs_anc:>14.3f}"
    )
    
    # Manuel vs Nouveau
    exact_match_pct_nouv = (metrics_df['man_nouv_exact_match'].sum() / len(metrics_df) * 100)
    accuracy_nouv = metrics_df['man_nouv_accuracy'].mean()
    jaccard_nouv = metrics_df['man_nouv_jaccard'].mean()
    levenshtein_nouv = metrics_df['man_nouv_levenshtein'].mean()
    lcs_nouv = metrics_df['man_nouv_lcs_ratio'].mean()
    
    report_lines.append(
        f"{'Manuel vs Nouveau':<25} {exact_match_pct_nouv:>13.1f}% "
        f"{accuracy_nouv:>14.3f} {jaccard_nouv:>14.3f} {levenshtein_nouv:>14.1f} {lcs_nouv:>14.3f}"
    )
    
    # Ancien vs Nouveau
    exact_match_pct_anc_nouv = (metrics_df['anc_nouv_exact_match'].sum() / len(metrics_df) * 100)
    accuracy_anc_nouv = metrics_df['anc_nouv_accuracy'].mean()
    jaccard_anc_nouv = metrics_df['anc_nouv_jaccard'].mean()
    levenshtein_anc_nouv = metrics_df['anc_nouv_levenshtein'].mean()
    lcs_anc_nouv = metrics_df['anc_nouv_lcs_ratio'].mean()
    
    report_lines.append(
        f"{'Ancien vs Nouveau':<25} {exact_match_pct_anc_nouv:>13.1f}% "
        f"{accuracy_anc_nouv:>14.3f} {jaccard_anc_nouv:>14.3f} {levenshtein_anc_nouv:>14.1f} {lcs_anc_nouv:>14.3f}"
    )
    
    report_lines.append("")
    
    # ========== TABLEAU 2 : DISTRIBUTION DES CHANGEMENTS ==========
    report_lines.append("-" * 80)
    report_lines.append("TABLEAU 2 : DISTRIBUTION DES CHANGEMENTS (Ancien → Nouveau)")
    report_lines.append("-" * 80)
    report_lines.append("")
    
    change_counts = metrics_df['change_category'].value_counts()
    total = len(metrics_df)
    
    report_lines.append(f"{'Catégorie':<30} {'Nombre':<15} {'Pourcentage':<15}")
    report_lines.append("-" * 60)
    
    for category in ['correction', 'regression', 'changement_lateral', 'stable_correct', 'stable_incorrect']:
        count = change_counts.get(category, 0)
        pct = (count / total * 100) if total > 0 else 0
        report_lines.append(f"{category:<30} {count:<15} {pct:>13.1f}%")
    
    report_lines.append("")
    
    # ========== TABLEAU 3 : ANALYSE PAR COMPLEXITÉ ==========
    report_lines.append("-" * 80)
    report_lines.append("TABLEAU 3 : MÉTRIQUES PAR COMPLEXITÉ")
    report_lines.append("-" * 80)
    report_lines.append("")
    
    report_lines.append(f"{'Complexité':<15} {'N':<10} {'Accuracy Anc':<15} {'Accuracy Nouv':<15} {'Jaccard Anc':<15} {'Jaccard Nouv':<15}")
    report_lines.append("-" * 85)
    
    for complexity in ['SIMPLE', 'MOYENNE', 'COMPLEXE']:
        subset = metrics_df[metrics_df['complexity'] == complexity]
        if len(subset) > 0:
            n = len(subset)
            acc_anc = subset['man_anc_accuracy'].mean()
            acc_nouv = subset['man_nouv_accuracy'].mean()
            jac_anc = subset['man_anc_jaccard'].mean()
            jac_nouv = subset['man_nouv_jaccard'].mean()
            
            report_lines.append(
                f"{complexity:<15} {n:<10} {acc_anc:>14.3f} {acc_nouv:>14.3f} "
                f"{jac_anc:>14.3f} {jac_nouv:>14.3f}"
            )
    
    report_lines.append("")
    
    # ========== TABLEAU 4 : ANALYSE PAR TYPE CULINAIRE ==========
    report_lines.append("-" * 80)
    report_lines.append("TABLEAU 4 : MÉTRIQUES PAR TYPE CULINAIRE")
    report_lines.append("-" * 80)
    report_lines.append("")
    
    report_lines.append(f"{'Type Culinaire':<20} {'N':<10} {'Accuracy Anc':<15} {'Accuracy Nouv':<15} {'Jaccard Anc':<15} {'Jaccard Nouv':<15}")
    report_lines.append("-" * 90)
    
    for cuisine_type in metrics_df['cuisine_type'].unique():
        subset = metrics_df[metrics_df['cuisine_type'] == cuisine_type]
        if len(subset) > 0:
            n = len(subset)
            acc_anc = subset['man_anc_accuracy'].mean()
            acc_nouv = subset['man_nouv_accuracy'].mean()
            jac_anc = subset['man_anc_jaccard'].mean()
            jac_nouv = subset['man_nouv_jaccard'].mean()
            
            report_lines.append(
                f"{cuisine_type:<20} {n:<10} {acc_anc:>14.3f} {acc_nouv:>14.3f} "
                f"{jac_anc:>14.3f} {jac_nouv:>14.3f}"
            )
    
    report_lines.append("")
    
    # ========== CONCLUSION ==========
    report_lines.append("-" * 80)
    report_lines.append("CONCLUSION")
    report_lines.append("-" * 80)
    report_lines.append("")
    
    improvement = accuracy_nouv - accuracy_anc
    if improvement > 0:
        report_lines.append(f"✅ Le nouveau modèle LLM montre une AMÉLIORATION de {improvement:.3f} en accuracy.")
    elif improvement < 0:
        report_lines.append(f"❌ Le nouveau modèle LLM montre une RÉGRESSION de {abs(improvement):.3f} en accuracy.")
    else:
        report_lines.append(f"➖ Le nouveau modèle LLM montre une performance IDENTIQUE au précédent.")
    
    report_lines.append("")
    report_lines.append(f"Corrections réussies : {change_counts.get('correction', 0)} recettes")
    report_lines.append(f"Régressions : {change_counts.get('regression', 0)} recettes")
    report_lines.append(f"Stables corrects : {change_counts.get('stable_correct', 0)} recettes")
    report_lines.append("")
    
    report_lines.append("=" * 80)
    report_lines.append("FIN DU RAPPORT")
    report_lines.append("=" * 80)
    
    # Sauvegarder le rapport
    report_text = '\n'.join(report_lines)
    report_file = os.path.join(output_dir, 'comparison_report.txt')
    
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write(report_text)
    
    print(report_text)
    print(f"\n✅ Rapport sauvegardé : {report_file}")
    
    return report_file


# ==============================================================================
# SECTION 5 : PIPELINE PRINCIPAL   recipes_csv: str,
  #  recipe_instructions_csv: str,
  #  recipe_ingredients_csv: str,
# ==============================================================================

def run_strategy_1_pipeline(
    recipes_df,
    recipe_instructions_df,
    recipe_ingredients_df,
    graphs_recipes_df,
    api_key: str,
    nb_recettes: int = 50,
    output_dir: str = "strategy_1_results",
    random_state: int = 42
):
    """
    Exécute le pipeline complet de la Stratégie 1
    
    Args:
        recipes_csv: Chemin vers recipes.csv
        recipe_instructions_csv: Chemin vers recipe_instructions.csv
        recipe_ingredients_csv: Chemin vers recipe_ingredients.csv
        graphs_recipes_csv: Chemin vers graphs_recipes.csv
        anthropic_api_key: Clé API Anthropic
        nb_recettes: Nombre de recettes à échantillonner (défaut: 50)
        output_dir: Répertoire de sortie
        random_state: Seed pour reproductibilité
    """
    print(f"\n{'#'*80}")
    print(f"# PIPELINE STRATÉGIE 1 - VALIDATION PAR ANNOTATIONS MANUELLES")
    print(f"# Nombre de recettes : {nb_recettes}")
    print(f"{'#'*80}\n")
    
    # Créer le répertoire de sortie
    os.makedirs(output_dir, exist_ok=True)
    
    # ========== ÉTAPE 1 : CHARGEMENT DES DONNÉES ==========
    #print("ÉTAPE 1/6 : Chargement des données...")
    #recipes_df = pd.read_csv(recipes_csv)
    #recipe_instructions_df = pd.read_csv(recipe_instructions_csv)
    #recipe_ingredients_df = pd.read_csv(recipe_ingredients_csv)
    #graphs_recipes_df = pd.read_csv(graphs_recipes_csv)
    
    print(f"  ✅ Recettes : {len(recipes_df)}")
    print(f"  ✅ Instructions : {len(recipe_instructions_df)}")
    print(f"  ✅ Ingrédients : {len(recipe_ingredients_df)}")
    print(f"  ✅ Graphes : {len(graphs_recipes_df)}")
    
    # ========== ÉTAPE 2 : ÉCHANTILLONNAGE STRATIFIÉ ==========
    print("\nÉTAPE 2/6 : Échantillonnage stratifié...")
    sampled_recipes = stratified_sampling(
        recipes_df,
        recipe_instructions_df,
        recipe_ingredients_df,
        graphs_recipes_df,
        nb_recettes=nb_recettes,
        random_state=random_state
    )
    
    # ========== ÉTAPE 3 : GÉNÉRATION DU FICHIER EXCEL ==========
    print("\nÉTAPE 3/6 : Génération du fichier Excel pour annotation manuelle...")
    excel_file = os.path.join(output_dir, f'echantillon_annotation_manuelle_{nb_recettes}recettes.xlsx')
    create_annotation_excel(sampled_recipes, excel_file)
    
    print("\n" + "="*80)
    print("⏸️  PAUSE REQUISE : ANNOTATION MANUELLE")
    print("="*80)
    print(f"\n📋 Actions à effectuer :")
    print(f"   1. Ouvrir le fichier : {excel_file}")
    print(f"   2. Remplir la colonne 'annotations_manuelles' pour chaque recette")
    print(f"   3. Sauvegarder le fichier Excel")
    print(f"   4. Revenir exécuter la suite du pipeline\n")
    
    user_input = input("Avez-vous terminé l'annotation manuelle ? (oui/non) : ").strip().lower()
    
    if user_input != 'oui':
        print("\n⏸️  Pipeline mis en pause. Reprenez après avoir complété les annotations manuelles.")
        return
    
    # ========== ÉTAPE 4 : RÉ-ANNOTATION LLM ==========
    print("\nÉTAPE 4/6 : Ré-annotation avec le LLM...")
    reannotated_df = reannotate_recipes_with_llm(
        sampled_recipes,
        api_key,
        output_dir=output_dir
    )
    
    # ========== ÉTAPE 5 : CHARGEMENT DES ANNOTATIONS MANUELLES ==========
    print("\nÉTAPE 5/6 : Chargement des annotations manuelles...")
    manual_annotations_df = load_manual_annotations(excel_file)
    
    # ========== ÉTAPE 6 : COMPARAISON ET RAPPORT ==========
    print("\nÉTAPE 6/6 : Comparaison tripartite et génération du rapport...")
    metrics_df = perform_tripartite_comparison(
        manual_annotations_df,
        reannotated_df,
        output_dir
    )
    
    report_file = generate_comparison_report(metrics_df, output_dir)
    
    print(f"\n{'#'*80}")
    print(f"# PIPELINE TERMINÉ AVEC SUCCÈS")
    print(f"{'#'*80}\n")
    print(f"📁 Tous les résultats sont dans : {output_dir}/")
    print(f"   - Fichier Excel d'annotation : {excel_file}")
    print(f"   - Métriques détaillées : {output_dir}/comparison_metrics_detailed.csv")
    print(f"   - Rapport de synthèse : {report_file}")
    print(f"\n✅ Stratégie 1 complétée !\n")





# ==============================================================================
# SECTION 6 : VISUALISATIONS ET GRAPHIQUES
# ==============================================================================


# Configuration globale pour des graphiques professionnels
plt.style.use('seaborn-v0_8-darkgrid')
sns.set_palette("husl")


def plot_metrics_comparison_radar(metrics_df: pd.DataFrame, output_dir: str = "strategy_1_results"):
    """
    Graphique radar comparant les 3 versions (Manuel, Ancien, Nouveau)
    sur les métriques clés
    """
    fig, ax = plt.subplots(figsize=(10, 10), subplot_kw=dict(projection='polar'))
    
    # Métriques à afficher
    categories = ['Accuracy', 'Jaccard', 'LCS Ratio']
    N = len(categories)
    
    # Calculer les moyennes
    values_ancien = [
        metrics_df['man_anc_accuracy'].mean(),
        metrics_df['man_anc_jaccard'].mean(),
        metrics_df['man_anc_lcs_ratio'].mean()
    ]
    
    values_nouveau = [
        metrics_df['man_nouv_accuracy'].mean(),
        metrics_df['man_nouv_jaccard'].mean(),
        metrics_df['man_nouv_lcs_ratio'].mean()
    ]
    
    # Angles pour chaque axe
    angles = [n / float(N) * 2 * np.pi for n in range(N)]
    values_ancien += values_ancien[:1]
    values_nouveau += values_nouveau[:1]
    angles += angles[:1]
    
    # Tracer
    ax.plot(angles, values_ancien, 'o-', linewidth=2, label='LLM Ancien', color='#E74C3C')
    ax.fill(angles, values_ancien, alpha=0.25, color='#E74C3C')
    
    ax.plot(angles, values_nouveau, 'o-', linewidth=2, label='LLM Nouveau', color='#2ECC71')
    ax.fill(angles, values_nouveau, alpha=0.25, color='#2ECC71')
    
    # Personnalisation
    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(categories, size=12)
    ax.set_ylim(0, 1)
    ax.set_title('Comparaison des Performances\nManuel comme Référence', size=16, weight='bold', pad=20)
    ax.legend(loc='upper right', bbox_to_anchor=(1.3, 1.1), fontsize=11)
    ax.grid(True)
    
    plt.tight_layout()
    output_path = os.path.join(output_dir, 'radar_metrics_comparison.png')
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"✅ Graphique radar sauvegardé : {output_path}")
    return output_path


def plot_change_categories_distribution(metrics_df: pd.DataFrame, output_dir: str = "strategy_1_results"):
    """
    Graphique en barres empilées montrant la distribution des catégories de changements
    avec des couleurs expressives
    """
    fig, ax = plt.subplots(figsize=(12, 7))
    
    # Compter les catégories
    change_counts = metrics_df['change_category'].value_counts()
    total = len(metrics_df)
    
    # Ordre et couleurs
    categories_order = ['stable_correct', 'correction', 'stable_incorrect', 'changement_lateral', 'regression']
    colors = {
        'stable_correct': '#27AE60',      # Vert foncé
        'correction': '#2ECC71',          # Vert clair
        'stable_incorrect': '#95A5A6',    # Gris
        'changement_lateral': '#F39C12',  # Orange
        'regression': '#E74C3C'           # Rouge
    }
    labels_fr = {
        'stable_correct': 'Stable Correct',
        'correction': 'Corrections',
        'stable_incorrect': 'Stable Incorrect',
        'changement_lateral': 'Changements Latéraux',
        'regression': 'Régressions'
    }
    
    # Préparer les données
    counts = [change_counts.get(cat, 0) for cat in categories_order]
    percentages = [(c / total * 100) for c in counts]
    colors_list = [colors[cat] for cat in categories_order]
    labels_list = [labels_fr[cat] for cat in categories_order]
    
    # Créer le graphique
    bars = ax.bar(range(len(categories_order)), counts, color=colors_list, edgecolor='black', linewidth=1.5)
    
    # Ajouter les valeurs et pourcentages sur les barres
    for i, (bar, count, pct) in enumerate(zip(bars, counts, percentages)):
        height = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2., height,
                f'{count}\n({pct:.1f}%)',
                ha='center', va='bottom', fontsize=11, weight='bold')
    
    # Personnalisation
    ax.set_xticks(range(len(categories_order)))
    ax.set_xticklabels(labels_list, rotation=15, ha='right', fontsize=11)
    ax.set_ylabel('Nombre de Recettes', fontsize=12, weight='bold')
    ax.set_title('Distribution des Changements (Ancien → Nouveau)\npar rapport à l\'Annotation Manuelle', 
                 fontsize=14, weight='bold', pad=15)
    ax.grid(axis='y', alpha=0.3)
    
    # Ajouter une ligne horizontale pour la moyenne
    ax.axhline(y=total/len(categories_order), color='gray', linestyle='--', linewidth=1, alpha=0.5, label='Moyenne')
    
    plt.tight_layout()
    output_path = os.path.join(output_dir, 'change_categories_distribution.png')
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"✅ Graphique de distribution sauvegardé : {output_path}")
    return output_path


def plot_accuracy_by_complexity(metrics_df: pd.DataFrame, output_dir: str = "strategy_1_results"):
    """
    Graphique en barres groupées comparant l'accuracy par niveau de complexité
    """
    fig, ax = plt.subplots(figsize=(12, 7))
    
    # Préparer les données
    complexities = ['SIMPLE', 'MOYENNE', 'COMPLEXE']
    ancien_accuracies = []
    nouveau_accuracies = []
    
    for complexity in complexities:
        subset = metrics_df[metrics_df['complexity'] == complexity]
        if len(subset) > 0:
            ancien_accuracies.append(subset['man_anc_accuracy'].mean())
            nouveau_accuracies.append(subset['man_nouv_accuracy'].mean())
        else:
            ancien_accuracies.append(0)
            nouveau_accuracies.append(0)
    
    # Position des barres
    x = np.arange(len(complexities))
    width = 0.35
    
    # Créer les barres
    bars1 = ax.bar(x - width/2, ancien_accuracies, width, label='LLM Ancien', 
                   color='#E74C3C', edgecolor='black', linewidth=1.5)
    bars2 = ax.bar(x + width/2, nouveau_accuracies, width, label='LLM Nouveau', 
                   color='#2ECC71', edgecolor='black', linewidth=1.5)
    
    # Ajouter les valeurs sur les barres
    for bars in [bars1, bars2]:
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                    f'{height:.3f}',
                    ha='center', va='bottom', fontsize=10, weight='bold')
    
    # Personnalisation
    ax.set_xlabel('Niveau de Complexité', fontsize=12, weight='bold')
    ax.set_ylabel('Accuracy Moyenne', fontsize=12, weight='bold')
    ax.set_title('Performance par Niveau de Complexité\n(vs Annotations Manuelles)', 
                 fontsize=14, weight='bold', pad=15)
    ax.set_xticks(x)
    ax.set_xticklabels(complexities, fontsize=11)
    ax.legend(fontsize=11, loc='lower left')
    ax.set_ylim(0, 1.1)
    ax.grid(axis='y', alpha=0.3)
    
    # Ajouter une ligne de référence à 0.85 (objectif)
    ax.axhline(y=0.85, color='gold', linestyle='--', linewidth=2, alpha=0.7, label='Objectif (85%)')
    ax.legend(fontsize=11, loc='lower left')
    
    plt.tight_layout()
    output_path = os.path.join(output_dir, 'accuracy_by_complexity.png')
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"✅ Graphique par complexité sauvegardé : {output_path}")
    return output_path


def plot_levenshtein_distribution(metrics_df: pd.DataFrame, output_dir: str = "strategy_1_results"):
    """
    Histogramme de distribution de la distance de Levenshtein
    comparant Ancien vs Nouveau
    """
    fig, axes = plt.subplots(1, 2, figsize=(14, 6))
    
    # Ancien
    axes[0].hist(metrics_df['man_anc_levenshtein'], bins=20, color='#E74C3C', 
                 edgecolor='black', alpha=0.7)
    axes[0].axvline(metrics_df['man_anc_levenshtein'].mean(), color='darkred', 
                    linestyle='--', linewidth=2, label=f"Moyenne: {metrics_df['man_anc_levenshtein'].mean():.1f}")
    axes[0].set_xlabel('Distance de Levenshtein', fontsize=11, weight='bold')
    axes[0].set_ylabel('Nombre de Recettes', fontsize=11, weight='bold')
    axes[0].set_title('LLM Ancien vs Manuel', fontsize=13, weight='bold')
    axes[0].legend(fontsize=10)
    axes[0].grid(axis='y', alpha=0.3)
    
    # Nouveau
    axes[1].hist(metrics_df['man_nouv_levenshtein'], bins=20, color='#2ECC71', 
                 edgecolor='black', alpha=0.7)
    axes[1].axvline(metrics_df['man_nouv_levenshtein'].mean(), color='darkgreen', 
                    linestyle='--', linewidth=2, label=f"Moyenne: {metrics_df['man_nouv_levenshtein'].mean():.1f}")
    axes[1].set_xlabel('Distance de Levenshtein', fontsize=11, weight='bold')
    axes[1].set_ylabel('Nombre de Recettes', fontsize=11, weight='bold')
    axes[1].set_title('LLM Nouveau vs Manuel', fontsize=13, weight='bold')
    axes[1].legend(fontsize=10)
    axes[1].grid(axis='y', alpha=0.3)
    
    fig.suptitle('Distribution de la Distance d\'Édition\n(Plus proche de 0 = Meilleur)', 
                 fontsize=15, weight='bold', y=1.02)
    
    plt.tight_layout()
    output_path = os.path.join(output_dir, 'levenshtein_distribution.png')
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"✅ Graphique de Levenshtein sauvegardé : {output_path}")
    return output_path


def plot_sequence_length_comparison(metrics_df: pd.DataFrame, output_dir: str = "strategy_1_results"):
    """
    Scatter plot comparant les longueurs des séquences
    Manuel vs Ancien et Manuel vs Nouveau
    """
    fig, axes = plt.subplots(1, 2, figsize=(14, 6))
    
    # Manuel vs Ancien
    axes[0].scatter(metrics_df['len_manual'], metrics_df['len_ancien'], 
                    alpha=0.6, s=80, color='#E74C3C', edgecolors='black', linewidth=0.5)
    
    # Ligne d'identité (parfait match)
    max_val = max(metrics_df['len_manual'].max(), metrics_df['len_ancien'].max())
    axes[0].plot([0, max_val], [0, max_val], 'k--', linewidth=2, alpha=0.5, label='Ligne d\'identité')
    
    axes[0].set_xlabel('Longueur Manuel', fontsize=11, weight='bold')
    axes[0].set_ylabel('Longueur LLM Ancien', fontsize=11, weight='bold')
    axes[0].set_title('Comparaison des Longueurs\nManuel vs Ancien', fontsize=13, weight='bold')
    axes[0].legend(fontsize=10)
    axes[0].grid(True, alpha=0.3)
    
    # Manuel vs Nouveau
    axes[1].scatter(metrics_df['len_manual'], metrics_df['len_nouveau'], 
                    alpha=0.6, s=80, color='#2ECC71', edgecolors='black', linewidth=0.5)
    
    # Ligne d'identité
    max_val = max(metrics_df['len_manual'].max(), metrics_df['len_nouveau'].max())
    axes[1].plot([0, max_val], [0, max_val], 'k--', linewidth=2, alpha=0.5, label='Ligne d\'identité')
    
    axes[1].set_xlabel('Longueur Manuel', fontsize=11, weight='bold')
    axes[1].set_ylabel('Longueur LLM Nouveau', fontsize=11, weight='bold')
    axes[1].set_title('Comparaison des Longueurs\nManuel vs Nouveau', fontsize=13, weight='bold')
    axes[1].legend(fontsize=10)
    axes[1].grid(True, alpha=0.3)
    
    fig.suptitle('Analyse des Longueurs de Séquences\n(Points sur la diagonale = longueurs identiques)', 
                 fontsize=15, weight='bold', y=1.02)
    
    plt.tight_layout()
    output_path = os.path.join(output_dir, 'sequence_length_comparison.png')
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"✅ Graphique des longueurs sauvegardé : {output_path}")
    return output_path


def plot_heatmap_metrics_correlation(metrics_df: pd.DataFrame, output_dir: str = "strategy_1_results"):
    """
    Heatmap de corrélation entre les différentes métriques
    pour identifier les relations entre mesures
    """
    fig, ax = plt.subplots(figsize=(12, 10))
    
    # Sélectionner les colonnes de métriques
    metrics_columns = [
        'man_anc_accuracy', 'man_anc_jaccard', 'man_anc_lcs_ratio', 'man_anc_levenshtein',
        'man_nouv_accuracy', 'man_nouv_jaccard', 'man_nouv_lcs_ratio', 'man_nouv_levenshtein',
        'len_manual', 'len_ancien', 'len_nouveau'
    ]
    
    # Calculer la matrice de corrélation
    correlation_matrix = metrics_df[metrics_columns].corr()
    
    # Labels plus lisibles
    labels = [
        'Acc Ancien', 'Jac Ancien', 'LCS Ancien', 'Lev Ancien',
        'Acc Nouveau', 'Jac Nouveau', 'LCS Nouveau', 'Lev Nouveau',
        'Len Manuel', 'Len Ancien', 'Len Nouveau'
    ]
    
    # Créer la heatmap
    sns.heatmap(correlation_matrix, annot=True, fmt='.2f', cmap='coolwarm', 
                center=0, square=True, linewidths=1, cbar_kws={"shrink": 0.8},
                xticklabels=labels, yticklabels=labels, ax=ax)
    
    ax.set_title('Matrice de Corrélation des Métriques\n(Rouge = Corrélation Positive, Bleu = Négative)', 
                 fontsize=14, weight='bold', pad=15)
    
    plt.tight_layout()
    output_path = os.path.join(output_dir, 'metrics_correlation_heatmap.png')
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"✅ Heatmap de corrélation sauvegardée : {output_path}")
    return output_path


def plot_improvement_waterfall(metrics_df: pd.DataFrame, output_dir: str = "strategy_1_results"):
    """
    Graphique en cascade (waterfall) montrant l'évolution des performances
    du LLM Ancien vers le LLM Nouveau
    """
    fig, ax = plt.subplots(figsize=(12, 7))
    
    # Calculer les métriques
    accuracy_ancien = metrics_df['man_anc_accuracy'].mean()
    accuracy_nouveau = metrics_df['man_nouv_accuracy'].mean()
    improvement = accuracy_nouveau - accuracy_ancien
    
    jaccard_ancien = metrics_df['man_anc_jaccard'].mean()
    jaccard_nouveau = metrics_df['man_nouv_jaccard'].mean()
    improvement_jaccard = jaccard_nouveau - jaccard_ancien
    
    lcs_ancien = metrics_df['man_anc_lcs_ratio'].mean()
    lcs_nouveau = metrics_df['man_nouv_lcs_ratio'].mean()
    improvement_lcs = lcs_nouveau - lcs_ancien
    
    # Données pour le waterfall
    categories = ['Accuracy\nAncien', 'Δ Accuracy', 'Accuracy\nNouveau', '', 
                  'Jaccard\nAncien', 'Δ Jaccard', 'Jaccard\nNouveau', '',
                  'LCS Ratio\nAncien', 'Δ LCS', 'LCS Ratio\nNouveau']
    
    values = [accuracy_ancien, improvement, accuracy_nouveau, 0,
              jaccard_ancien, improvement_jaccard, jaccard_nouveau, 0,
              lcs_ancien, improvement_lcs, lcs_nouveau]
    
    colors = ['#3498DB', '#2ECC71' if improvement >= 0 else '#E74C3C', '#3498DB', 'white',
              '#3498DB', '#2ECC71' if improvement_jaccard >= 0 else '#E74C3C', '#3498DB', 'white',
              '#3498DB', '#2ECC71' if improvement_lcs >= 0 else '#E74C3C', '#3498DB']
    
    # Créer le graphique
    bars = ax.bar(range(len(categories)), values, color=colors, edgecolor='black', linewidth=1.5)
    
    # Ajouter les valeurs sur les barres
    for i, (bar, value) in enumerate(zip(bars, values)):
        if value != 0:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                    f'{value:.3f}',
                    ha='center', va='bottom' if value > 0 else 'top', 
                    fontsize=10, weight='bold')
    
    # Personnalisation
    ax.set_xticks(range(len(categories)))
    ax.set_xticklabels(categories, fontsize=10)
    ax.set_ylabel('Score', fontsize=12, weight='bold')
    ax.set_title('Évolution des Performances : LLM Ancien → LLM Nouveau\n(vs Annotations Manuelles)', 
                 fontsize=14, weight='bold', pad=15)
    ax.set_ylim(0, 1.1)
    ax.grid(axis='y', alpha=0.3)
    
    # Ajouter une légende
    from matplotlib.patches import Patch
    legend_elements = [
        Patch(facecolor='#3498DB', edgecolor='black', label='Score de Base'),
        Patch(facecolor='#2ECC71', edgecolor='black', label='Amélioration'),
        Patch(facecolor='#E74C3C', edgecolor='black', label='Régression')
    ]
    ax.legend(handles=legend_elements, loc='upper left', fontsize=10)
    
    plt.tight_layout()
    output_path = os.path.join(output_dir, 'improvement_waterfall.png')
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"✅ Graphique waterfall sauvegardé : {output_path}")
    return output_path


def plot_cuisine_type_performance(metrics_df: pd.DataFrame, output_dir: str = "strategy_1_results"):
    """
    Graphique en barres horizontales montrant la performance par type de cuisine
    """
    fig, ax = plt.subplots(figsize=(12, 8))
    
    # Préparer les données
    cuisine_types = metrics_df['cuisine_type'].unique()
    ancien_scores = []
    nouveau_scores = []
    n_recipes = []
    
    for cuisine in cuisine_types:
        subset = metrics_df[metrics_df['cuisine_type'] == cuisine]
        if len(subset) > 0:
            ancien_scores.append(subset['man_anc_accuracy'].mean())
            nouveau_scores.append(subset['man_nouv_accuracy'].mean())
            n_recipes.append(len(subset))
    
    # Trier par performance du nouveau
    sorted_indices = np.argsort(nouveau_scores)[::-1]
    cuisine_types = [cuisine_types[i] for i in sorted_indices]
    ancien_scores = [ancien_scores[i] for i in sorted_indices]
    nouveau_scores = [nouveau_scores[i] for i in sorted_indices]
    n_recipes = [n_recipes[i] for i in sorted_indices]
    
    # Position des barres
    y = np.arange(len(cuisine_types))
    height = 0.35
    
    # Créer les barres horizontales
    bars1 = ax.barh(y - height/2, ancien_scores, height, label='LLM Ancien', 
                    color='#E74C3C', edgecolor='black', linewidth=1.5)
    bars2 = ax.barh(y + height/2, nouveau_scores, height, label='LLM Nouveau', 
                    color='#2ECC71', edgecolor='black', linewidth=1.5)
    
    # Ajouter les valeurs et le nombre de recettes
    for i, (bar1, bar2, n) in enumerate(zip(bars1, bars2, n_recipes)):
        # Ancien
        width1 = bar1.get_width()
        ax.text(width1, bar1.get_y() + bar1.get_height()/2.,
                f' {width1:.3f}',
                ha='left', va='center', fontsize=9, weight='bold')
        
        # Nouveau
        width2 = bar2.get_width()
        ax.text(width2, bar2.get_y() + bar2.get_height()/2.,
                f' {width2:.3f}',
                ha='left', va='center', fontsize=9, weight='bold')
        
        # Nombre de recettes
        ax.text(-0.02, i, f'(n={n})', ha='right', va='center', fontsize=8, style='italic')
    
    # Personnalisation
    ax.set_yticks(y)
    ax.set_yticklabels(cuisine_types, fontsize=11)
    ax.set_xlabel('Accuracy Moyenne', fontsize=12, weight='bold')
    ax.set_title('Performance par Type de Cuisine\n(vs Annotations Manuelles)', 
                 fontsize=14, weight='bold', pad=15)
    ax.legend(fontsize=11, loc='lower right')
    ax.set_xlim(0, 1.1)
    ax.grid(axis='x', alpha=0.3)
    
    # Ligne de référence objectif
    ax.axvline(x=0.85, color='gold', linestyle='--', linewidth=2, alpha=0.7, label='Objectif (85%)')
    ax.legend(fontsize=11, loc='lower right')
    
    plt.tight_layout()
    output_path = os.path.join(output_dir, 'cuisine_type_performance.png')
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"✅ Graphique par type de cuisine sauvegardé : {output_path}")
    return output_path


def plot_error_analysis_sunburst(metrics_df: pd.DataFrame, output_dir: str = "strategy_1_results"):
    """
    Graphique sunburst (diagramme circulaire hiérarchique) montrant
    la distribution des erreurs par complexité et par catégorie de changement
    """
    try:
        import plotly.graph_objects as go
        
        # Préparer les données hiérarchiques
        labels = ['Tous']
        parents = ['']
        values = [len(metrics_df)]
        colors = ['#34495E']
        
        # Niveau 1: Complexité
        for complexity in ['SIMPLE', 'MOYENNE', 'COMPLEXE']:
            subset = metrics_df[metrics_df['complexity'] == complexity]
            if len(subset) > 0:
                labels.append(complexity)
                parents.append('Tous')
                values.append(len(subset))
                colors.append('#3498DB' if complexity == 'SIMPLE' else '#9B59B6' if complexity == 'MOYENNE' else '#E67E22')
                
                # Niveau 2: Catégories de changement
                for category in ['correction', 'regression', 'stable_correct', 'stable_incorrect', 'changement_lateral']:
                    cat_subset = subset[subset['change_category'] == category]
                    if len(cat_subset) > 0:
                        labels.append(f"{complexity}-{category}")
                        parents.append(complexity)
                        values.append(len(cat_subset))
                        
                        # Couleurs selon la catégorie
                        if category == 'correction' or category == 'stable_correct':
                            colors.append('#2ECC71')
                        elif category == 'regression':
                            colors.append('#E74C3C')
                        elif category == 'stable_incorrect':
                            colors.append('#95A5A6')
                        else:
                            colors.append('#F39C12')
        
        # Créer le graphique sunburst
        fig = go.Figure(go.Sunburst(
            labels=labels,
            parents=parents,
            values=values,
            marker=dict(colors=colors, line=dict(color='white', width=2)),
            textinfo='label+percent parent',
            hovertemplate='<b>%{label}</b><br>Nombre: %{value}<br>Pourcentage: %{percentParent}<extra></extra>'
        ))
        
        fig.update_layout(
            title='Distribution Hiérarchique des Résultats<br>Complexité → Catégorie de Changement',
            title_font_size=16,
            width=900,
            height=900
        )
        
        output_path = os.path.join(output_dir, 'error_analysis_sunburst.html')
        fig.write_html(output_path)
        
        print(f"✅ Graphique sunburst sauvegardé : {output_path}")
        return output_path
        
    except ImportError:
        print("⚠️  Plotly non installé. Installation requise : pip install plotly")
        return None


def generate_all_visualizations(metrics_df: pd.DataFrame, output_dir: str = "strategy_1_results"):
    """
    Génère toutes les visualisations en une seule fois
    
    Args:
        metrics_df: DataFrame des métriques de comparaison
        output_dir: Répertoire de sortie
    
    Returns:
        Liste des chemins des fichiers générés
    """
    print(f"\n{'='*80}")
    print(f"GÉNÉRATION DE TOUTES LES VISUALISATIONS")
    print(f"{'='*80}\n")
    
    output_paths = []
    
    # 1. Radar chart
    output_paths.append(plot_metrics_comparison_radar(metrics_df, output_dir))
    
    # 2. Distribution des changements
    output_paths.append(plot_change_categories_distribution(metrics_df, output_dir))
    
    # 3. Performance par complexité
    output_paths.append(plot_accuracy_by_complexity(metrics_df, output_dir))
    
    # 4. Distribution Levenshtein
    output_paths.append(plot_levenshtein_distribution(metrics_df, output_dir))
    
    # 5. Comparaison des longueurs
    output_paths.append(plot_sequence_length_comparison(metrics_df, output_dir))
    
    # 6. Heatmap de corrélation
    output_paths.append(plot_heatmap_metrics_correlation(metrics_df, output_dir))
    
    # 7. Waterfall d'amélioration
    output_paths.append(plot_improvement_waterfall(metrics_df, output_dir))
    
    # 8. Performance par type de cuisine
    output_paths.append(plot_cuisine_type_performance(metrics_df, output_dir))
    
    # 9. Sunburst (si plotly disponible)
    sunburst_path = plot_error_analysis_sunburst(metrics_df, output_dir)
    if sunburst_path:
        output_paths.append(sunburst_path)
    
    print(f"\n{'='*80}")
    print(f"✅ TOUTES LES VISUALISATIONS GÉNÉRÉES")
    print(f"{'='*80}")
    print(f"Total de graphiques créés : {len(output_paths)}")
    print(f"Répertoire : {output_dir}/")
    
    return output_paths




import pandas as pd
import json
import re
import time
import os
from typing import List, Dict, Tuple
from datetime import datetime
from openai import OpenAI
import numpy as np
import ast
from collections import defaultdict


""" **************************************************************************************************************************

Pipeline complet pour la Stratégie 2 : Validation Structurelle des Graphes
Vérification automatique de la cohérence structurelle des graphes de recettes

VERSION AMÉLIORÉE avec :
- Classification des recettes par category et complexity
- Statistiques Test 1 enrichies (par variante, category, complexity)
- Outputs simplifiés : rapport.md + dataset_synthese.csv + dataset_resume.csv

    **************************************************************************************************************************
""" 


# ==============================================================================
# SECTION 1 : FONCTIONS DE CLASSIFICATION
# ==============================================================================

def classify_cuisine_type(title: str) -> str:
    """
    Classifie le type de cuisine d'une recette uniquement à partir du titre
    
    Args:
        title: Titre de la recette
    
    Returns:
        'bakery', 'stew', 'quick_prep', ou 'other'
    """
    if pd.isna(title):
        return 'other'
    
    title_lower = str(title).lower()
    
    # Mots-clés pour la pâtisserie/boulangerie
    bakery_keywords = [
        # Gâteaux et desserts
        'cake', 'cupcake', 'cheesecake', 'torte', 'gateau',
        # Cookies et biscuits
        'cookie', 'biscuit', 'brownie', 'bar', 'square',
        # Pains et viennoiseries
        'bread', 'bun', 'roll', 'loaf', 'baguette', 'croissant', 'danish', 'scone', 'muffin',
        # Tartes et pâtisseries
        'pie', 'tart', 'pastry', 'puff', 'strudel', 'turnover', 'cobbler', 'crisp', 'crumble',
        # Autres desserts au four
        'pudding', 'souffle', 'meringue', 'macaron', 'eclair', 'profiterole',
        # Ingrédients typiques
        'chocolate', 'vanilla', 'cinnamon', 'frosting', 'icing'
    ]
    
    # Mots-clés pour plats mijotés/cuisinés lentement
    stew_keywords = [
        # Types de plats
        'stew', 'soup', 'broth', 'chowder', 'bisque', 'consommé', 'gumbo', 'chili',
        'casserole', 'pot roast', 'ragout', 'tagine',
        # Techniques de cuisson lente
        'braised', 'slow cooked', 'slow-cooked', 'crock pot', 'crockpot',
        # Plats typiques
        'curry', 'goulash', 'bourguignon', 'stroganoff', 'fricassee',
        # Indicateurs
        'slow cooker', 'dutch oven'
    ]
    
    # Mots-clés pour préparations rapides
    quick_prep_keywords = [
        # Salades
        'salad', 'slaw', 'coleslaw', 'caesar', 'greek salad', 'caprese',
        # Sandwichs et wraps
        'sandwich', 'panini', 'sub', 'hoagie', 'wrap', 'burrito', 'quesadilla', 'taco',
        # Bols et assemblages
        'bowl', 'poke', 'buddha bowl', 'grain bowl', 'rice bowl', 'noodle bowl',
        # Smoothies et boissons
        'smoothie', 'shake', 'juice', 'drink', 'beverage', 'cocktail', 'mocktail',
        # Snacks et apéritifs
        'dip', 'spread', 'hummus', 'guacamole', 'salsa', 'bruschetta', 'crostini',
        'appetizer', 'finger food',
        # Préparations crues/froides
        'no-bake', 'no bake', 'raw', 'fresh', 'cold', 'chilled',
        'ceviche', 'tartare', 'carpaccio',
        # Indicateurs de rapidité
        'quick', 'easy', 'simple', '5 minute', '10 minute', '15 minute', 'instant',
        '5-minute', '10-minute', '15-minute'
    ]
    
    # Compter les occurrences de chaque catégorie
    bakery_score = sum(1 for keyword in bakery_keywords if keyword in title_lower)
    stew_score = sum(1 for keyword in stew_keywords if keyword in title_lower)
    quick_score = sum(1 for keyword in quick_prep_keywords if keyword in title_lower)
    
    # Retourner la catégorie avec le score le plus élevé
    scores = {
        'bakery': bakery_score,
        'stew': stew_score,
        'quick_prep': quick_score
    }
    
    max_score = max(scores.values())
    
    # Si aucun mot-clé trouvé, retourner 'other'
    if max_score == 0:
        return 'other'
    
    # Retourner la catégorie avec le score maximum
    for category, score in scores.items():
        if score == max_score:
            return category
    
    return 'other'


def classify_complexity(steps) -> str:
    """
    Classifie la complexité d'une recette selon le nombre d'étapes
    
    Args:
        steps: Nombre d'étapes (number_of_steps)
    
    Returns:
        'simple', 'moyenne', 'elevee', ou 'unknown'
    """
    if pd.isna(steps):
        return 'unknown'
    
    try:
        steps = int(steps)
        if steps <= 10:
            return 'simple'
        elif steps <= 20:
            return 'moyenne'
        else:
            return 'elevee'
    except (ValueError, TypeError):
        return 'unknown'


# ==============================================================================
# SECTION 2 : UTILITAIRES ET HELPERS
# ==============================================================================

def parse_actions_column(actions_value):
    """
    Parse robustement la colonne 'actions' qui peut être sous différents formats
    
    Args:
        actions_value: Valeur de la colonne actions (string, list, ou autre)
    
    Returns:
        Liste d'actions
    """
    if pd.isna(actions_value):
        return []
    
    # Si c'est déjà une liste
    if isinstance(actions_value, list):
        return actions_value
    
    # Si c'est une string représentant une liste
    if isinstance(actions_value, str):
        try:
            # Essayer de parser comme du JSON/Python
            parsed = ast.literal_eval(actions_value)
            if isinstance(parsed, list):
                return parsed
        except (ValueError, SyntaxError):
            # Si échec, essayer de split par virgules
            if ',' in actions_value:
                return [a.strip() for a in actions_value.split(',') if a.strip()]
            # Sinon retourner comme élément unique
            return [actions_value.strip()] if actions_value.strip() else []
    
    return []


def create_output_directory(output_dir: str):
    """Crée le répertoire de sortie s'il n'existe pas"""
    os.makedirs(output_dir, exist_ok=True)


# ==============================================================================
# SECTION 3 : TEST 1 - CALCUL DE LA TAILLE DES LISTES D'ACTIONS (AMÉLIORÉ)
# ==============================================================================

def test_1_calculate_action_lengths(
    graphs_df: pd.DataFrame,
    recipes_df: pd.DataFrame
) -> Tuple[Dict, pd.DataFrame]:
    """
    Test 1: Calcule les statistiques de taille des listes d'actions
    AMÉLIORÉ avec statistiques complètes par variante, category et complexity
    
    Pour chaque niveau (global, par variante, par category, par complexity):
    - Longueur moyenne, médiane, écart-type, min, max, Q1, Q3
    - Outliers détectés (count et pourcentage)
    
    Args:
        graphs_df: DataFrame des graphes (colonnes: id, title, actions, type, type_2)
        recipes_df: DataFrame des recettes avec category et complexity
    
    Returns:
        Tuple (stats_dict, dataset_resume_df)
    """
    print(f"\n{'='*80}")
    print(f"TEST 1 : CALCUL DE LA TAILLE DES LISTES D'ACTIONS")
    print(f"{'='*80}\n")
    
    # Parser les actions
    graphs_df = graphs_df.copy()
    graphs_df['actions_parsed'] = graphs_df['actions'].apply(parse_actions_column)
    graphs_df['action_length'] = graphs_df['actions_parsed'].apply(len)
    
    # Merger avec recipes pour obtenir category et complexity
    graphs_df = graphs_df.merge(
        recipes_df[['id', 'category', 'complexity']], 
        on='id', 
        how='left'
    )
    
    def calculate_detailed_stats(subset_df, subset_name=""):
        """
        Calcule les statistiques complètes pour un sous-ensemble de données
        
        Args:
            subset_df: DataFrame filtré
            subset_name: Nom du sous-ensemble pour l'affichage
        
        Returns:
            Dictionnaire de statistiques
        """
        if len(subset_df) == 0:
            return None
        
        # Statistiques de base
        stats = {
            'count': int(len(subset_df)),
            'longueur_min': int(subset_df['action_length'].min()),
            'longueur_max': int(subset_df['action_length'].max()),
            'longueur_moyenne': float(subset_df['action_length'].mean()),
            'longueur_mediane': float(subset_df['action_length'].median()),
            'ecart_type': float(subset_df['action_length'].std()),
            'q1': float(subset_df['action_length'].quantile(0.25)),
            'q3': float(subset_df['action_length'].quantile(0.75))
        }
        
        # Calcul des outliers (> Q3 + 3*IQR ou < Q1 - 3*IQR)
        iqr = stats['q3'] - stats['q1']
        upper_bound = stats['q3'] + 3 * iqr
        lower_bound = stats['q1'] - 3 * iqr
        
        outliers_mask = (subset_df['action_length'] > upper_bound) | (subset_df['action_length'] < lower_bound)
        outliers_count = outliers_mask.sum()
        
        stats['outliers_count'] = int(outliers_count)
        stats['outliers_percentage'] = float(outliers_count / len(subset_df) * 100) if len(subset_df) > 0 else 0.0
        stats['upper_bound'] = float(upper_bound)
        stats['lower_bound'] = float(lower_bound)
        
        return stats
    
    # ========== STATISTIQUES GLOBALES ==========
    print("📊 STATISTIQUES GLOBALES")
    print("-" * 80)
    
    stats_globales = calculate_detailed_stats(graphs_df, "GLOBAL")
    
    print(f"Total graphes: {stats_globales['count']:,}")
    print(f"Longueur moyenne: {stats_globales['longueur_moyenne']:.2f} ± {stats_globales['ecart_type']:.2f} actions")
    print(f"Médiane: {stats_globales['longueur_mediane']:.1f}")
    print(f"Min: {stats_globales['longueur_min']}, Max: {stats_globales['longueur_max']}")
    print(f"Quartiles: Q1={stats_globales['q1']:.1f}, Q3={stats_globales['q3']:.1f}")
    print(f"Outliers détectés: {stats_globales['outliers_count']:,} ({stats_globales['outliers_percentage']:.2f}%)")
    print(f"  Bornes: [{stats_globales['lower_bound']:.1f}, {stats_globales['upper_bound']:.1f}]")
    
    # ========== STATISTIQUES PAR VARIANTE ==========
    print(f"\n📊 STATISTIQUES PAR VARIANTE")
    print("-" * 80)
    
    stats_by_variant = {}
    for variant_type in sorted(graphs_df['type_2'].dropna().unique()):
        subset = graphs_df[graphs_df['type_2'] == variant_type]
        stats = calculate_detailed_stats(subset, variant_type)
        
        if stats:
            stats_by_variant[variant_type] = stats
            print(f"\n  📌 {variant_type.upper()}")
            print(f"     Count: {stats['count']:,}")
            print(f"     Moyenne: {stats['longueur_moyenne']:.2f} ± {stats['ecart_type']:.2f}")
            print(f"     Médiane: {stats['longueur_mediane']:.1f}")
            print(f"     Min-Max: [{stats['longueur_min']}, {stats['longueur_max']}]")
            print(f"     Quartiles: Q1={stats['q1']:.1f}, Q3={stats['q3']:.1f}")
            print(f"     Outliers: {stats['outliers_count']:,} ({stats['outliers_percentage']:.2f}%)")
            print(f"     Bornes: [{stats['lower_bound']:.1f}, {stats['upper_bound']:.1f}]")
    
    # ========== STATISTIQUES PAR CATEGORY ==========
    print(f"\n📊 STATISTIQUES PAR CATÉGORIE DE CUISINE")
    print("-" * 80)
    
    stats_by_category = {}
    for category in sorted(graphs_df['category'].dropna().unique()):
        subset = graphs_df[graphs_df['category'] == category]
        stats = calculate_detailed_stats(subset, category)
        
        if stats:
            stats_by_category[category] = stats
            print(f"\n  📌 {category.upper()}")
            print(f"     Count: {stats['count']:,}")
            print(f"     Moyenne: {stats['longueur_moyenne']:.2f} ± {stats['ecart_type']:.2f}")
            print(f"     Médiane: {stats['longueur_mediane']:.1f}")
            print(f"     Min-Max: [{stats['longueur_min']}, {stats['longueur_max']}]")
            print(f"     Quartiles: Q1={stats['q1']:.1f}, Q3={stats['q3']:.1f}")
            print(f"     Outliers: {stats['outliers_count']:,} ({stats['outliers_percentage']:.2f}%)")
            print(f"     Bornes: [{stats['lower_bound']:.1f}, {stats['upper_bound']:.1f}]")
    
    # ========== STATISTIQUES PAR COMPLEXITY ==========
    print(f"\n📊 STATISTIQUES PAR NIVEAU DE COMPLEXITÉ")
    print("-" * 80)
    
    stats_by_complexity = {}
    complexity_order = ['simple', 'moyenne', 'elevee', 'unknown']
    for complexity in complexity_order:
        subset = graphs_df[graphs_df['complexity'] == complexity]
        if len(subset) > 0:
            stats = calculate_detailed_stats(subset, complexity)
            
            if stats:
                stats_by_complexity[complexity] = stats
                print(f"\n  📌 {complexity.upper()}")
                print(f"     Count: {stats['count']:,}")
                print(f"     Moyenne: {stats['longueur_moyenne']:.2f} ± {stats['ecart_type']:.2f}")
                print(f"     Médiane: {stats['longueur_mediane']:.1f}")
                print(f"     Min-Max: [{stats['longueur_min']}, {stats['longueur_max']}]")
                print(f"     Quartiles: Q1={stats['q1']:.1f}, Q3={stats['q3']:.1f}")
                print(f"     Outliers: {stats['outliers_count']:,} ({stats['outliers_percentage']:.2f}%)")
                print(f"     Bornes: [{stats['lower_bound']:.1f}, {stats['upper_bound']:.1f}]")
    
    print(f"\n✅ Test 1 terminé avec succès")
    
    # Compiler toutes les stats
    all_stats = {
        'globales': stats_globales,
        'par_variante': stats_by_variant,
        'par_category': stats_by_category,
        'par_complexity': stats_by_complexity
    }
    
    # Créer le dataset_resume pour ce dataset (sera complété par les autres datasets)
    dataset_resume_partial = graphs_df[['id', 'action_length', 'type_2']].copy()
    dataset_resume_partial.columns = ['id', 'nombre_actions', 'variante']
    
    return all_stats, dataset_resume_partial

# ==============================================================================
# SECTION 4 : TEST 2 - VARIANTE PRINCIPALE VS NOMBRE D'INSTRUCTIONS
# ==============================================================================

def test_2_validate_principale_vs_steps(
    recipes_df: pd.DataFrame,
    graphs_df: pd.DataFrame
) -> Tuple[Dict, pd.DataFrame]:
    """
    Test 2: Vérifie la cohérence entre le nombre d'actions de la variante principale
    et le nombre d'instructions de la recette
    
    Args:
        recipes_df: DataFrame des recettes
        graphs_df: DataFrame des graphes
    
    Returns:
        Tuple (stats_dict, flags_df)
    """
    print(f"\n{'='*80}")
    print(f"TEST 2 : COMPARAISON VARIANTE PRINCIPALE VS NOMBRE D'INSTRUCTIONS")
    print(f"{'='*80}\n")
    
    # Filtrer uniquement les variantes principales
    principales = graphs_df[graphs_df['type_2'] == 'variante_principale'].copy()
    
    # Parser les actions
    principales['actions_parsed'] = principales['actions'].apply(parse_actions_column)
    principales['action_length'] = principales['actions_parsed'].apply(len)
    
    # Merger avec recipes
    merged = principales.merge(
        recipes_df[['id', 'number_of_steps']], 
        on='id', 
        how='left'
    )
    
    # Calculer le ratio
    merged['ratio'] = merged['action_length'] / merged['number_of_steps'].replace(0, np.nan)
    
    # Définir les seuils de validation
    RATIO_MIN = 0.8
    RATIO_MAX = 2.5
    
    # Identifier les flags
    merged['flag'] = 'CONFORME'
    merged.loc[merged['ratio'] < RATIO_MIN, 'flag'] = 'FLAG_RATIO_FAIBLE'
    merged.loc[merged['ratio'] > RATIO_MAX, 'flag'] = 'FLAG_RATIO_ELEVE'
    merged.loc[merged['ratio'].isna(), 'flag'] = 'FLAG_CRITIQUE'  # number_of_steps = 0 ou NaN
    
    # Statistiques
    stats = {
        'total_principales': int(len(merged)),
        'ratio_moyen': float(merged['ratio'].mean()),
        'ratio_median': float(merged['ratio'].median()),
        'ratio_std': float(merged['ratio'].std()),
        'CONFORME_count': int((merged['flag'] == 'CONFORME').sum()),
        'CONFORME_pct': float((merged['flag'] == 'CONFORME').sum() / len(merged) * 100) if len(merged) > 0 else 0.0,
        'FLAG_RATIO_FAIBLE_count': int((merged['flag'] == 'FLAG_RATIO_FAIBLE').sum()),
        'FLAG_RATIO_FAIBLE_pct': float((merged['flag'] == 'FLAG_RATIO_FAIBLE').sum() / len(merged) * 100) if len(merged) > 0 else 0.0,
        'FLAG_RATIO_ELEVE_count': int((merged['flag'] == 'FLAG_RATIO_ELEVE').sum()),
        'FLAG_RATIO_ELEVE_pct': float((merged['flag'] == 'FLAG_RATIO_ELEVE').sum() / len(merged) * 100) if len(merged) > 0 else 0.0,
        'FLAG_CRITIQUE_count': int((merged['flag'] == 'FLAG_CRITIQUE').sum()),
        'FLAG_CRITIQUE_pct': float((merged['flag'] == 'FLAG_CRITIQUE').sum() / len(merged) * 100) if len(merged) > 0 else 0.0
    }
    
    print(f"Total variantes principales: {stats['total_principales']:,}")
    print(f"Ratio moyen (actions/instructions): {stats['ratio_moyen']:.2f}")
    print(f"Ratio médian: {stats['ratio_median']:.2f}")
    print(f"\nDistribution des flags:")
    print(f"  ✅ CONFORME: {stats['CONFORME_count']:,} ({stats['CONFORME_pct']:.2f}%)")
    print(f"  ⚠️  FLAG_RATIO_FAIBLE: {stats['FLAG_RATIO_FAIBLE_count']:,} ({stats['FLAG_RATIO_FAIBLE_pct']:.2f}%)")
    print(f"  ⚠️  FLAG_RATIO_ELEVE: {stats['FLAG_RATIO_ELEVE_count']:,} ({stats['FLAG_RATIO_ELEVE_pct']:.2f}%)")
    print(f"  ❌ FLAG_CRITIQUE: {stats['FLAG_CRITIQUE_count']:,} ({stats['FLAG_CRITIQUE_pct']:.2f}%)")
    
    print(f"\n✅ Test 2 terminé")
    
    # Préparer le dataframe des flags pour dataset_synthese
    flags_df = merged[merged['flag'] != 'CONFORME'][['id', 'flag']].copy()
    flags_df['test'] = '2'
    flags_df['test_name'] = 'Comparaison VARIANTE PRINCIPALE VS NOMBRE D\'INSTRUCTIONS'
    flags_df['delta_distribution'] = np.nan
    
    return stats, flags_df


# ==============================================================================
# SECTION 5 : TEST 3 - VARIANTE INGREDIENTS VS TAILLE THÉORIQUE
# ==============================================================================

def test_3_validate_ingredients_variant(
    recipes_df: pd.DataFrame,
    graphs_df: pd.DataFrame
) -> Tuple[Dict, pd.DataFrame]:
    """
    Test 3: Vérifie que la variante ingredients a bien ajouté des gestes
    par rapport à la variante principale
    
    Args:
        recipes_df: DataFrame des recettes
        graphs_df: DataFrame des graphes
    
    Returns:
        Tuple (stats_dict, flags_df)
    """
    print(f"\n{'='*80}")
    print(f"TEST 3 : VALIDATION VARIANTE INGREDIENTS")
    print(f"{'='*80}\n")
    
    # Extraire principales et ingredients
    principales = graphs_df[graphs_df['type_2'] == 'variante_principale'].copy()
    ingredients = graphs_df[graphs_df['type_2'] == 'variante_ingredients'].copy()
    
    # Parser les actions
    principales['actions_parsed'] = principales['actions'].apply(parse_actions_column)
    principales['len_principale'] = principales['actions_parsed'].apply(len)
    
    ingredients['actions_parsed'] = ingredients['actions'].apply(parse_actions_column)
    ingredients['len_ingredient'] = ingredients['actions_parsed'].apply(len)
    
    # Merger
    merged = principales[['id', 'len_principale']].merge(
        ingredients[['id', 'len_ingredient']],
        on='id',
        how='inner'
    )
    
    # Calculer le delta
    merged['delta'] = merged['len_ingredient'] - merged['len_principale']
    merged['delta_distribution'] = merged['delta']
    
    # Flags
    merged['flag'] = 'CONFORME'
    merged.loc[merged['delta'] == 0, 'flag'] = 'FLAG_AUCUN_AJOUT'
    merged.loc[merged['delta'] > 10, 'flag'] = 'FLAG_TROP_AJOUT'
    merged.loc[merged['delta'] < 0, 'flag'] = 'FLAG_CRITIQUE_NEGATIF'
    
    # Statistiques
    stats = {
        'total_paires': int(len(merged)),
        'delta_moyen': float(merged['delta'].mean()),
        'delta_median': float(merged['delta'].median()),
        'delta_std': float(merged['delta'].std()),
        'delta_min': int(merged['delta'].min()),
        'delta_max': int(merged['delta'].max()),
        'CONFORME_count': int((merged['flag'] == 'CONFORME').sum()),
        'CONFORME_pct': float((merged['flag'] == 'CONFORME').sum() / len(merged) * 100) if len(merged) > 0 else 0.0,
        'FLAG_AUCUN_AJOUT_count': int((merged['flag'] == 'FLAG_AUCUN_AJOUT').sum()),
        'FLAG_AUCUN_AJOUT_pct': float((merged['flag'] == 'FLAG_AUCUN_AJOUT').sum() / len(merged) * 100) if len(merged) > 0 else 0.0,
        'FLAG_TROP_AJOUT_count': int((merged['flag'] == 'FLAG_TROP_AJOUT').sum()),
        'FLAG_TROP_AJOUT_pct': float((merged['flag'] == 'FLAG_TROP_AJOUT').sum() / len(merged) * 100) if len(merged) > 0 else 0.0,
        'FLAG_CRITIQUE_NEGATIF_count': int((merged['flag'] == 'FLAG_CRITIQUE_NEGATIF').sum()),
        'FLAG_CRITIQUE_NEGATIF_pct': float((merged['flag'] == 'FLAG_CRITIQUE_NEGATIF').sum() / len(merged) * 100) if len(merged) > 0 else 0.0
    }
    
    print(f"Total paires principale-ingredient: {stats['total_paires']:,}")
    print(f"Delta moyen (ingredient - principale): {stats['delta_moyen']:.2f} actions")
    print(f"Delta médian: {stats['delta_median']:.1f}")
    print(f"Delta min: {stats['delta_min']}, max: {stats['delta_max']}")
    print(f"\nDistribution des flags:")
    print(f"  ✅ CONFORME (1-10 ajouts): {stats['CONFORME_count']:,} ({stats['CONFORME_pct']:.2f}%)")
    print(f"  ⚠️  FLAG_AUCUN_AJOUT: {stats['FLAG_AUCUN_AJOUT_count']:,} ({stats['FLAG_AUCUN_AJOUT_pct']:.2f}%)")
    print(f"  ⚠️  FLAG_TROP_AJOUT (>10): {stats['FLAG_TROP_AJOUT_count']:,} ({stats['FLAG_TROP_AJOUT_pct']:.2f}%)")
    print(f"  ❌ FLAG_CRITIQUE_NEGATIF: {stats['FLAG_CRITIQUE_NEGATIF_count']:,} ({stats['FLAG_CRITIQUE_NEGATIF_pct']:.2f}%)")
    
    print(f"\n✅ Test 3 terminé")
    
    # Préparer le dataframe des flags
    flags_df = merged[merged['flag'] != 'CONFORME'][['id', 'flag', 'delta_distribution']].copy()
    flags_df['test'] = '3'
    flags_df['test_name'] = 'Validation VARIANTE INGREDIENTS'
    
    return stats, flags_df

# ==============================================================================
# SECTION 6 : TEST 4A - PERMUTATIONS VS PRINCIPALE
# ==============================================================================
from difflib import SequenceMatcher
import ast


def parse_actions_column(actions_value):
    """
    Parse robustement la colonne 'actions'
    """
    if pd.isna(actions_value):
        return []
    
    if isinstance(actions_value, list):
        return actions_value
    
    if isinstance(actions_value, str):
        try:
            parsed = ast.literal_eval(actions_value)
            if isinstance(parsed, list):
                return parsed
        except (ValueError, SyntaxError):
            if ',' in actions_value:
                return [a.strip() for a in actions_value.split(',') if a.strip()]
            return [actions_value.strip()] if actions_value.strip() else []
    
    return []


def test_4a_validate_permutations(
    graphs_df: pd.DataFrame
) -> Tuple[Dict, pd.DataFrame]:
    """
    Test 4A CORRIGÉ : Vérifie que les variantes permutations sont valides
    
    Une permutation valide doit :
    - Avoir les MÊMES éléments que la principale (overlap = 1.0)
    - Avoir un ORDRE DIFFÉRENT (levenshtein > 0)
    
    Flags :
    - CONFORME : vraie permutation (mêmes éléments, ordre différent)
    - FLAG_AUCUNE_PERMUTATION : liste identique (pas de permutation générée)
    - FLAG_ELEMENTS_DIFFERENTS : éléments différents (bug dans le pipeline)
    
    Args:
        graphs_df: DataFrame des graphes
    
    Returns:
        Tuple (stats_dict, flags_df)
    """
    print(f"\n{'='*80}")
    print(f"TEST 4A : COMPARAISON PERMUTATIONS VS PRINCIPALE")
    print(f"{'='*80}\n")
    
    def calculate_overlap(list1, list2):
        """Calcule l'overlap (Jaccard similarity) sur les ensembles"""
        if len(list1) == 0 and len(list2) == 0:
            return 1.0
        set1, set2 = set(list1), set(list2)
        intersection = len(set1 & set2)
        union = len(set1 | set2)
        return intersection / union if union > 0 else 0.0
    
    def calculate_levenshtein(list1, list2):
        """
        Calcule la distance de Levenshtein normalisée entre deux listes
        Retourne 0 si les listes sont identiques (même ordre)
        """
        if list1 == list2:
            return 0
        
        if len(list1) == 0:
            return len(list2)
        if len(list2) == 0:
            return len(list1)
        
        # Utiliser SequenceMatcher pour calcul rapide
        ratio = SequenceMatcher(None, list1, list2).ratio()
        return int((1 - ratio) * max(len(list1), len(list2)))
    
    def lists_are_identical(list1, list2):
        """Vérifie si deux listes sont exactement identiques (même ordre)"""
        return list1 == list2
    
    # Extraire principales et permutations
    principales = graphs_df[graphs_df['type_2'] == 'variante_principale'].copy()
    permutations = graphs_df[graphs_df['type_2'] == 'variante_permutation'].copy()
    
    print(f"Variantes principales : {len(principales):,}")
    print(f"Variantes permutations : {len(permutations):,}")
    
    # Parser les actions
    principales['actions_parsed'] = principales['actions'].apply(parse_actions_column)
    permutations['actions_parsed'] = permutations['actions'].apply(parse_actions_column)
    
    # Merger
    merged = permutations[['id', 'type_2', 'actions_parsed']].merge(
        principales[['id', 'actions_parsed']],
        on='id',
        how='inner',
        suffixes=('_perm', '_princ')
    )
    
    print(f"Paires permutation-principale : {len(merged):,}")
    
    # Calculer les métriques
    print("\nCalcul des métriques...")
    
    merged['overlap'] = merged.apply(
        lambda row: calculate_overlap(row['actions_parsed_perm'], row['actions_parsed_princ']),
        axis=1
    )
    
    merged['levenshtein'] = merged.apply(
        lambda row: calculate_levenshtein(row['actions_parsed_perm'], row['actions_parsed_princ']),
        axis=1
    )
    
    merged['is_identical'] = merged.apply(
        lambda row: lists_are_identical(row['actions_parsed_perm'], row['actions_parsed_princ']),
        axis=1
    )
    
    # ========== NOUVELLE LOGIQUE DE FLAGS ==========
    merged['flag'] = 'CONFORME'
    
    # FLAG 1 : Éléments différents → bug dans le pipeline de génération
    merged.loc[merged['overlap'] < 1.0, 'flag'] = 'FLAG_ELEMENTS_DIFFERENTS'
    
    # FLAG 2 : Liste identique (même ordre) → pas de permutation générée
    merged.loc[
        (merged['overlap'] == 1.0) & (merged['is_identical'] == True),
        'flag'
    ] = 'FLAG_AUCUNE_PERMUTATION'
    
    # ========== STATISTIQUES (COMPATIBLE AVEC LE RAPPORT) ==========
    total = len(merged)
    
    conforme_count = int((merged['flag'] == 'CONFORME').sum())
    aucune_perm_count = int((merged['flag'] == 'FLAG_AUCUNE_PERMUTATION').sum())
    elements_diff_count = int((merged['flag'] == 'FLAG_ELEMENTS_DIFFERENTS').sum())
    
    stats = {
        'total_permutations': total,
        'overlap_moyen': float(merged['overlap'].mean()),
        'overlap_median': float(merged['overlap'].median()),
        'levenshtein_moyen': float(merged['levenshtein'].mean()),
        'levenshtein_median': float(merged['levenshtein'].median()),
        
        # Clés COMPATIBLES avec generate_strategy_2_report_markdown
        'CONFORME_count': conforme_count,
        'conformes_pct': float(conforme_count / total * 100) if total > 0 else 0.0,
        
        # Ancien nom: FLAG_SIMILARITE_BASSE → Nouveau: FLAG_ELEMENTS_DIFFERENTS
        'FLAG_SIMILARITE_BASSE_count': elements_diff_count,
        'FLAG_SIMILARITE_BASSE_pct': float(elements_diff_count / total * 100) if total > 0 else 0.0,
        
        # Ancien nom: FLAG_IDENTIQUE → Nouveau: FLAG_AUCUNE_PERMUTATION
        'FLAG_IDENTIQUE_count': aucune_perm_count,
        'FLAG_IDENTIQUE_pct': float(aucune_perm_count / total * 100) if total > 0 else 0.0,
    }
    
    # ========== AFFICHAGE ==========
    print(f"\n{'='*60}")
    print(f"RÉSULTATS")
    print(f"{'='*60}")
    print(f"Total permutations analysées : {stats['total_permutations']:,}")
    print(f"Overlap moyen : {stats['overlap_moyen']:.3f}")
    print(f"Distance Levenshtein moyenne : {stats['levenshtein_moyen']:.2f}")
    
    print(f"\n📊 Distribution des flags :")
    print(f"  ✅ CONFORME (vraie permutation) : {conforme_count:,} ({stats['conformes_pct']:.2f}%)")
    print(f"  ⚠️  FLAG_AUCUNE_PERMUTATION (liste identique) : {aucune_perm_count:,} ({stats['FLAG_IDENTIQUE_pct']:.2f}%)")
    print(f"  ❌ FLAG_ELEMENTS_DIFFERENTS (éléments différents) : {elements_diff_count:,} ({stats['FLAG_SIMILARITE_BASSE_pct']:.2f}%)")
    
    print(f"\n✅ Test 4A terminé")
    
    # ========== PRÉPARATION DU DATAFRAME DES FLAGS ==========
    # Mapper les nouveaux noms vers les anciens pour compatibilité avec dataset_synthese
    flag_mapping = {
        'FLAG_AUCUNE_PERMUTATION': 'FLAG_IDENTIQUE',
        'FLAG_ELEMENTS_DIFFERENTS': 'FLAG_SIMILARITE_BASSE'
    }
    
    flags_df = merged[merged['flag'] != 'CONFORME'][['id', 'flag']].copy()
    flags_df['flag'] = flags_df['flag'].map(lambda x: flag_mapping.get(x, x))
    flags_df['test'] = '4A'
    flags_df['test_name'] = 'Comparaison PERMUTATIONS VS PRINCIPALE'
    flags_df['delta_distribution'] = np.nan
    
    return stats, flags_df



# ==============================================================================
# SECTION 7 : TEST 4B - VARIANTE INGREDIENTS VS PRINCIPALE (SIMILARITÉ)
# ==============================================================================

def test_4b_validate_ingredients_similarity(
    graphs_df: pd.DataFrame
) -> Tuple[Dict, pd.DataFrame]:
    """
    Test 4B: Vérifie que la variante ingredients contient bien la principale + gestes
    
    Args:
        graphs_df: DataFrame des graphes
    
    Returns:
        Tuple (stats_dict, flags_df)
    """
    print(f"\n{'='*80}")
    print(f"TEST 4B : VARIANTE INGREDIENTS VS PRINCIPALE (SIMILARITÉ)")
    print(f"{'='*80}\n")
    
    def calculate_overlap(list1, list2):
        """Calcule l'overlap (Jaccard similarity)"""
        if len(list1) == 0 and len(list2) == 0:
            return 1.0
        set1, set2 = set(list1), set(list2)
        intersection = len(set1 & set2)
        union = len(set1 | set2)
        return intersection / union if union > 0 else 0.0
    
    def flatten_actions(actions):
        """
        Aplati une liste d'actions potentiellement imbriquées.
        Exemple :
            ['mix', ['chop', 'onion']] → ['mix', 'chop onion']
        """
        flattened = []
        for a in actions:
            if isinstance(a, list):
                flattened.append(" ".join(str(x) for x in a))
            else:
                flattened.append(str(a))
        return flattened
    
    def count_added_gestures(ingredient_actions, principale_actions):
        """Compte les gestes ajoutés"""
        set_ingr = set(ingredient_actions)
        set_princ = set(principale_actions)
        return len(set_ingr - set_princ)
    
    # Extraire principales et ingredients
    principales = graphs_df[graphs_df['type_2'] == 'variante_principale'].copy()
    ingredients = graphs_df[graphs_df['type_2'] == 'variante_ingredients'].copy()
    
    # Parser les actions
    principales['actions_parsed'] = principales['actions'].apply(parse_actions_column)
    ingredients['actions_parsed'] = ingredients['actions'].apply(parse_actions_column)
    
    # Merger
    merged = ingredients[['id', 'actions_parsed']].merge(
        principales[['id', 'actions_parsed']],
        on='id',
        how='inner',
        suffixes=('_ingr', '_princ')
    )
    
    # Calculer les métriques
    merged['overlap'] = merged.apply(
        lambda row: calculate_overlap(flatten_actions(row['actions_parsed_ingr']), flatten_actions(row['actions_parsed_princ'])),
        axis=1
    )
    merged['gestes_ajoutes'] = merged.apply(
        lambda row: count_added_gestures(flatten_actions(row['actions_parsed_ingr']), flatten_actions(row['actions_parsed_princ'])),
        axis=1
    )
    
    # Flags
    merged['flag'] = 'CONFORME'
    merged.loc[merged['overlap'] < 0.7, 'flag'] = 'FLAG_SIMILARITE_BASSE'
    merged.loc[merged['gestes_ajoutes'] == 0, 'flag'] = 'FLAG_AUCUN_AJOUT'
    merged.loc[merged['gestes_ajoutes'] > 10, 'flag'] = 'FLAG_GESTES_INCORRECTS'
    
    # Statistiques
    stats = {
        'total_paires': int(len(merged)),
        'overlap_moyen': float(merged['overlap'].mean()),
        'gestes_ajoutes_moyen': float(merged['gestes_ajoutes'].mean()),
        'gestes_ajoutes_median': float(merged['gestes_ajoutes'].median()),
        'CONFORME_count': int((merged['flag'] == 'CONFORME').sum()),
        'conformes_pct': float((merged['flag'] == 'CONFORME').sum() / len(merged) * 100) if len(merged) > 0 else 0.0,
        'FLAG_SIMILARITE_BASSE_count': int((merged['flag'] == 'FLAG_SIMILARITE_BASSE').sum()),
        'FLAG_SIMILARITE_BASSE_pct': float((merged['flag'] == 'FLAG_SIMILARITE_BASSE').sum() / len(merged) * 100) if len(merged) > 0 else 0.0,
        'FLAG_AUCUN_AJOUT_count': int((merged['flag'] == 'FLAG_AUCUN_AJOUT').sum()),
        'FLAG_AUCUN_AJOUT_pct': float((merged['flag'] == 'FLAG_AUCUN_AJOUT').sum() / len(merged) * 100) if len(merged) > 0 else 0.0,
        'FLAG_GESTES_INCORRECTS_count': int((merged['flag'] == 'FLAG_GESTES_INCORRECTS').sum()),
        'FLAG_GESTES_INCORRECTS_pct': float((merged['flag'] == 'FLAG_GESTES_INCORRECTS').sum() / len(merged) * 100) if len(merged) > 0 else 0.0
    }
    
    print(f"Total paires principale-ingredient: {stats['total_paires']:,}")
    print(f"Overlap moyen: {stats['overlap_moyen']:.3f}")
    print(f"Gestes ajoutés (moyenne): {stats['gestes_ajoutes_moyen']:.2f}")
    print(f"\nDistribution des flags:")
    print(f"  ✅ CONFORME: {stats['CONFORME_count']:,} ({stats['conformes_pct']:.2f}%)")
    print(f"  ⚠️  FLAG_SIMILARITE_BASSE (<0.7): {stats['FLAG_SIMILARITE_BASSE_count']:,} ({stats['FLAG_SIMILARITE_BASSE_pct']:.2f}%)")
    print(f"  ⚠️  FLAG_AUCUN_AJOUT: {stats['FLAG_AUCUN_AJOUT_count']:,} ({stats['FLAG_AUCUN_AJOUT_pct']:.2f}%)")
    print(f"  ⚠️  FLAG_GESTES_INCORRECTS (>10): {stats['FLAG_GESTES_INCORRECTS_count']:,} ({stats['FLAG_GESTES_INCORRECTS_pct']:.2f}%)")
    
    print(f"\n✅ Test 4B terminé")
    
    # Préparer le dataframe des flags
    flags_df = merged[merged['flag'] != 'CONFORME'][['id', 'flag']].copy()
    flags_df['test'] = '4B'
    flags_df['test_name'] = 'Variante INGREDIENTS VS PRINCIPALE (Similarité)'
    flags_df['delta_distribution'] = np.nan
    
    return stats, flags_df


# ==============================================================================
# SECTION 8 : TEST 5 - COMPARAISON DES 3 DATASETS
# ==============================================================================

def test_5_compare_three_datasets(
    graphs_brut_df: pd.DataFrame,
    graphs_with_gestures_df: pd.DataFrame,
    graphs_without_gestures_df: pd.DataFrame
) -> Tuple[Dict, pd.DataFrame]:
    """
    Test 5: Compare les 3 datasets pour vérifier la cohérence du nettoyage
    
    Args:
        graphs_brut_df: Dataset D1 (brut/avant nettoyage)
        graphs_with_gestures_df: Dataset D2 (nettoyé avec non-gestes)
        graphs_without_gestures_df: Dataset D3 (nettoyé sans non-gestes)
    
    Returns:
        Tuple (stats_dict, flags_df)
    """
    print(f"\n{'='*80}")
    print(f"TEST 5 : COMPARAISON DES 3 DATASETS")
    print(f"{'='*80}\n")
    
    # Parser les actions pour chaque dataset
    for df in [graphs_brut_df, graphs_with_gestures_df, graphs_without_gestures_df]:
        df['actions_parsed'] = df['actions'].apply(parse_actions_column)
        df['action_length'] = df['actions_parsed'].apply(len)
    
    # Filtrer uniquement les variantes principales pour comparaison
    d1_princ = graphs_brut_df[graphs_brut_df['type_2'] == 'variante_principale'].copy()
    d2_princ = graphs_with_gestures_df[graphs_with_gestures_df['type_2'] == 'variante_principale'].copy()
    d3_princ = graphs_without_gestures_df[graphs_without_gestures_df['type_2'] == 'variante_principale'].copy()
    
    # Merger les 3 datasets
    merged = d1_princ[['id', 'action_length']].merge(
        d2_princ[['id', 'action_length']],
        on='id',
        how='inner',
        suffixes=('_d1', '_d2')
    ).merge(
        d3_princ[['id', 'action_length']],
        on='id',
        how='inner'
    )
    merged.rename(columns={'action_length': 'action_length_d3'}, inplace=True)
    
    # Calculer les ratios
    merged['R1_D2_D1'] = merged['action_length_d2'] / merged['action_length_d1'].replace(0, np.nan)
    merged['R2_D3_D2'] = merged['action_length_d3'] / merged['action_length_d2'].replace(0, np.nan)
    merged['R3_D3_D1'] = merged['action_length_d3'] / merged['action_length_d1'].replace(0, np.nan)
    
    # Flags : vérifier que D1 >= D2 >= D3
    merged['flag'] = 'CONFORME'
    merged.loc[
        (merged['action_length_d1'] < merged['action_length_d2']) |
        (merged['action_length_d2'] < merged['action_length_d3']),
        'flag'
    ] = 'FLAG_ORDRE_INCOHERENT'
    
    # Statistiques
    stats = {
        'total_recettes_communes': int(len(merged)),
        'dataset_D1_brut': {
            'taille_moyenne': float(merged['action_length_d1'].mean()),
            'taille_mediane': float(merged['action_length_d1'].median())
        },
        'dataset_D2_cleaned_with_gestures': {
            'taille_moyenne': float(merged['action_length_d2'].mean()),
            'taille_mediane': float(merged['action_length_d2'].median())
        },
        'dataset_D3_cleaned_without_gestures': {
            'taille_moyenne': float(merged['action_length_d3'].mean()),
            'taille_mediane': float(merged['action_length_d3'].median())
        },
        'ratios_moyens': {
            'R1_D2_D1': float(merged['R1_D2_D1'].mean()),
            'R2_D3_D2': float(merged['R2_D3_D2'].mean()),
            'R3_D3_D1': float(merged['R3_D3_D1'].mean())
        },
        'flags': {
            'FLAG_ORDRE_INCOHERENT': int((merged['flag'] == 'FLAG_ORDRE_INCOHERENT').sum()),
            'pct_incoherent': float((merged['flag'] == 'FLAG_ORDRE_INCOHERENT').sum() / len(merged) * 100) if len(merged) > 0 else 0.0,
            'pct_conformes': float((merged['flag'] == 'CONFORME').sum() / len(merged) * 100) if len(merged) > 0 else 0.0
        }
    }
    
    print(f"Total recettes communes aux 3 datasets: {stats['total_recettes_communes']:,}")
    print(f"\nTaille moyenne des séquences (variante principale):")
    print(f"  D1 (brut): {stats['dataset_D1_brut']['taille_moyenne']:.2f} actions")
    print(f"  D2 (nettoyé avec): {stats['dataset_D2_cleaned_with_gestures']['taille_moyenne']:.2f} actions")
    print(f"  D3 (nettoyé sans): {stats['dataset_D3_cleaned_without_gestures']['taille_moyenne']:.2f} actions")
    print(f"\nRatios moyens:")
    print(f"  R1 (D2/D1): {stats['ratios_moyens']['R1_D2_D1']:.3f}")
    print(f"  R2 (D3/D2): {stats['ratios_moyens']['R2_D3_D2']:.3f}")
    print(f"  R3 (D3/D1): {stats['ratios_moyens']['R3_D3_D1']:.3f}")
    print(f"\nFlags:")
    print(f"  ❌ FLAG_ORDRE_INCOHERENT (D1<D2 ou D2<D3): {stats['flags']['FLAG_ORDRE_INCOHERENT']} ({stats['flags']['pct_incoherent']:.2f}%)")
    print(f"  ✅ CONFORME: {stats['flags']['pct_conformes']:.2f}%")
    
    print(f"\n✅ Test 5 terminé")
    
    # Préparer le dataframe des flags
    flags_df = merged[merged['flag'] != 'CONFORME'][['id', 'flag']].copy()
    flags_df['test'] = '5'
    flags_df['test_name'] = 'Comparaison des 3 DATASETS'
    flags_df['delta_distribution'] = np.nan
    
    return stats, flags_df


# ==============================================================================
# SECTION 9 : TEST 6 - COHÉRENCE GLOBALE PAR RECETTE
# ==============================================================================

def test_6_validate_recipe_coherence(
    graphs_df: pd.DataFrame
) -> Tuple[Dict, pd.DataFrame]:
    """
    Test 6: Vérifie la cohérence globale par recette
    (chaque recette doit avoir exactement 1 variante principale, etc.)
    
    Args:
        graphs_df: DataFrame des graphes
    
    Returns:
        Tuple (stats_dict, flags_df)
    """
    print(f"\n{'='*80}")
    print(f"TEST 6 : COHÉRENCE GLOBALE PAR RECETTE")
    print(f"{'='*80}\n")
    
    # Compter les variantes par recette
    variant_counts = graphs_df.groupby(['id', 'type_2']).size().unstack(fill_value=0)
    
    # Créer un dataframe des recettes
    recipe_coherence = pd.DataFrame({
        'id': variant_counts.index
    })
    
    # Vérifier la présence de la variante principale
    if 'variante_principale' in variant_counts.columns:
        recipe_coherence['nb_principale'] = variant_counts['variante_principale'].values
    else:
        recipe_coherence['nb_principale'] = 0
    
    # Flags
    recipe_coherence['flag'] = 'CONFORME'
    recipe_coherence.loc[recipe_coherence['nb_principale'] == 0, 'flag'] = 'FLAG_NO_PRINCIPALE'
    recipe_coherence.loc[recipe_coherence['nb_principale'] > 1, 'flag'] = 'FLAG_MULTIPLE_PRINCIPALE'
    
    # Statistiques
    stats = {
        'total_recettes': int(len(recipe_coherence)),
        'flags': {
            'FLAG_NO_PRINCIPALE': int((recipe_coherence['flag'] == 'FLAG_NO_PRINCIPALE').sum()),
            'FLAG_MULTIPLE_PRINCIPALE': int((recipe_coherence['flag'] == 'FLAG_MULTIPLE_PRINCIPALE').sum()),
            'pct_conformes': float((recipe_coherence['flag'] == 'CONFORME').sum() / len(recipe_coherence) * 100) if len(recipe_coherence) > 0 else 0.0
        }
    }
    
    print(f"Total recettes analysées: {stats['total_recettes']:,}")
    print(f"\nFlags:")
    print(f"  ❌ FLAG_NO_PRINCIPALE: {stats['flags']['FLAG_NO_PRINCIPALE']}")
    print(f"  ❌ FLAG_MULTIPLE_PRINCIPALE: {stats['flags']['FLAG_MULTIPLE_PRINCIPALE']}")
    print(f"  ✅ CONFORME: {stats['flags']['pct_conformes']:.2f}%")
    
    print(f"\n✅ Test 6 terminé")
    
    # Préparer le dataframe des flags
    flags_df = recipe_coherence[recipe_coherence['flag'] != 'CONFORME'][['id', 'flag']].copy()
    flags_df['test'] = '6'
    flags_df['test_name'] = 'Cohérence GLOBALE par recette'
    flags_df['delta_distribution'] = np.nan
    
    return stats, flags_df


# ==============================================================================
# SECTION 10 : CONSTRUCTION DES DATASETS FINAUX
# ==============================================================================
def count_actions_vectorized(actions_series: pd.Series) -> np.ndarray:
    """
    Version vectorisée pour comptage ultra-rapide.
    Utilise numpy pour traiter toute la série d'un coup.
    
    Args:
        actions_series: Série pandas contenant les actions
    
    Returns:
        Array numpy avec le nombre d'actions
    """
    def _count_single(val):
        if pd.isna(val):
            return 0
        if isinstance(val, list):
            return len(val)
        if isinstance(val, str):
            s = val.strip()
            if s in ('[]', ''):
                return 0
            if s.startswith('[') and s.endswith(']'):
                inner = s[1:-1].strip()
                return (inner.count(',') + 1) if inner else 0
            if ',' in s:
                return s.count(',') + 1
            return 1 if s else 0
        return 0
    
    # Utiliser np.vectorize pour appliquer la fonction de manière optimisée
    vectorized_count = np.vectorize(_count_single)
    return vectorized_count(actions_series.values)


def build_dataset_resume(
    graphs_brut_df: pd.DataFrame,
    graphs_with_gestures_df: pd.DataFrame,
    graphs_without_gestures_df: pd.DataFrame,
    show_progress: bool = True
) -> pd.DataFrame:
    """
    Version ULTRA-RAPIDE avec numpy vectorisé
    
    Args:
        graphs_brut_df: Dataset brut
        graphs_with_gestures_df: Dataset nettoyé avec non-gestes
        graphs_without_gestures_df: Dataset nettoyé sans non-gestes
        show_progress: Afficher la progression
    
    Returns:
        DataFrame du résumé
    """
    if show_progress:
        print(f"\n{'='*80}")
        print(f"CONSTRUCTION DU DATASET_RESUME (VERSION ULTRA-RAPIDE)")
        print(f"{'='*80}\n")
    
    all_dfs = []
    
    datasets = [
        (graphs_brut_df, 'avant_nettoyage'),
        (graphs_with_gestures_df, 'avec_non_geste'),
        (graphs_without_gestures_df, 'sans_non_geste')
    ]
    
    for df, type_dataset in datasets:
        if show_progress:
            print(f"  📊 Traitement de '{type_dataset}'... ({len(df):,} lignes)")
        
        # Créer directement le DataFrame résultat
        result_df = pd.DataFrame({
            'id': df['id'].values,
            'nombre_actions': count_actions_vectorized(df['actions']),
            'variante': df['type_2'].values,
            'type_dataset': type_dataset
        })
        
        all_dfs.append(result_df)
        
        if show_progress:
            print(f"     ✅ Terminé")
    
    # Concaténation finale
    dataset_resume = pd.concat(all_dfs, ignore_index=True)
    dataset_resume = dataset_resume[['id', 'nombre_actions', 'variante', 'type_dataset']]
    
    if show_progress:
        print(f"\n✅ Dataset résumé créé : {len(dataset_resume):,} lignes")
    
    return dataset_resume



def build_dataset_synthese(all_flags_dfs: List[pd.DataFrame]) -> pd.DataFrame:
    """
    Construit le dataset_synthese.csv en agrégeant tous les flags
    
    Colonnes: id | test | test_name | flag | delta_distribution
    
    Args:
        all_flags_dfs: Liste de DataFrames contenant les flags de chaque test
    
    Returns:
        DataFrame de synthèse
    """
    print(f"\n{'='*80}")
    print(f"CONSTRUCTION DU DATASET_SYNTHESE")
    print(f"{'='*80}\n")
    
    # Concaténer tous les flags
    dataset_synthese = pd.concat(all_flags_dfs, ignore_index=True)
    
    # S'assurer que les colonnes sont dans le bon ordre
    dataset_synthese = dataset_synthese[['id', 'test', 'test_name', 'flag', 'delta_distribution']]
    
    # Trier par id puis par test
    dataset_synthese = dataset_synthese.sort_values(['id', 'test']).reset_index(drop=True)
    
    print(f"✅ Dataset synthèse créé : {len(dataset_synthese):,} flags détectés")
    print(f"   Colonnes : {list(dataset_synthese.columns)}")
    print(f"\nDistribution des flags par test :")
    for test_num in sorted(dataset_synthese['test'].unique()):
        count = len(dataset_synthese[dataset_synthese['test'] == test_num])
        print(f"   Test {test_num}: {count:,} flags")
    
    return dataset_synthese


# ==============================================================================
# SECTION 11 : GÉNÉRATION DU RAPPORT MARKDOWN
# ==============================================================================

def generate_strategy_2_report_markdown(
    all_stats: Dict,
    output_dir: str
) -> str:
    """
    Génère un rapport Markdown complet de la Stratégie 2
    
    Args:
        all_stats: Dictionnaire contenant toutes les statistiques
        output_dir: Répertoire de sortie
    
    Returns:
        Chemin du fichier rapport
    """
    print(f"\n{'='*80}")
    print(f"GÉNÉRATION DU RAPPORT MARKDOWN")
    print(f"{'='*80}\n")
    
    report_lines = []
    
    # Header
    report_lines.append("# Rapport de Validation Structurelle - Stratégie 2")
    report_lines.append("")
    report_lines.append(f"**Date de génération** : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report_lines.append("")
    report_lines.append("---")
    report_lines.append("")
    
    # Introduction
    report_lines.append("## Introduction")
    report_lines.append("")
    report_lines.append("Ce rapport présente les résultats de la validation structurelle des graphes de recettes.")
    report_lines.append("La validation comprend 6 tests automatisés vérifiant la cohérence et la qualité des données.")
    report_lines.append("")
    report_lines.append("---")
    report_lines.append("")
    
    # Test 1 - Statistiques de tailles SUR LES 3 DATASETS
    report_lines.append("## Test 1 : Calcul de la Taille des Listes d'Actions")
    report_lines.append("")
    report_lines.append("Le Test 1 a été exécuté sur les **3 datasets** pour comparer l'évolution des statistiques.")
    report_lines.append("")
    
    # Fonction helper pour générer les tables de stats
    def add_stats_section(stats_dict, dataset_name, report_lines):
        """Ajoute une section de statistiques au rapport"""
        
        report_lines.append(f"### 📊 {dataset_name}")
        report_lines.append("")
        
        # Statistiques globales
        if 'globales' in stats_dict:
            report_lines.append("#### Statistiques Globales")
            report_lines.append("")
            stats = stats_dict['globales']
            report_lines.append(f"- **Total de graphes** : {stats['count']:,}")
            report_lines.append(f"- **Longueur moyenne** : {stats['longueur_moyenne']:.2f} ± {stats['ecart_type']:.2f} actions")
            report_lines.append(f"- **Médiane** : {stats['longueur_mediane']:.1f}")
            report_lines.append(f"- **Min** : {stats['longueur_min']}, **Max** : {stats['longueur_max']}")
            report_lines.append(f"- **Quartiles** : Q1={stats['q1']:.1f}, Q3={stats['q3']:.1f}")
            report_lines.append(f"- **Outliers** : {stats['outliers_count']:,} ({stats['outliers_percentage']:.2f}%)")
            report_lines.append(f"- **Bornes outliers** : [{stats['lower_bound']:.1f}, {stats['upper_bound']:.1f}]")
            report_lines.append("")
        
        # Statistiques par variante
        if 'par_variante' in stats_dict and stats_dict['par_variante']:
            report_lines.append("#### Statistiques par Variante")
            report_lines.append("")
            report_lines.append("| Variante | Count | Moyenne ± Écart-type | Médiane | Min-Max | Q1-Q3 | Outliers | % Outliers |")
            report_lines.append("|----------|-------|----------------------|---------|---------|-------|----------|------------|")
            for variant, vstats in sorted(stats_dict['par_variante'].items()):
                report_lines.append(
                    f"| {variant} | {vstats['count']:,} | {vstats['longueur_moyenne']:.2f} ± {vstats['ecart_type']:.2f} | "
                    f"{vstats['longueur_mediane']:.1f} | [{vstats['longueur_min']}, {vstats['longueur_max']}] | "
                    f"[{vstats['q1']:.1f}, {vstats['q3']:.1f}] | {vstats['outliers_count']:,} | {vstats['outliers_percentage']:.2f}% |"
                )
            report_lines.append("")
        
        # Statistiques par catégorie
        if 'par_category' in stats_dict and stats_dict['par_category']:
            report_lines.append("#### Statistiques par Catégorie de Cuisine")
            report_lines.append("")
            report_lines.append("| Catégorie | Count | Moyenne ± Écart-type | Médiane | Min-Max | Q1-Q3 | Outliers | % Outliers |")
            report_lines.append("|-----------|-------|----------------------|---------|---------|-------|----------|------------|")
            for category, cstats in sorted(stats_dict['par_category'].items()):
                report_lines.append(
                    f"| {category} | {cstats['count']:,} | {cstats['longueur_moyenne']:.2f} ± {cstats['ecart_type']:.2f} | "
                    f"{cstats['longueur_mediane']:.1f} | [{cstats['longueur_min']}, {cstats['longueur_max']}] | "
                    f"[{cstats['q1']:.1f}, {cstats['q3']:.1f}] | {cstats['outliers_count']:,} | {cstats['outliers_percentage']:.2f}% |"
                )
            report_lines.append("")
        
        # Statistiques par complexité
        if 'par_complexity' in stats_dict and stats_dict['par_complexity']:
            report_lines.append("#### Statistiques par Niveau de Complexité")
            report_lines.append("")
            report_lines.append("| Complexité | Count | Moyenne ± Écart-type | Médiane | Min-Max | Q1-Q3 | Outliers | % Outliers |")
            report_lines.append("|------------|-------|----------------------|---------|---------|-------|----------|------------|")
            complexity_order = ['simple', 'moyenne', 'elevee', 'unknown']
            for complexity in complexity_order:
                if complexity in stats_dict['par_complexity']:
                    cpstats = stats_dict['par_complexity'][complexity]
                    report_lines.append(
                        f"| {complexity} | {cpstats['count']:,} | {cpstats['longueur_moyenne']:.2f} ± {cpstats['ecart_type']:.2f} | "
                        f"{cpstats['longueur_mediane']:.1f} | [{cpstats['longueur_min']}, {cpstats['longueur_max']}] | "
                        f"[{cpstats['q1']:.1f}, {cpstats['q3']:.1f}] | {cpstats['outliers_count']:,} | {cpstats['outliers_percentage']:.2f}% |"
                    )
            report_lines.append("")
        
        report_lines.append("")
    
    # Ajouter les statistiques pour chaque dataset
    if 'test1_brut' in all_stats:
        add_stats_section(all_stats['test1_brut'], "Dataset BRUT (Avant Nettoyage)", report_lines)
    
    if 'test1_avec' in all_stats:
        add_stats_section(all_stats['test1_avec'], "Dataset AVEC Non-Gestes", report_lines)
    
    if 'test1_sans' in all_stats:
        add_stats_section(all_stats['test1_sans'], "Dataset SANS Non-Gestes", report_lines)
    
    # Comparaison des 3 datasets (Test 1)
    if 'test1_brut' in all_stats and 'test1_avec' in all_stats and 'test1_sans' in all_stats:
        report_lines.append("### 📊 Comparaison des 3 Datasets (Test 1)")
        report_lines.append("")
        report_lines.append("| Métrique | Brut | Avec Non-Gestes | Sans Non-Gestes |")
        report_lines.append("|----------|------|-----------------|-----------------|")
        report_lines.append(
            f"| Longueur moyenne | {all_stats['test1_brut']['globales']['longueur_moyenne']:.2f} | "
            f"{all_stats['test1_avec']['globales']['longueur_moyenne']:.2f} | "
            f"{all_stats['test1_sans']['globales']['longueur_moyenne']:.2f} |"
        )
        report_lines.append(
            f"| Écart-type | {all_stats['test1_brut']['globales']['ecart_type']:.2f} | "
            f"{all_stats['test1_avec']['globales']['ecart_type']:.2f} | "
            f"{all_stats['test1_sans']['globales']['ecart_type']:.2f} |"
        )
        report_lines.append(
            f"| Médiane | {all_stats['test1_brut']['globales']['longueur_mediane']:.1f} | "
            f"{all_stats['test1_avec']['globales']['longueur_mediane']:.1f} | "
            f"{all_stats['test1_sans']['globales']['longueur_mediane']:.1f} |"
        )
        report_lines.append(
            f"| Outliers (%) | {all_stats['test1_brut']['globales']['outliers_percentage']:.2f}% | "
            f"{all_stats['test1_avec']['globales']['outliers_percentage']:.2f}% | "
            f"{all_stats['test1_sans']['globales']['outliers_percentage']:.2f}% |"
        )
        report_lines.append("")
    
    report_lines.append("---")
    report_lines.append("")
    
    # Test 2
    if 'test2' in all_stats:
        report_lines.append("## Test 2 : Comparaison Variante Principale vs Nombre d'Instructions")
        report_lines.append("")
        stats = all_stats['test2']
        report_lines.append(f"- **Total variantes principales** : {stats['total_principales']:,}")
        report_lines.append(f"- **Ratio moyen (actions/instructions)** : {stats['ratio_moyen']:.2f}")
        report_lines.append(f"- **Ratio médian** : {stats['ratio_median']:.2f}")
        report_lines.append("")
        report_lines.append("### Distribution des Flags")
        report_lines.append("")
        report_lines.append("| Flag | Count | Pourcentage |")
        report_lines.append("|------|-------|-------------|")
        report_lines.append(f"| ✅ CONFORME | {stats['CONFORME_count']:,} | {stats['CONFORME_pct']:.2f}% |")
        report_lines.append(f"| ⚠️ FLAG_RATIO_FAIBLE | {stats['FLAG_RATIO_FAIBLE_count']:,} | {stats['FLAG_RATIO_FAIBLE_pct']:.2f}% |")
        report_lines.append(f"| ⚠️ FLAG_RATIO_ELEVE | {stats['FLAG_RATIO_ELEVE_count']:,} | {stats['FLAG_RATIO_ELEVE_pct']:.2f}% |")
        report_lines.append(f"| ❌ FLAG_CRITIQUE | {stats['FLAG_CRITIQUE_count']:,} | {stats['FLAG_CRITIQUE_pct']:.2f}% |")
        report_lines.append("")
        report_lines.append("---")
        report_lines.append("")
    
    # Test 3
    if 'test3' in all_stats:
        report_lines.append("## Test 3 : Validation Variante Ingrédients")
        report_lines.append("")
        stats = all_stats['test3']
        report_lines.append(f"- **Total paires principale-ingredient** : {stats['total_paires']:,}")
        report_lines.append(f"- **Delta moyen (ingredient - principale)** : {stats['delta_moyen']:.2f} actions")
        report_lines.append(f"- **Delta médian** : {stats['delta_median']:.1f}")
        report_lines.append(f"- **Delta min** : {stats['delta_min']}, **Delta max** : {stats['delta_max']}")
        report_lines.append("")
        report_lines.append("### Distribution des Flags")
        report_lines.append("")
        report_lines.append("| Flag | Count | Pourcentage |")
        report_lines.append("|------|-------|-------------|")
        report_lines.append(f"| ✅ CONFORME (1-10 ajouts) | {stats['CONFORME_count']:,} | {stats['CONFORME_pct']:.2f}% |")
        report_lines.append(f"| ⚠️ FLAG_AUCUN_AJOUT | {stats['FLAG_AUCUN_AJOUT_count']:,} | {stats['FLAG_AUCUN_AJOUT_pct']:.2f}% |")
        report_lines.append(f"| ⚠️ FLAG_TROP_AJOUT (>10) | {stats['FLAG_TROP_AJOUT_count']:,} | {stats['FLAG_TROP_AJOUT_pct']:.2f}% |")
        report_lines.append(f"| ❌ FLAG_CRITIQUE_NEGATIF | {stats['FLAG_CRITIQUE_NEGATIF_count']:,} | {stats['FLAG_CRITIQUE_NEGATIF_pct']:.2f}% |")
        report_lines.append("")
        report_lines.append("---")
        report_lines.append("")
    
    # Test 4A
    if 'test4a' in all_stats:
        report_lines.append("## Test 4A : Comparaison Permutations vs Principale")
        report_lines.append("")
        stats = all_stats['test4a']
        report_lines.append(f"- **Total permutations** : {stats['total_permutations']:,}")
        report_lines.append(f"- **Overlap moyen** : {stats['overlap_moyen']:.3f}")
        report_lines.append(f"- **Distance Levenshtein moyenne** : {stats['levenshtein_moyen']:.2f}")
        report_lines.append("")
        report_lines.append("### Distribution des Flags")
        report_lines.append("")
        report_lines.append("| Flag | Count | Pourcentage |")
        report_lines.append("|------|-------|-------------|")
        report_lines.append(f"| ✅ CONFORME (overlap 0.6-1.0) | {stats['CONFORME_count']:,} | {stats['conformes_pct']:.2f}% |")
        report_lines.append(f"| ⚠️ FLAG_SIMILARITE_BASSE (<0.6) | {stats['FLAG_SIMILARITE_BASSE_count']:,} | {stats['FLAG_SIMILARITE_BASSE_pct']:.2f}% |")
        report_lines.append(f"| ⚠️ FLAG_IDENTIQUE (=1.0) | {stats['FLAG_IDENTIQUE_count']:,} | {stats['FLAG_IDENTIQUE_pct']:.2f}% |")
        report_lines.append("")
        report_lines.append("---")
        report_lines.append("")
    
    # Test 4B
    if 'test4b' in all_stats:
        report_lines.append("## Test 4B : Variante Ingrédients vs Principale (Similarité)")
        report_lines.append("")
        stats = all_stats['test4b']
        report_lines.append(f"- **Total paires** : {stats['total_paires']:,}")
        report_lines.append(f"- **Overlap moyen** : {stats['overlap_moyen']:.3f}")
        report_lines.append(f"- **Gestes ajoutés (moyenne)** : {stats['gestes_ajoutes_moyen']:.2f}")
        report_lines.append(f"- **Gestes ajoutés (médiane)** : {stats['gestes_ajoutes_median']:.1f}")
        report_lines.append("")
        report_lines.append("### Distribution des Flags")
        report_lines.append("")
        report_lines.append("| Flag | Count | Pourcentage |")
        report_lines.append("|------|-------|-------------|")
        report_lines.append(f"| ✅ CONFORME | {stats['CONFORME_count']:,} | {stats['conformes_pct']:.2f}% |")
        report_lines.append(f"| ⚠️ FLAG_SIMILARITE_BASSE (<0.7) | {stats['FLAG_SIMILARITE_BASSE_count']:,} | {stats['FLAG_SIMILARITE_BASSE_pct']:.2f}% |")
        report_lines.append(f"| ⚠️ FLAG_AUCUN_AJOUT | {stats['FLAG_AUCUN_AJOUT_count']:,} | {stats['FLAG_AUCUN_AJOUT_pct']:.2f}% |")
        report_lines.append(f"| ⚠️ FLAG_GESTES_INCORRECTS (>10) | {stats['FLAG_GESTES_INCORRECTS_count']:,} | {stats['FLAG_GESTES_INCORRECTS_pct']:.2f}% |")
        report_lines.append("")
        report_lines.append("---")
        report_lines.append("")
    
    # Test 5
    if 'test5' in all_stats:
        report_lines.append("## Test 5 : Comparaison des 3 Datasets")
        report_lines.append("")
        stats = all_stats['test5']
        report_lines.append(f"- **Total recettes communes** : {stats['total_recettes_communes']:,}")
        report_lines.append("")
        report_lines.append("### Taille Moyenne des Séquences (variante principale)")
        report_lines.append("")
        report_lines.append("| Dataset | Moyenne | Médiane |")
        report_lines.append("|---------|---------|---------|")
        report_lines.append(f"| D1 (brut/avant nettoyage) | {stats['dataset_D1_brut']['taille_moyenne']:.2f} | {stats['dataset_D1_brut']['taille_mediane']:.1f} |")
        report_lines.append(f"| D2 (avec non-gestes) | {stats['dataset_D2_cleaned_with_gestures']['taille_moyenne']:.2f} | {stats['dataset_D2_cleaned_with_gestures']['taille_mediane']:.1f} |")
        report_lines.append(f"| D3 (sans non-gestes) | {stats['dataset_D3_cleaned_without_gestures']['taille_moyenne']:.2f} | {stats['dataset_D3_cleaned_without_gestures']['taille_mediane']:.1f} |")
        report_lines.append("")
        report_lines.append("### Ratios Moyens")
        report_lines.append("")
        report_lines.append("| Ratio | Valeur |")
        report_lines.append("|-------|--------|")
        report_lines.append(f"| R1 (D2/D1) | {stats['ratios_moyens']['R1_D2_D1']:.3f} |")
        report_lines.append(f"| R2 (D3/D2) | {stats['ratios_moyens']['R2_D3_D2']:.3f} |")
        report_lines.append(f"| R3 (D3/D1) | {stats['ratios_moyens']['R3_D3_D1']:.3f} |")
        report_lines.append("")
        report_lines.append("### Flags")
        report_lines.append("")
        report_lines.append("| Flag | Count | Pourcentage |")
        report_lines.append("|------|-------|-------------|")
        report_lines.append(f"| ❌ FLAG_ORDRE_INCOHERENT (D1<D2 ou D2<D3) | {stats['flags']['FLAG_ORDRE_INCOHERENT']} | {stats['flags']['pct_incoherent']:.2f}% |")
        report_lines.append(f"| ✅ CONFORME | - | {stats['flags']['pct_conformes']:.2f}% |")
        report_lines.append("")
        report_lines.append("---")
        report_lines.append("")
    
    # Test 6
    if 'test6' in all_stats:
        report_lines.append("## Test 6 : Cohérence Globale par Recette")
        report_lines.append("")
        stats = all_stats['test6']
        report_lines.append(f"- **Total recettes analysées** : {stats['total_recettes']:,}")
        report_lines.append("")
        report_lines.append("### Flags")
        report_lines.append("")
        report_lines.append("| Flag | Count |")
        report_lines.append("|------|-------|")
        report_lines.append(f"| ❌ FLAG_NO_PRINCIPALE | {stats['flags']['FLAG_NO_PRINCIPALE']} |")
        report_lines.append(f"| ❌ FLAG_MULTIPLE_PRINCIPALE | {stats['flags']['FLAG_MULTIPLE_PRINCIPALE']} |")
        report_lines.append(f"| ✅ CONFORME | {stats['flags']['pct_conformes']:.2f}% |")
        report_lines.append("")
        report_lines.append("---")
        report_lines.append("")
    
    # Conclusion
    report_lines.append("## Conclusion")
    report_lines.append("")
    report_lines.append("### Récapitulatif des Flags Critiques")
    report_lines.append("")
    
    total_flags = 0
    if 'test2' in all_stats:
        total_flags += all_stats['test2']['FLAG_CRITIQUE_count']
    if 'test3' in all_stats:
        total_flags += all_stats['test3']['FLAG_CRITIQUE_NEGATIF_count']
    if 'test5' in all_stats:
        total_flags += all_stats['test5']['flags']['FLAG_ORDRE_INCOHERENT']
    if 'test6' in all_stats:
        total_flags += all_stats['test6']['flags']['FLAG_NO_PRINCIPALE']
        total_flags += all_stats['test6']['flags']['FLAG_MULTIPLE_PRINCIPALE']
    
    report_lines.append(f"**Total de flags critiques détectés** : {total_flags}")
    report_lines.append("")
    report_lines.append("### Recommandations")
    report_lines.append("")
    report_lines.append("1. **Examiner les recettes avec flags critiques** : Priorité absolue pour les FLAGS_CRITIQUE détectés dans les tests 2, 3, 5 et 6")
    report_lines.append("2. **Vérifier l'intégrité du pipeline de nettoyage** : Les incohérences dans le Test 5 indiquent des problèmes potentiels")
    report_lines.append("3. **Corriger les variantes avec incohérences structurelles** : Utiliser le dataset_synthese.csv pour identifier et corriger")
    report_lines.append("4. **Analyser les patterns d'erreurs** : Identifier les types de recettes problématiques (par category/complexity)")
    report_lines.append("")
    report_lines.append("---")
    report_lines.append("")
    report_lines.append("## Fichiers Générés")
    report_lines.append("")
    report_lines.append("- **strategy_2_report.md** : Ce rapport")
    report_lines.append("- **dataset_synthese.csv** : Tous les flags détectés par test et par recette")
    report_lines.append("- **dataset_resume.csv** : Nombre d'actions par recette/variante/dataset")
    report_lines.append("")
    report_lines.append("---")
    report_lines.append("")
    report_lines.append(f"**Fin du rapport** - Généré le {datetime.now().strftime('%Y-%m-%d à %H:%M:%S')}")
    
    # Sauvegarder le rapport
    report_text = '\n'.join(report_lines)
    report_file = os.path.join(output_dir, 'strategy_2_report.md')
    
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write(report_text)
    
    print(f"✅ Rapport Markdown sauvegardé : {report_file}")
    print(f"   Longueur : {len(report_lines)} lignes")
    
    return report_file



# ==============================================================================
# SECTION 12 : PIPELINE PRINCIPAL
# ==============================================================================

def run_strategy_2_pipeline(
    recipes_csv: str,
    graphs_recipes_csv: str,
    graphs_cleaned_with_gestures_csv: str,
    graphs_cleaned_without_gestures_csv: str,
    output_dir: str = "strategy_2_results"
):
    """
    Exécute le pipeline complet de la Stratégie 2 (VERSION AMÉLIORÉE)
    
    Args:
        recipes_csv: Chemin vers recipes.csv
        graphs_recipes_csv: Chemin vers graphs_recipes.csv (brut)
        graphs_cleaned_with_gestures_csv: Chemin vers graphs_recipes_cleaned_with_gestures.csv
        graphs_cleaned_without_gestures_csv: Chemin vers graphs_recipes_cleaned_without_non_gestures.csv
        output_dir: Répertoire de sortie
    """
    print(f"\n{'#'*80}")
    print(f"# PIPELINE STRATÉGIE 2 - VALIDATION STRUCTURELLE (VERSION AMÉLIORÉE)")
    print(f"{'#'*80}\n")
    
    # Créer le répertoire de sortie
    create_output_directory(output_dir)
    
    # ========== CHARGEMENT DES DONNÉES ==========
    print("📂 CHARGEMENT DES DONNÉES...")
    recipes_df = pd.read_csv(recipes_csv)
    graphs_brut_df = pd.read_csv(graphs_recipes_csv)
    graphs_with_gestures_df = pd.read_csv(graphs_cleaned_with_gestures_csv)
    graphs_without_gestures_df = pd.read_csv(graphs_cleaned_without_gestures_csv)
    
    print(f"  ✅ Recettes: {len(recipes_df):,}")
    print(f"  ✅ Graphes bruts: {len(graphs_brut_df):,}")
    print(f"  ✅ Graphes nettoyés (avec): {len(graphs_with_gestures_df):,}")
    print(f"  ✅ Graphes nettoyés (sans): {len(graphs_without_gestures_df):,}")
    
    # ========== AJOUT DES COLONNES CATEGORY ET COMPLEXITY ==========
    print(f"\n🏷️  CLASSIFICATION DES RECETTES...")
    recipes_df['category'] = recipes_df['title'].apply(classify_cuisine_type)
    recipes_df['complexity'] = recipes_df['number_of_steps'].apply(classify_complexity)
    
    print(f"  ✅ Colonne 'category' ajoutée")
    print(f"     Distribution : {dict(recipes_df['category'].value_counts())}")
    print(f"  ✅ Colonne 'complexity' ajoutée")
    print(f"     Distribution : {dict(recipes_df['complexity'].value_counts())}")
    
    # Sauvegarder le recipes.csv enrichi (optionnel, pour référence)
    recipes_enriched_path = os.path.join(output_dir, 'recipes_enriched.csv')
    recipes_df.to_csv(recipes_enriched_path, index=False)
    print(f"  ✅ Recipes enrichi sauvegardé : {recipes_enriched_path}")
    
    # ========== EXÉCUTION DES TESTS ==========
    all_stats = {}
    all_flags_dfs = []
    
    # Test 1: Calcul des tailles sur les 3 DATASETS
    print(f"\n{'='*80}")
    print(f"EXÉCUTION DU TEST 1 SUR LES 3 DATASETS")
    print(f"{'='*80}")
    
    # Test 1 - Dataset brut (avant nettoyage)
    print(f"\n🔷 TEST 1 - DATASET BRUT (AVANT NETTOYAGE)")
    stats_test1_brut, _ = test_1_calculate_action_lengths(graphs_brut_df, recipes_df)
    all_stats['test1_brut'] = stats_test1_brut
    
    # Test 1 - Dataset nettoyé avec non-gestes
    print(f"\n🔷 TEST 1 - DATASET AVEC NON-GESTES")
    stats_test1_avec, _ = test_1_calculate_action_lengths(graphs_with_gestures_df, recipes_df)
    all_stats['test1_avec'] = stats_test1_avec
    
    # Test 1 - Dataset nettoyé sans non-gestes
    print(f"\n🔷 TEST 1 - DATASET SANS NON-GESTES")
    stats_test1_sans, _ = test_1_calculate_action_lengths(graphs_without_gestures_df, recipes_df)
    all_stats['test1_sans'] = stats_test1_sans
    
    # Test 2: Variante principale vs nombre d'instructions
    stats_test2, flags_test2 = test_2_validate_principale_vs_steps(recipes_df, graphs_with_gestures_df)
    all_stats['test2'] = stats_test2
    all_flags_dfs.append(flags_test2)
    
    # Test 3: Variante ingredients vs taille théorique
    stats_test3, flags_test3 = test_3_validate_ingredients_variant(recipes_df, graphs_with_gestures_df)
    all_stats['test3'] = stats_test3
    all_flags_dfs.append(flags_test3)
    
    # Test 4A: Permutations vs principale
    stats_test4a, flags_test4a = test_4a_validate_permutations(graphs_with_gestures_df)
    all_stats['test4a'] = stats_test4a
    all_flags_dfs.append(flags_test4a)
    
    # Test 4B: Variante ingredients vs principale (similarité)
    stats_test4b, flags_test4b = test_4b_validate_ingredients_similarity(graphs_with_gestures_df)
    all_stats['test4b'] = stats_test4b
    all_flags_dfs.append(flags_test4b)
    
    # Test 5: Comparaison des 3 datasets
    stats_test5, flags_test5 = test_5_compare_three_datasets(
        graphs_brut_df,
        graphs_with_gestures_df,
        graphs_without_gestures_df
    )
    all_stats['test5'] = stats_test5
    all_flags_dfs.append(flags_test5)
    
    # Test 6: Cohérence globale par recette
    stats_test6, flags_test6 = test_6_validate_recipe_coherence(graphs_with_gestures_df)
    all_stats['test6'] = stats_test6
    all_flags_dfs.append(flags_test6)
    
    # ========== CONSTRUCTION DES DATASETS FINAUX ==========
    
    # Dataset résumé
    dataset_resume = build_dataset_resume(
        graphs_brut_df,
        graphs_with_gestures_df,
        graphs_without_gestures_df
    )
    resume_path = os.path.join(output_dir, 'dataset_resume.csv')
    dataset_resume.to_csv(resume_path, index=False)
    print(f"\n💾 Dataset résumé sauvegardé : {resume_path}")
    
    # Dataset synthèse
    dataset_synthese = build_dataset_synthese(all_flags_dfs)
    synthese_path = os.path.join(output_dir, 'dataset_synthese.csv')
    dataset_synthese.to_csv(synthese_path, index=False)
    print(f"💾 Dataset synthèse sauvegardé : {synthese_path}")
    
    # ========== GÉNÉRATION DU RAPPORT MARKDOWN ==========
    report_path = generate_strategy_2_report_markdown(all_stats, output_dir)
    
    # ========== RÉSUMÉ FINAL ==========
    print(f"\n{'#'*80}")
    print(f"# PIPELINE TERMINÉ AVEC SUCCÈS")
    print(f"{'#'*80}\n")
    print(f"📁 Tous les résultats sont dans : {output_dir}/")
    print(f"\n📄 Fichiers générés :")
    print(f"   1. {report_path}")
    print(f"   2. {synthese_path}")
    print(f"   3. {resume_path}")
    


"""
Pipeline complet pour la Stratégie 3 : Détection des Successions Illogiques
Identification des séquences d'actions sémantiquement incohérentes

Auteur: Pipeline de validation de qualité des données
Date: 2025-01-07
"""

import pandas as pd
import json
import numpy as np
import os
from typing import Dict, List, Tuple, Set
from datetime import datetime
from collections import Counter, defaultdict
import ast


# ==============================================================================
# SECTION 1 : TAXONOMIE DES VERBES ET CLASSIFICATION
# ==============================================================================

import json
from typing import Set

class VerbTaxonomy:
    """
    Taxonomie complète des verbes culinaires en 11 catégories fonctionnelles
    Version étendue avec classification exhaustive
    """
    
    PREPARATION_INITIALE = {
        'wash', 'rinse', 'clean', 'peel', 'trim', 'core', 'pit', 
        'bone', 'devein', 'scale', 'hull', 'shell', 'shuck', 'gut', 
        'beard', 'behead', 'skin', 'stem', 'seed', 'stone', 'fish', 
        'gill', 'head', 'eviscerate', 'fillet', 'filet', 'butcher', 
        'cavity', 'scrub', 'debone'
    }
    
    TRANSFORMATION_MECANIQUE = {
        'chop', 'dice', 'slice', 'cut', 'mince', 'julienne', 'cube', 
        'halve', 'quarter', 'grate', 'shred', 'crush', 'grind', 'pound', 
        'tenderize', 'mash', 'puree', 'shave', 'zest', 'chiffonade', 
        'brunoise', 'concasse', 'carve', 'chip', 'chunk', 'crumb', 
        'crumble', 'flake', 'smash', 'pulverise', 'splitter', 'hacken', 
        'wedge', 'wedge slice', 'crosscut', 'crosshatch', 'trisect', 
        'diagonal', 'pieces', 'break', 'crack', 'snap', 'tear', 
        'split', 'cleave', 'slash', 'slit', 'gash', 'nick', 'score', 
        'incision', 'chisel', 'plane', 'shear', 'mill', 'process', 
        'blitz', 'blenderize', 'whiz', 'whizz', 'texturize', 'fragment',
        'powder', 'powderize', 'crunch', 'splinter', 'shatter'
    }
    
    MELANGE_COMBINAISON = {
        'mix', 'stir', 'combine', 'whisk', 'beat', 'whip', 'fold', 
        'blend', 'toss', 'incorporate', 'emulsify', 'cream', 'knead',
        'amalgamate', 'meld', 'mingle', 'marry', 'muddle', 'agitate',
        'churn', 'swirl', 'swish', 'scramble', 'massage', 'work',
        'pulsate', 'pulse', 'claw', 'froth', 'foam', 'bubble',
        'aerate', 'fluff', 'lighten', 'premix', 'remix', 'overmix'
    }
    
    TRANSFERT_MANIPULATION = {
        'pour', 'add', 'place', 'put', 'transfer', 'spread', 'layer', 
        'arrange', 'fill', 'stuff', 'wrap', 'coat', 'brush', 'drizzle', 
        'sprinkle', 'dust', 'top', 'spoon', 'scoop', 'ladle', 'dab',
        'dollop', 'dot', 'drop', 'splash', 'splatter', 'spill', 'trickle',
        'drip', 'dribble', 'dump', 'plop', 'plonk', 'plunk', 'splat',
        'sploosh', 'splodge', 'splotch', 'dish', 'portion', 'stack',
        'pile', 'mound', 'nest', 'nestle', 'cradle', 'tuck', 'enfold',
        'enrobe', 'encase', 'encapsulate', 'package', 'bag', 'baggie',
        'bundle', 'wrapper', 'cover', 'blanket', 'tent', 'foil', 'unfoil',
        'cap', 'cork', 'center', 'position', 'slide', 'move', 'return',
        'retrieve', 'apply', 'anoint', 'bathe', 'dredge', 'drench',
        'smother', 'smear', 'smudge', 'grease', 'ungrease', 'butter',
        'flour', 'egg', 'egg wash', 'batter', 'crumb', 'bread',
        'moisten', 'moisturize', 'wet', 'dampen', 'damp', 'soak',
        'spritz', 'spray', 'mist', 'squirt', 'douse', 'drape',
        'blanket', 'film', 'glaze', 'frost', 'ice', 'sugar',
        'candy', 'caramelize', 'spice', 'season', 'preseason',
        'funnel', 'siphon', 'decant', 'ladle', 'blob', 'puddle'
    }
    
    CUISSON_ACTIVE = {
        'sauté', 'sautee', 'saute', 'stir-fry', 'stir fry', 'pan-fry', 
        'pan fry', 'sear', 'brown', 'caramelize', 'reduce', 'flip', 
        'turn', 'baste', 'fry', 'deep-fry', 'deep fry', 'shallow fry',
        'griddle', 'sweat', 'char', 'blacken', 'singe', 'scorch',
        'blister', 'crisp', 'crackle', 'sizzle', 'deglaze', 'toss',
        'agitate', 'rotate', 'shake'
    }
    
    CUISSON_PASSIVE = {
        'bake', 'roast', 'boil', 'simmer', 'steam', 'broil', 'grill', 
        'poach', 'braise', 'slow-cook', 'slow cook', 'pressure-cook', 
        'pressure cook', 'smoke', 'barbecue', 'toast', 'stew', 'coddle',
        'parboil', 'parbake', 'parcook', 'blanch', 'scald', 'flambe',
        'flame', 'fire', 'blast', 'blaze', 'burn', 'cook', 'cooks',
        'air fry', 'microwave', 'nuke', 'prebake', 'precook', 'overbake',
        'overcook', 'underbake', 'undercook', 'dry roast', 'torch',
        'gratinee', 'brÃ»lÃ©e', 'roil', 'bubble', 'shimmer'
    }
    
    TRANSFORMATION_THERMIQUE = {
        'cool', 'chill', 'freeze', 'refrigerate', 'thaw', 'defrost', 
        'warm', 'heat', 'reheat', 'melt', 'temper', 'preheat', 'prewarm',
        'room', 'flash', 'refresh', 'ice', 'congeal', 'solidify',
        'crystallize', 'set', 'firm', 'harden', 'soften', 'liquefy',
        'gel', 'coagulate', 'curdle', 'condense', 'evaporate', 'reduce',
        'thicken', 'thin', 'dissolve', 'melt', 'freeze', 'freezer',
        'dehydrate', 'dry', 'parch', 'wilt', 'shrink', 'swell',
        'expand', 'bloom', 'hydrate', 'rehydrate', 'restore'
    }
    
    ATTENTE_REPOS = {
        'rest', 'set', 'settle', 'rise', 'proof', 'ferment', 'marinate', 
        'pickle', 'cure', 'age', 'soak', 'steep', 'mature', 'ripen',
        'autolyse', 'leaven', 'inoculate', 'culture', 'preserve', 'can',
        'confit', 'brine', 'macerate', 'mull', 'presoak', 'leave', 'hang',
        'infuse', 'brew', 'percolate', 'activate', 'develop', 'mellow',
        'season', 'idle', 'wait', 'stand', 'absorb', 'adhere', 'bind'
    }
    
    EXTRACTION_SEPARATION = {
        'drain', 'strain', 'press', 'squeeze', 'filter', 'sift', 
        'separate', 'skim', 'extract', 'wring', 'decant', 'siphon',
        'leach', 'degas', 'deglaze', 'degrease', 'discharge', 'clarify',
        'declump', 'render', 'distill', 'percolate', 'perk', 'seep',
        'ooze', 'drip', 'blot', 'wipe', 'scum', 'remove', 'discard',
        'skim', 'scrape', 'express', 'extrude', 'force', 'juice'
    }
    
    FINITION_SERVICE = {
        'garnish', 'plate', 'serve', 'decorate', 'present', 'unmold', 
        'display', 'portion', 'ornament', 'dress', 'glaze', 'unglaze',
        'frost', 'brûlée', 'finish', 'nap', 'rim', 'mark', 'decorate',
        'embellish', 'arrange', 'present', 'drizzle', 'dust', 'top',
        'crown', 'adorn', 'burnish', 'polish', 'buff', 'shine',
        'gloss', 'glaze', 'varnish', 'coat', 'tipping', 'tip', 'tip out'
    }
    
    ACTIONS_SPECIALES = {
        'roll', 'shape', 'form', 'mold', 'mould', 'score', 'pierce', 
        'prick', 'skewer', 'thread', 'tie', 'secure', 'truss', 'bind',
        'braid', 'plait', 'twine', 'weave', 'coil', 'spiral', 'spiralize',
        'curl', 'twist', 'crimp', 'flute', 'pleat', 'fold', 'pinwheel',
        'snake', 'spool', 'rope', 'string', 'lace', 'interlace', 'dock',
        'butterfly', 'spatchcock', 'supreme', 'tournee', 'quenelle',
        'dimple', 'indent', 'ridge', 'fan', 'feather', 'fringe', 'ruffle',
        'fluff', 'nest', 'dome', 'mound', 'build', 'assemble', 'construct',
        'laminate', 'interleave', 'wrap', 'unwrap', 'unroll', 'unfurl',
        'uncurl', 'untwist', 'unwind', 'untie', 'untruss', 'unthread',
        'unskewer', 'unspit', 'unstack', 'unstuff', 'unshell', 'open',
        'close', 'seal', 'unseal', 'attach', 'detach', 'hook', 'unhook',
        'clump', 'cluster', 'bunch', 'gather', 'collect', 'compact',
        'compress', 'flatten', 'smooth', 'level', 'even', 'straighten',
        'round', 'square', 'curve', 'bend', 'angle', 'tilt', 'incline',
        'lean', 'prop', 'support', 'suspend', 'dangle', 'hang',
        'scallop', 'frizzle', 'corrugate', 'wrinkle', 'crinkle', 'crimp',
        'ruffle', 'plump', 'puff', 'inflate', 'deflate', 'compress',
        'sandwich', 'layer', 'stack', 'tier', 'graduate', 'overlap',
        'interlock', 'mesh', 'entwine', 'braid', 'knot', 'loop',
        'stitch', 'sew', 'needle', 'pin', 'tack', 'staple', 'clip',
        'clamp', 'grip', 'hold', 'stabilize', 'anchor', 'fix',
        'mount', 'erect', 'raise', 'lift', 'elevate', 'lower',
        'descend', 'drop', 'sink', 'submerge', 'immerse', 'dunk',
        'bathe', 'wash', 'rinse', 'flush', 'cleanse', 'purify',
        'sterilize', 'sanitize', 'disinfect', 'pasteurize',
        'taper', 'sharpen', 'dull', 'blunt', 'hone', 'whet',
        'strop', 'grind', 'file', 'rasp', 'scrape', 'scour',
        'burnish', 'polish', 'smooth', 'roughen', 'rough', 'coarsen',
        'texturize', 'dimple', 'corrugate', 'emboss', 'imprint',
        'stamp', 'press', 'mold', 'cast', 'sculpt', 'carve',
        'engrave', 'etch', 'inscribe', 'mark', 'label', 'tag',
        'identify', 'recognize', 'distinguish', 'differentiate',
        'convert', 'transform', 'modify', 'alter', 'change',
        'adjust', 'adapt', 'customize', 'personalize', 'tailor',
        'fit', 'size', 'scale', 'proportion', 'balance', 'equalize',
        'level', 'calibrate', 'tune', 'regulate', 'control',
        'moderate', 'modulate', 'temper', 'adjust', 'fine-tune',
        'tweak', 'refine', 'perfect', 'optimize', 'improve',
        'enhance', 'enrich', 'fortify', 'strengthen', 'reinforce',
        'bolster', 'support', 'buttress', 'brace', 'shore',
        'prop', 'cradle', 'nestle', 'cuddle', 'snug', 'snuggle',
        'tuck', 'wrap', 'envelope', 'encircle', 'surround',
        'encompass', 'enclose', 'contain', 'hold', 'cradle',
        'seat', 'position', 'situate', 'locate', 'place'
    }
    
    # Mapping catégorie -> nom
    CATEGORY_NAMES = {
        'PREPARATION_INITIALE': 'Préparation Initiale',
        'TRANSFORMATION_MECANIQUE': 'Transformation Mécanique',
        'MELANGE_COMBINAISON': 'Mélange/Combinaison',
        'TRANSFERT_MANIPULATION': 'Transfert/Manipulation',
        'CUISSON_ACTIVE': 'Cuisson Active',
        'CUISSON_PASSIVE': 'Cuisson Passive',
        'TRANSFORMATION_THERMIQUE': 'Transformation Thermique',
        'ATTENTE_REPOS': 'Attente/Repos',
        'EXTRACTION_SEPARATION': 'Extraction/Séparation',
        'FINITION_SERVICE': 'Finition/Service',
        'ACTIONS_SPECIALES': 'Actions Spéciales'
    }
    
    @classmethod
    def get_category(cls, verb: str) -> str:
        """
        Retourne la catégorie d'un verbe
        
        Args:
            verb: Verbe à classifier
        
        Returns:
            Nom de la catégorie ou 'UNKNOWN'
        """
        verb_lower = verb.lower().strip()
        
        if verb_lower in cls.PREPARATION_INITIALE:
            return 'PREPARATION_INITIALE'
        elif verb_lower in cls.TRANSFORMATION_MECANIQUE:
            return 'TRANSFORMATION_MECANIQUE'
        elif verb_lower in cls.MELANGE_COMBINAISON:
            return 'MELANGE_COMBINAISON'
        elif verb_lower in cls.TRANSFERT_MANIPULATION:
            return 'TRANSFERT_MANIPULATION'
        elif verb_lower in cls.CUISSON_ACTIVE:
            return 'CUISSON_ACTIVE'
        elif verb_lower in cls.CUISSON_PASSIVE:
            return 'CUISSON_PASSIVE'
        elif verb_lower in cls.TRANSFORMATION_THERMIQUE:
            return 'TRANSFORMATION_THERMIQUE'
        elif verb_lower in cls.ATTENTE_REPOS:
            return 'ATTENTE_REPOS'
        elif verb_lower in cls.EXTRACTION_SEPARATION:
            return 'EXTRACTION_SEPARATION'
        elif verb_lower in cls.FINITION_SERVICE:
            return 'FINITION_SERVICE'
        elif verb_lower in cls.ACTIONS_SPECIALES:
            return 'ACTIONS_SPECIALES'
        else:
            return 'UNKNOWN'
    
    @classmethod
    def get_all_verbs(cls) -> Set[str]:
        """Retourne l'ensemble de tous les verbes connus"""
        all_verbs = set()
        for category_set in [
            cls.PREPARATION_INITIALE, cls.TRANSFORMATION_MECANIQUE,
            cls.MELANGE_COMBINAISON, cls.TRANSFERT_MANIPULATION,
            cls.CUISSON_ACTIVE, cls.CUISSON_PASSIVE,
            cls.TRANSFORMATION_THERMIQUE, cls.ATTENTE_REPOS,
            cls.EXTRACTION_SEPARATION, cls.FINITION_SERVICE,
            cls.ACTIONS_SPECIALES
        ]:
            all_verbs.update(category_set)
        return all_verbs
    
    @classmethod
    def get_category_stats(cls) -> dict:
        """Retourne les statistiques par catégorie"""
        stats = {}
        for attr_name in dir(cls):
            if attr_name.isupper() and not attr_name.startswith('CATEGORY'):
                category_set = getattr(cls, attr_name)
                if isinstance(category_set, set):
                    stats[attr_name] = {
                        'name': cls.CATEGORY_NAMES.get(attr_name, attr_name),
                        'count': len(category_set)
                    }
        return stats
    
    @classmethod
    def export_taxonomy(cls, output_file: str):
        """Exporte la taxonomie complète en JSON"""
        taxonomy = {}
        for attr_name in dir(cls):
            if attr_name.isupper() and not attr_name.startswith('CATEGORY'):
                category_set = getattr(cls, attr_name)
                if isinstance(category_set, set):
                    taxonomy[attr_name] = {
                        'name': cls.CATEGORY_NAMES.get(attr_name, attr_name),
                        'verbs': sorted(list(category_set)),
                        'count': len(category_set)
                    }
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(taxonomy, f, indent=2, ensure_ascii=False)
        
        print(f"✅ Taxonomie exportée : {output_file}")
        print(f"📊 Total verbes : {len(cls.get_all_verbs())}")





# ==============================================================================
# SECTION 2 : RÈGLES DE SUCCESSION
# ==============================================================================

class SuccessionRules:
    """
    Règles de succession impossibles et suspectes
    """
    
    # Règles Type A : ERREURS CERTAINES
    TYPE_A_RULES = {
        'A1_IRREVERSIBILITE_TEMPORELLE': {
            'description': 'FINITION → PRÉPARATION INITIALE',
            'from_category': 'FINITION_SERVICE',
            'to_category': 'PREPARATION_INITIALE',
            'severity': 'CRITICAL',
            'examples': ['serve → wash', 'plate → peel', 'garnish → clean']
        },
        'A2_IRREVERSIBILITE_PHYSIQUE': {
            'description': 'CUISSON → TRANSFORMATION MÉCANIQUE',
            'from_categories': ['CUISSON_ACTIVE', 'CUISSON_PASSIVE'],
            'to_category': 'TRANSFORMATION_MECANIQUE',
            'severity': 'CRITICAL',
            'examples': ['bake → chop', 'sauté → dice', 'roast → slice']
        },
        'A3_ILLOGISME_FINITION_MELANGE': {
            'description': 'FINITION → MÉLANGE',
            'from_category': 'FINITION_SERVICE',
            'to_category': 'MELANGE_COMBINAISON',
            'severity': 'CRITICAL',
            'examples': ['serve → mix', 'plate → stir', 'garnish → combine']
        },
        'A4_ILLOGISME_FINITION_CUISSON': {
            'description': 'FINITION → CUISSON',
            'from_category': 'FINITION_SERVICE',
            'to_categories': ['CUISSON_ACTIVE', 'CUISSON_PASSIVE'],
            'severity': 'CRITICAL',
            'examples': ['serve → bake', 'garnish → sauté', 'plate → roast']
        }
    }
    
    # Règles Type B : SUSPICIONS
    TYPE_B_RULES = {
        'B1_SAUT_DE_PHASE': {
            'description': 'PRÉPARATION → FINITION (sans étapes intermédiaires)',
            'from_categories': ['PREPARATION_INITIALE', 'TRANSFORMATION_MECANIQUE'],
            'to_category': 'FINITION_SERVICE',
            'severity': 'SUSPICIOUS',
            'context_check': 'check_intermediate_actions',
            'examples': ['chop → serve', 'wash → plate']
        },
        'B2_ORDRE_INVERSE': {
            'description': 'TRANSFERT → TRANSFORMATION MÉCANIQUE',
            'from_category': 'TRANSFERT_MANIPULATION',
            'to_category': 'TRANSFORMATION_MECANIQUE',
            'severity': 'SUSPICIOUS',
            'examples': ['pour → chop', 'add → dice']
        },
        'B3_CUISSON_SANS_PREPARATION': {
            'description': 'CUISSON comme première action',
            'severity': 'SUSPICIOUS',
            'context_check': 'check_first_action',
            'examples': ['bake (first)', 'boil (first)']
        },
        'B4_MELANGE_APRES_CUISSON_PASSIVE': {
            'description': 'CUISSON PASSIVE → MÉLANGE (sans EXTRACTION)',
            'from_category': 'CUISSON_PASSIVE',
            'to_category': 'MELANGE_COMBINAISON',
            'severity': 'SUSPICIOUS',
            'context_check': 'check_extraction_between',
            'examples': ['bake → mix', 'roast → stir']
        },
        'B5_TRANSFORMATION_APRES_TRANSFERT_FINAL': {
            'description': 'TRANSFERT (plat) → TRANSFORMATION MÉCANIQUE',
            'from_category': 'TRANSFERT_MANIPULATION',
            'to_category': 'TRANSFORMATION_MECANIQUE',
            'severity': 'SUSPICIOUS',
            'context_check': 'check_if_final_transfer',
            'examples': ['plate → slice', 'serve → cut']
        }
    }
    
    # Patterns valides mais rares (liste blanche)
    WHITELIST_PATTERNS = [
        ['bake', 'cool', 'slice', 'toast'],  # Pain grillé
        ['cook', 'mash', 'mix', 'cook'],  # Purées re-cuites
        ['boil', 'chop', 'drain', 'sauté'],  # Multitasking
        ['bake', 'cool', 'slice', 'serve'],  # Gâteaux
        ['drain', 'rinse', 'drain'],  # Pâtes/légumineuses
    ]
    
    @classmethod
    def check_type_a_violation(
        cls, 
        cat_from: str, 
        cat_to: str
    ) -> Tuple[bool, str, str]:
        """
        Vérifie si une transition viole une règle Type A
        
        Returns:
            (violation_detected, rule_id, description)
        """
        # A1: FINITION → PRÉPARATION
        if cat_from == 'FINITION_SERVICE' and cat_to == 'PREPARATION_INITIALE':
            return (True, 'A1_IRREVERSIBILITE_TEMPORELLE', 
                    cls.TYPE_A_RULES['A1_IRREVERSIBILITE_TEMPORELLE']['description'])
        
        # A2: CUISSON → TRANSFORMATION
        if cat_from in ['CUISSON_ACTIVE', 'CUISSON_PASSIVE'] and cat_to == 'TRANSFORMATION_MECANIQUE':
            return (True, 'A2_IRREVERSIBILITE_PHYSIQUE',
                    cls.TYPE_A_RULES['A2_IRREVERSIBILITE_PHYSIQUE']['description'])
        
        # A3: FINITION → MÉLANGE
        if cat_from == 'FINITION_SERVICE' and cat_to == 'MELANGE_COMBINAISON':
            return (True, 'A3_ILLOGISME_FINITION_MELANGE',
                    cls.TYPE_A_RULES['A3_ILLOGISME_FINITION_MELANGE']['description'])
        
        # A4: FINITION → CUISSON
        if cat_from == 'FINITION_SERVICE' and cat_to in ['CUISSON_ACTIVE', 'CUISSON_PASSIVE']:
            return (True, 'A4_ILLOGISME_FINITION_CUISSON',
                    cls.TYPE_A_RULES['A4_ILLOGISME_FINITION_CUISSON']['description'])
        
        return (False, '', '')
    
    @classmethod
    def check_type_b_violation(
        cls,
        cat_from: str,
        cat_to: str,
        context: Dict = None
    ) -> Tuple[bool, str, str]:
        """
        Vérifie si une transition viole une règle Type B
        
        Returns:
            (violation_detected, rule_id, description)
        """
        # B1: PRÉPARATION → FINITION (saut de phase)
        if cat_from in ['PREPARATION_INITIALE', 'TRANSFORMATION_MECANIQUE'] and cat_to == 'FINITION_SERVICE':
            if context and context.get('has_intermediate_actions', False):
                return (False, '', '')
            return (True, 'B1_SAUT_DE_PHASE',
                    cls.TYPE_B_RULES['B1_SAUT_DE_PHASE']['description'])
        
        # B2: TRANSFERT → TRANSFORMATION
        if cat_from == 'TRANSFERT_MANIPULATION' and cat_to == 'TRANSFORMATION_MECANIQUE':
            return (True, 'B2_ORDRE_INVERSE',
                    cls.TYPE_B_RULES['B2_ORDRE_INVERSE']['description'])
        
        # B4: CUISSON PASSIVE → MÉLANGE
        if cat_from == 'CUISSON_PASSIVE' and cat_to == 'MELANGE_COMBINAISON':
            if context and context.get('has_extraction_between', False):
                return (False, '', '')
            return (True, 'B4_MELANGE_APRES_CUISSON_PASSIVE',
                    cls.TYPE_B_RULES['B4_MELANGE_APRES_CUISSON_PASSIVE']['description'])
        
        # B5: TRANSFERT → TRANSFORMATION (après service)
        if cat_from == 'TRANSFERT_MANIPULATION' and cat_to == 'TRANSFORMATION_MECANIQUE':
            return (True, 'B5_TRANSFORMATION_APRES_TRANSFERT_FINAL',
                    cls.TYPE_B_RULES['B5_TRANSFORMATION_APRES_TRANSFERT_FINAL']['description'])
        
        return (False, '', '')


# ==============================================================================
# SECTION 3 : UTILITAIRES
# ==============================================================================

def parse_actions_column(actions_value):
    """Parse robustement la colonne 'actions'"""
    if pd.isna(actions_value):
        return []
    
    if isinstance(actions_value, list):
        return actions_value
    
    if isinstance(actions_value, str):
        try:
            parsed = ast.literal_eval(actions_value)
            if isinstance(parsed, list):
                return parsed
        except (ValueError, SyntaxError):
            if ',' in actions_value:
                return [a.strip() for a in actions_value.split(',') if a.strip()]
            return [actions_value.strip()] if actions_value.strip() else []
    
    return []


def create_output_directory(output_dir: str):
    """Crée le répertoire de sortie s'il n'existe pas"""
    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(os.path.join(output_dir, 'errors'), exist_ok=True)
    os.makedirs(os.path.join(output_dir, 'statistics'), exist_ok=True)


# ==============================================================================
# SECTION 4 : ANNOTATION DES SÉQUENCES
# ==============================================================================

def annotate_sequence_with_categories(
    actions: List[str]
) -> List[Tuple[str, str]]:
    """
    Annote chaque action avec sa catégorie
    
    Args:
        actions: Liste d'actions
    
    Returns:
        Liste de tuples (action, category)
    """
    annotated = []
    for action in actions:
        category = VerbTaxonomy.get_category(action)
        annotated.append((action, category))
    
    return annotated


# ==============================================================================
# SECTION 5 : DÉTECTION DES VIOLATIONS
# ==============================================================================

def detect_violations(
    recipe_id: int,
    title: str,
    actions: List[str],
    annotated_sequence: List[Tuple[str, str]]
) -> Dict:
    """
    Détecte toutes les violations dans une séquence
    
    Args:
        recipe_id: ID de la recette
        title: Titre de la recette
        actions: Liste d'actions
        annotated_sequence: Séquence annotée [(action, category), ...]
    
    Returns:
        Dictionnaire avec les violations détectées
    """
    violations = {
        'recipe_id': recipe_id,
        'title': title,
        'sequence_length': len(actions),
        'unknown_verbs': [],
        'type_a_violations': [],
        'type_b_violations': [],
        'repetitions': []
    }
    
    # Détecter les verbes inconnus
    for action, category in annotated_sequence:
        if category == 'UNKNOWN':
            violations['unknown_verbs'].append(action)
    
    # Détecter les répétitions suspectes (>3 fois consécutives)
    i = 0
    while i < len(actions):
        action = actions[i]
        count = 1
        while i + count < len(actions) and actions[i + count] == action:
            count += 1
        
        if count > 3:
            violations['repetitions'].append({
                'action': action,
                'count': count,
                'position': i
            })
        
        i += count
    
    # Détecter les violations Type A et Type B
    for i in range(len(annotated_sequence) - 1):
        action_i, cat_i = annotated_sequence[i]
        action_j, cat_j = annotated_sequence[i + 1]
        
        # Contexte pour analyse
        context_window = annotated_sequence[max(0, i-2):min(len(annotated_sequence), i+4)]
        
        # Type A
        is_violation_a, rule_id_a, desc_a = SuccessionRules.check_type_a_violation(cat_i, cat_j)
        if is_violation_a:
            violations['type_a_violations'].append({
                'position': i,
                'action_from': action_i,
                'category_from': cat_i,
                'action_to': action_j,
                'category_to': cat_j,
                'rule_id': rule_id_a,
                'rule_description': desc_a,
                'context': [f"{a}({c})" for a, c in context_window]
            })
        
        # Type B
        context_info = {
            'has_intermediate_actions': len(annotated_sequence) > 3,
            'has_extraction_between': any(
                cat == 'EXTRACTION_SEPARATION' 
                for _, cat in annotated_sequence[i:i+2]
            )
        }
        
        is_violation_b, rule_id_b, desc_b = SuccessionRules.check_type_b_violation(
            cat_i, cat_j, context_info
        )
        if is_violation_b:
            violations['type_b_violations'].append({
                'position': i,
                'action_from': action_i,
                'category_from': cat_i,
                'action_to': action_j,
                'category_to': cat_j,
                'rule_id': rule_id_b,
                'rule_description': desc_b,
                'context': [f"{a}({c})" for a, c in context_window]
            })
    
    return violations


# ==============================================================================
# SECTION 6 : ANALYSE DU DATASET
# ==============================================================================

def analyze_dataset(
    graphs_df: pd.DataFrame,
    output_dir: str = "strategy_3_results"
) -> Tuple[pd.DataFrame, Dict]:
    """
    Analyse complète du dataset pour détecter les violations
    
    Args:
        graphs_df: DataFrame des graphes
        output_dir: Répertoire de sortie
    
    Returns:
        (violations_df, statistics)
    """
    print(f"\n{'='*80}")
    print(f"ANALYSE DU DATASET - DÉTECTION DES VIOLATIONS")
    print(f"{'='*80}\n")
    
    # Parser les actions
    if 'actions_parsed' not in graphs_df.columns:
        graphs_df['actions_parsed'] = graphs_df['actions'].apply(parse_actions_column)
    
    all_violations = []
    unknown_verbs_set = set()
    
    total_recipes = len(graphs_df)
    
    print(f"Traitement de {total_recipes:,} graphes...\n")
    
    for idx, row in graphs_df.iterrows():
        if idx % 10000 == 0 and idx > 0:
            print(f"  Progression : {idx:,}/{total_recipes:,} ({idx/total_recipes*100:.1f}%)")
        
        recipe_id = row['id']
        title = row.get('title', f"Recipe_{recipe_id}")
        actions = row['actions_parsed']
        
        if not actions or len(actions) == 0:
            continue
        
        # Annoter la séquence
        annotated_seq = annotate_sequence_with_categories(actions)
        
        # Détecter les violations
        violations = detect_violations(recipe_id, title, actions, annotated_seq)
        
        # Collecter les verbes inconnus
        unknown_verbs_set.update(violations['unknown_verbs'])
        
        # Ajouter aux résultats si des violations détectées
        if (len(violations['type_a_violations']) > 0 or 
            len(violations['type_b_violations']) > 0 or 
            len(violations['repetitions']) > 0 or
            len(violations['unknown_verbs']) > 0):
            all_violations.append(violations)
    
    print(f"\n✅ Analyse terminée : {len(all_violations):,} graphes avec anomalies détectées\n")
    
    # Créer le DataFrame des violations
    violations_df = pd.DataFrame(all_violations)
    
    # Calculer les statistiques
    statistics = {
        'total_graphs_analyzed': total_recipes,
        'graphs_with_violations': len(all_violations),
        'violation_rate': len(all_violations) / total_recipes * 100 if total_recipes > 0 else 0,
        'total_type_a_violations': sum(len(v['type_a_violations']) for v in all_violations),
        'total_type_b_violations': sum(len(v['type_b_violations']) for v in all_violations),
        'total_repetitions': sum(len(v['repetitions']) for v in all_violations),
        'total_unknown_verbs': len(unknown_verbs_set),
        'unknown_verbs_list': sorted(list(unknown_verbs_set))
    }
    
    # Statistiques par règle
    type_a_by_rule = defaultdict(int)
    type_b_by_rule = defaultdict(int)
    
    for violations in all_violations:
        for v in violations['type_a_violations']:
            type_a_by_rule[v['rule_id']] += 1
        for v in violations['type_b_violations']:
            type_b_by_rule[v['rule_id']] += 1
    
    statistics['type_a_by_rule'] = dict(type_a_by_rule)
    statistics['type_b_by_rule'] = dict(type_b_by_rule)
    
    # Top 10 successions problématiques
    problematic_successions_a = Counter()
    problematic_successions_b = Counter()
    
    for violations in all_violations:
        for v in violations['type_a_violations']:
            key = f"{v['action_from']} → {v['action_to']}"
            problematic_successions_a[key] += 1
        for v in violations['type_b_violations']:
            key = f"{v['action_from']} → {v['action_to']}"
            problematic_successions_b[key] += 1
    
    statistics['top_10_type_a'] = problematic_successions_a.most_common(10)
    statistics['top_10_type_b'] = problematic_successions_b.most_common(10)
    
    # Afficher les résultats
    print(f"STATISTIQUES GLOBALES:")
    print(f"  - Total graphes analysés : {statistics['total_graphs_analyzed']:,}")
    print(f"  - Graphes avec violations : {statistics['graphs_with_violations']:,} ({statistics['violation_rate']:.2f}%)")
    print(f"  - Violations Type A (critiques) : {statistics['total_type_a_violations']:,}")
    print(f"  - Violations Type B (suspicions) : {statistics['total_type_b_violations']:,}")
    print(f"  - Répétitions suspectes : {statistics['total_repetitions']:,}")
    print(f"  - Verbes inconnus : {statistics['total_unknown_verbs']}")
    
    print(f"\nDistribution des violations Type A par règle:")
    for rule_id, count in statistics['type_a_by_rule'].items():
        print(f"  - {rule_id}: {count:,}")
    
    print(f"\nDistribution des violations Type B par règle:")
    for rule_id, count in statistics['type_b_by_rule'].items():
        print(f"  - {rule_id}: {count:,}")
    
    print(f"\nTop 5 successions problématiques (Type A):")
    for succession, count in statistics['top_10_type_a'][:5]:
        print(f"  - {succession}: {count:,} occurrences")
    
    print(f"\nTop 5 successions problématiques (Type B):")
    for succession, count in statistics['top_10_type_b'][:5]:
        print(f"  - {succession}: {count:,} occurrences")
    
    # Sauvegarder les statistiques
    stats_file = os.path.join(output_dir, 'statistics', 'global_statistics.json')
    with open(stats_file, 'w', encoding='utf-8') as f:
        # Convertir les tuples en listes pour JSON
        stats_copy = statistics.copy()
        stats_copy['top_10_type_a'] = [[k, v] for k, v in statistics['top_10_type_a']]
        stats_copy['top_10_type_b'] = [[k, v] for k, v in statistics['top_10_type_b']]
        json.dump(stats_copy, f, indent=2, ensure_ascii=False)
    
    print(f"\n✅ Statistiques sauvegardées : {stats_file}")
    
    return violations_df, statistics


# ==============================================================================
# SECTION 7 : EXPORT DES RÉSULTATS
# ==============================================================================

def export_violations(
    violations_df: pd.DataFrame,
    output_dir: str = "strategy_3_results"
):
    """
    Exporte les violations dans des fichiers CSV séparés
    
    Args:
        violations_df: DataFrame des violations
        output_dir: Répertoire de sortie
    """
    print(f"\n{'='*80}")
    print(f"EXPORT DES VIOLATIONS")
    print(f"{'='*80}\n")
    
    # Erreurs Type A (critiques)
    type_a_rows = []
    for _, row in violations_df.iterrows():
        for violation in row['type_a_violations']:
            type_a_rows.append({
                'recipe_id': row['recipe_id'],
                'title': row['title'],
                'position': violation['position'],
                'action_from': violation['action_from'],
                'category_from': violation['category_from'],
                'action_to': violation['action_to'],
                'category_to': violation['category_to'],
                'rule_id': violation['rule_id'],
                'rule_description': violation['rule_description'],
                'context': ' | '.join(violation['context'])
            })
    
    if len(type_a_rows) > 0:
        type_a_df = pd.DataFrame(type_a_rows)
        type_a_file = os.path.join(output_dir, 'errors', 'type_a_critical_errors.csv')
        type_a_df.to_csv(type_a_file, index=False, encoding='utf-8')
        print(f"✅ Erreurs Type A exportées : {type_a_file} ({len(type_a_rows):,} violations)")
    
    # Erreurs Type B (suspicions)
    type_b_rows = []
    for _, row in violations_df.iterrows():
        for violation in row['type_b_violations']:
            type_b_rows.append({
                'recipe_id': row['recipe_id'],
                'title': row['title'],
                'position': violation['position'],
                'action_from': violation['action_from'],
                'category_from': violation['category_from'],
                'action_to': violation['action_to'],
                'category_to': violation['category_to'],
                'rule_id': violation['rule_id'],
                'rule_description': violation['rule_description'],
                'context': ' | '.join(violation['context'])
            })
    
    if len(type_b_rows) > 0:
        type_b_df = pd.DataFrame(type_b_rows)
        type_b_file = os.path.join(output_dir, 'errors', 'type_b_suspicious_errors.csv')
        type_b_df.to_csv(type_b_file, index=False, encoding='utf-8')
        print(f"✅ Erreurs Type B exportées : {type_b_file} ({len(type_b_rows):,} violations)")
    
    # Répétitions suspectes
    repetitions_rows = []
    for _, row in violations_df.iterrows():
        for rep in row['repetitions']:
            repetitions_rows.append({
                'recipe_id': row['recipe_id'],
                'title': row['title'],
                'action': rep['action'],
                'count': rep['count'],
                'position': rep['position']
            })
    
    if len(repetitions_rows) > 0:
        rep_df = pd.DataFrame(repetitions_rows)
        rep_file = os.path.join(output_dir, 'errors', 'suspicious_repetitions.csv')
        rep_df.to_csv(rep_file, index=False, encoding='utf-8')
        print(f"✅ Répétitions exportées : {rep_file} ({len(repetitions_rows):,} cas)")
    
    # Résumé par recette
    summary_rows = []
    for _, row in violations_df.iterrows():
        summary_rows.append({
            'recipe_id': row['recipe_id'],
            'title': row['title'],
            'sequence_length': row['sequence_length'],
            'nb_type_a': len(row['type_a_violations']),
            'nb_type_b': len(row['type_b_violations']),
            'nb_repetitions': len(row['repetitions']),
            'nb_unknown_verbs': len(row['unknown_verbs']),
            'quality_score': 1.0 - (2 * len(row['type_a_violations']) + len(row['type_b_violations'])) / max(row['sequence_length'], 1)
        })
    
    summary_df = pd.DataFrame(summary_rows)
    summary_file = os.path.join(output_dir, 'statistics', 'recipe_quality_scores.csv')
    summary_df.to_csv(summary_file, index=False, encoding='utf-8')
    print(f"✅ Scores de qualité exportés : {summary_file}")
    
    print(f"\n{'='*80}\n")


# ==============================================================================
# SECTION 8 : GÉNÉRATION DU RAPPORT
# ==============================================================================

def generate_strategy_3_report(
    statistics: Dict,
    output_dir: str = "strategy_3_results"
) -> str:
    """
    Génère un rapport de synthèse de la Stratégie 3
    
    Args:
        statistics: Dictionnaire des statistiques
        output_dir: Répertoire de sortie
    
    Returns:
        Chemin du fichier de rapport
    """
    print(f"\n{'='*80}")
    print(f"GÉNÉRATION DU RAPPORT DE SYNTHÈSE")
    print(f"{'='*80}\n")
    
    report_lines = []
    
    report_lines.append("=" * 80)
    report_lines.append("RAPPORT DE DÉTECTION DES SUCCESSIONS ILLOGIQUES - STRATÉGIE 3")
    report_lines.append("=" * 80)
    report_lines.append("")
    report_lines.append(f"Date de génération : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report_lines.append("")
    
    # Statistiques globales
    report_lines.append("-" * 80)
    report_lines.append("STATISTIQUES GLOBALES")
    report_lines.append("-" * 80)
    report_lines.append("")
    report_lines.append(f"Total de graphes analysés : {statistics['total_graphs_analyzed']:,}")
    report_lines.append(f"Graphes avec violations : {statistics['graphs_with_violations']:,} ({statistics['violation_rate']:.2f}%)")
    report_lines.append(f"Graphes conformes : {statistics['total_graphs_analyzed'] - statistics['graphs_with_violations']:,} ({100 - statistics['violation_rate']:.2f}%)")
    report_lines.append("")
    report_lines.append(f"Total violations Type A (critiques) : {statistics['total_type_a_violations']:,}")
    report_lines.append(f"Total violations Type B (suspicions) : {statistics['total_type_b_violations']:,}")
    report_lines.append(f"Total répétitions suspectes : {statistics['total_repetitions']:,}")
    report_lines.append(f"Total verbes inconnus : {statistics['total_unknown_verbs']}")
    report_lines.append("")
    
    # Distribution Type A
    report_lines.append("-" * 80)
    report_lines.append("VIOLATIONS TYPE A (ERREURS CERTAINES)")
    report_lines.append("-" * 80)
    report_lines.append("")
    for rule_id, count in statistics['type_a_by_rule'].items():
        pct = (count / statistics['total_type_a_violations'] * 100) if statistics['total_type_a_violations'] > 0 else 0
        report_lines.append(f"  {rule_id}: {count:,} occurrences ({pct:.1f}%)")
    report_lines.append("")
    
    report_lines.append("Top 10 successions Type A les plus fréquentes:")
    for i, (succession, count) in enumerate(statistics['top_10_type_a'], 1):
        report_lines.append(f"  {i:2d}. {succession}: {count:,} occurrences")
    report_lines.append("")
    
    # Distribution Type B
    report_lines.append("-" * 80)
    report_lines.append("VIOLATIONS TYPE B (SUSPICIONS)")
    report_lines.append("-" * 80)
    report_lines.append("")
    for rule_id, count in statistics['type_b_by_rule'].items():
        pct = (count / statistics['total_type_b_violations'] * 100) if statistics['total_type_b_violations'] > 0 else 0
        report_lines.append(f"  {rule_id}: {count:,} occurrences ({pct:.1f}%)")
    report_lines.append("")
    
    report_lines.append("Top 10 successions Type B les plus fréquentes:")
    for i, (succession, count) in enumerate(statistics['top_10_type_b'], 1):
        report_lines.append(f"  {i:2d}. {succession}: {count:,} occurrences")
    report_lines.append("")
    
    # Verbes inconnus - Format sur plusieurs colonnes
    if statistics['total_unknown_verbs'] > 0:
        report_lines.append("-" * 80)
        report_lines.append("VERBES INCONNUS")
        report_lines.append("-" * 80)
        report_lines.append("")
        report_lines.append(f"Total : {statistics['total_unknown_verbs']} verbes")
        report_lines.append("")
        
        # Formater sur plusieurs colonnes (8 par ligne)
        unknown_verbs = statistics['unknown_verbs_list']
        items_per_line = 8
        col_width = max(len(verb) for verb in unknown_verbs) + 2 if unknown_verbs else 15
        
        for i in range(0, len(unknown_verbs), items_per_line):
            line_items = unknown_verbs[i:i+items_per_line]
            formatted_line = ''.join(f"{verb:<{col_width}}" for verb in line_items)
            report_lines.append(formatted_line.rstrip())
        
        report_lines.append("")
    
    # Conclusion
    report_lines.append("=" * 80)
    report_lines.append("CONCLUSION")
    report_lines.append("=" * 80)
    report_lines.append("")
    
    if statistics['violation_rate'] < 5:
        report_lines.append("✅ Le dataset présente une excellente qualité sémantique (<5% de violations).")
    elif statistics['violation_rate'] < 10:
        report_lines.append("⚠️  Le dataset présente une qualité acceptable (5-10% de violations).")
    else:
        report_lines.append("❌ Le dataset présente des problèmes significatifs (>10% de violations).")
    
    report_lines.append("")
    report_lines.append("Recommandations:")
    report_lines.append("1. Examiner et corriger toutes les violations Type A (erreurs certaines)")
    report_lines.append("2. Investiguer les violations Type B les plus fréquentes")
    report_lines.append("3. Enrichir la taxonomie avec les verbes inconnus fréquents")
    report_lines.append("4. Ré-annoter les recettes avec scores de qualité < 0.5")
    report_lines.append("")
    report_lines.append("=" * 80)
    report_lines.append("FIN DU RAPPORT")
    report_lines.append("=" * 80)
    
    # Sauvegarder le rapport
    report_text = '\n'.join(report_lines)
    report_file = os.path.join(output_dir, 'strategy_3_report.txt')
    
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write(report_text)
    
    print(report_text)
    print(f"\n✅ Rapport sauvegardé : {report_file}")
    
    return report_file


# ==============================================================================
# SECTION 9 : PIPELINE PRINCIPAL
# ==============================================================================

def run_strategy_3_pipeline(
    graphs_recipes_csv: str,
    output_dir: str = "strategy_3_results"
):
    """
    Exécute le pipeline complet de la Stratégie 3
    
    Args:
        graphs_recipes_csv: Chemin vers graphs_recipes.csv
        output_dir: Répertoire de sortie
    """
    print(f"\n{'#'*80}")
    print(f"# PIPELINE STRATÉGIE 3 - DÉTECTION DES SUCCESSIONS ILLOGIQUES")
    print(f"{'#'*80}\n")
    
    # Créer le répertoire de sortie
    create_output_directory(output_dir)
    
    # Étape 1 : Export de la taxonomie
    print("ÉTAPE 1/5 : Export de la taxonomie des verbes...")
    taxonomy_file = os.path.join(output_dir, 'verb_taxonomy.json')
    VerbTaxonomy.export_taxonomy(taxonomy_file)
    
    # Étape 2 : Chargement des données
    print("\nÉTAPE 2/5 : Chargement des données...")
    graphs_df = pd.read_csv(graphs_recipes_csv)
    print(f"  ✅ Graphes chargés : {len(graphs_df):,}")
    
    # Étape 3 : Analyse et détection
    print("\nÉTAPE 3/5 : Analyse et détection des violations...")
    violations_df, statistics = analyze_dataset(graphs_df, output_dir)
    
    # Étape 4 : Export des violations
    print("\nÉTAPE 4/5 : Export des violations...")
    export_violations(violations_df, output_dir)
    
    # Étape 5 : Génération du rapport
    print("\nÉTAPE 5/5 : Génération du rapport final...")
    generate_strategy_3_report(statistics, output_dir)
    
    print(f"\n{'#'*80}")
    print(f"# PIPELINE TERMINÉ AVEC SUCCÈS")
    print(f"{'#'*80}\n")
    print(f"📁 Tous les résultats sont dans : {output_dir}/")
    print(f"   - Taxonomie : {output_dir}/verb_taxonomy.json")
    print(f"   - Erreurs : {output_dir}/errors/")
    print(f"   - Statistiques : {output_dir}/statistics/")
    print(f"   - Rapport : {output_dir}/strategy_3_report.txt")
    print(f"\n✅ Stratégie 3 complétée !\n")